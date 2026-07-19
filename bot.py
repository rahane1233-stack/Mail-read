"""
╔══════════════════════════════════════════════════════════════════╗
║          NEROXA SHOP BOT — Professional Reply Keyboard UI        ║
║   aiogram 3.x  |  Firebase  |  100% Reply Keyboard Navigation   ║
╚══════════════════════════════════════════════════════════════════╝
"""

import asyncio
import html as html_lib
import io
import json
import logging
import os
import random
import re
import sys
import time
import uuid
from collections import defaultdict, deque
from datetime import datetime
from pathlib import Path
from typing import Any, Dict, List, Optional, Tuple

import requests as _requests
import urllib.request as _urllib_request

import firebase_admin
import openpyxl
from openpyxl.styles import Alignment, Font, PatternFill
import pyotp
import pytz
from aiohttp import web
from aiogram import BaseMiddleware, Bot, Dispatcher, F, Router
from aiogram.client.default import DefaultBotProperties
from aiogram.enums import ParseMode
from aiogram.filters import Command
from aiogram.fsm.context import FSMContext
from aiogram.fsm.state import State, StatesGroup
from aiogram.fsm.storage.memory import MemoryStorage
from aiogram.types import (
    BufferedInputFile,
    CallbackQuery,
    Document,
    InlineKeyboardButton,
    InlineKeyboardMarkup,
    KeyboardButton,
    Message,
    ReplyKeyboardMarkup,
    TelegramObject,
)
from aiogram.webhook.aiohttp_server import SimpleRequestHandler, setup_application
from dotenv import load_dotenv
from firebase_admin import credentials, db as fdb, storage

load_dotenv()

# ╔══════════════════════════════════════════════════════════════════╗
# ║              ⚙️  CONFIGURATION — আপনার মান এখানে বসান           ║
# ╚══════════════════════════════════════════════════════════════════╝

# ── ১. Telegram ─────────────────────────────────────────────────────
BOT_TOKEN_VALUE          = "8727550482:AAGoinHbp-0Svxa3djY8UDqQkwFFrIwltQk"
ADMIN_IDS_VALUE          = "8502686983"

# ── ২. Firebase ─────────────────────────────────────────────────────
FIREBASE_DB_URL_VALUE    = "https://mail-shop-bot-5f4f3-default-rtdb.asia-southeast1.firebasedatabase.app/"
FIREBASE_BUCKET_VALUE    = "mail-shop-bot-5f4f3.appspot.com"

# Firebase credentials dict (private key \n properly handled)
_FIREBASE_CREDS = {
    "type": "service_account",
    "project_id": "mail-shop-bot-5f4f3",
    "private_key_id": "4f863ad97d1b1fcd30b9ac1a7921c0d005e1128b",
    "private_key": "-----BEGIN PRIVATE KEY-----\nMIIEvgIBADANBgkqhkiG9w0BAQEFAASCBKgwggSkAgEAAoIBAQCnDt3kAAJ7PIgF\n9b2QrAYEkf/gd0XCA34ziu8GHmuvxepMTbxmLCfZIOOflosyG/KDWZqpq546QPyQ\nmG3KmdGoCr7zPQAIv+fXCWd31kXV7IDDmRVDVoPdn7s3x44hca0O4XrB2m7wsWC0\nTCxE8/FXnHHzWS2p0Sqd896Da7LNsVXFsr6ic4JX8FGlE1ebYQiMSroUejOjk2tt\n8qV/eXgPT3WbLXWQM6cPz7Z+x4AbAwGvxeybtwoR6kwRwUH+SH47fZJ3ryKY9wVK\n9pJRMEXa9uC1r4rMUfyZZ0+fHzqSTKixVZ1cYY8AG1GOomJR7NgZ+oQpy1UA2L3B\nzTPo5TBdAgMBAAECggEAP6cXfsdOKryuq26t0xDonhcvIszvZHRGQsdeObuflnLX\nykkYTunmKQIyGN2YnfguGEQs/RcqC9I1Kbcapkajrt6hUTbd63eLk9C+ftfC9jbN\n/Tk389dkGS7CfAdqLW4N3YymZHShLs63JRudBozYWWR/upQxJPJQxaxlDTgdAehB\nRqJZtKAa0SP40MpcVcqZaWpom4wmgToUSQD2ZqnpVFesNnfrY6+0glcIfakFxTFA\nxI1+1nL0i+I0TmMy1GSQH1ABZrvN3tGrceiHfqC/v7VaLBeBbjphsiNOtM+DYw2C\nFiFOEm0VFyFk+MCLvVhm5xFdZ5bNkPvQcT7h0OgVFQKBgQDRILZKgufinJIUlAgE\n9KVyvQ/3cASjLZMehDCtZXgr3gcsbuZYuZUii5nRYfvZ3TYpgBD5JiXQmEhgaL5E\nRv+Iuw+tpzi5gyzqfFK6xr6N3YZsFUrAExAItehBbRbib1h4i6F2a0oZu0dggaMA\npqSs+ZPTEB6EdmSJ/4k6P49e4wKBgQDMgEnHt6eYyfLPNbFkWDGVF3n5JvVC5slY\nVTcqJRgrb2NoGfoPljICdeIRJcK5z2B2PeSL6mukIJEN4pyk6xLS3z9qJzECXXHV\nTb5EZ253CiFNS7TFlHcryXI7X9NXCKYJVlrthg9vkQ5V2IspOSgJV334EHobbt9X\nIJRhVK0XvwKBgQChhk2mMYPfQSWHZWroQPnFLIg3irraOmpGiL18QEJYR45s4F9k\nKjspSoAM/ExlUvnxhewWNEPC4MwOQviqjdfzCOCsNNuYVdwMZOgVQUqPEoov0yZA\nQrkAsVfpqNOjI3NG8DpO18GNLFyOCrMW3p+UxtRJkkqv7y3qdIfOiKc5FwKBgHI1\nn37roa0h8/onWXfaDW7fmp48VLNVYtNzXAisiNAROGo2P8KetjVLksLS3Oqa15uR\nu2cst4sFKR2hFqzAIFYmmir10lgoyd8/uOhI/5d5z9l0U3QZE2kf6y0fuk8cJGaI\nlOWcflhnoaLt+eI+6o41D8QPp7JUfGUTa+rjuHvPAoGBANCqwghW7dAit/itf9lC\ngfdJh9m7crtXdR2cwsuMgbZZ1dqG7NyQVe6o9Y1s5g1rt8rkmDttBI2ppx6ALofZ\nJEY9KOYdr3/2UhbEoEgx5TIvEsbCUceWBDLFGGQwFyTHKRIvvjmdyqY7fj/jNS/c\ny3Mjp+l2j95TP8qmA1kLMLjK\n-----END PRIVATE KEY-----\n",
    "client_email": "firebase-adminsdk-fbsvc@mail-shop-bot-5f4f3.iam.gserviceaccount.com",
    "client_id": "101800195206257118014",
    "auth_uri": "https://accounts.google.com/o/oauth2/auth",
    "token_uri": "https://oauth2.googleapis.com/token",
    "auth_provider_x509_cert_url": "https://www.googleapis.com/oauth2/v1/certs",
    "client_x509_cert_url": "https://www.googleapis.com/robot/v1/metadata/x509/firebase-adminsdk-fbsvc%40mail-shop-bot-5f4f3.iam.gserviceaccount.com",
    "universe_domain": "googleapis.com",
}
FIREBASE_CREDS_JSON_VALUE = json.dumps(_FIREBASE_CREDS)

# ── ৩. Optional ─────────────────────────────────────────────────────
FORCE_JOIN_CHANNEL_VALUE = "https://t.me/NeroxaOfficial"
SUPPORT_USERNAME_VALUE   = "@NeroxaOfficial"

# ── ৪. Dongvan Mail API ─────────────────────────────────────────────
DONGVAN_API_KEY_VALUE    = "ChXbDXJGgQMYFwqYfRANwLc7i"

# ════════════════════════════════════════════════════════════════════
os.environ.setdefault("BOT_TOKEN",                BOT_TOKEN_VALUE)
os.environ.setdefault("ADMIN_IDS",                ADMIN_IDS_VALUE)
os.environ.setdefault("FIREBASE_DATABASE_URL",    FIREBASE_DB_URL_VALUE)
os.environ.setdefault("FIREBASE_STORAGE_BUCKET",  FIREBASE_BUCKET_VALUE)
os.environ.setdefault("FIREBASE_CREDENTIALS_JSON",FIREBASE_CREDS_JSON_VALUE)
os.environ.setdefault("FORCE_JOIN_CHANNEL",       FORCE_JOIN_CHANNEL_VALUE)
os.environ.setdefault("SUPPORT_USERNAME",         SUPPORT_USERNAME_VALUE)
os.environ.setdefault("DONGVAN_API_KEY",          DONGVAN_API_KEY_VALUE)
# ════════════════════════════════════════════════════════════════════

# ══════════════════════════════════════════════════════════════════
# UNICODE BOLD HELPER
# ══════════════════════════════════════════════════════════════════

def _b(text: str) -> str:
    """Convert ASCII letters/digits → Unicode Mathematical Bold Sans-Serif.
    Emojis, spaces and punctuation are left as-is, so _b("🏠 Home") → "🏠 𝗛𝗼𝗺𝗲".
    """
    out = []
    for ch in text:
        if "A" <= ch <= "Z":
            out.append(chr(0x1D5D4 + ord(ch) - ord("A")))
        elif "a" <= ch <= "z":
            out.append(chr(0x1D5EE + ord(ch) - ord("a")))
        elif "0" <= ch <= "9":
            out.append(chr(0x1D7EC + ord(ch) - ord("0")))
        else:
            out.append(ch)
    return "".join(out)


logging.basicConfig(
    level=logging.INFO,
    format="%(asctime)s | %(levelname)-8s | %(name)s | %(message)s",
    handlers=[logging.StreamHandler(sys.stdout)],
)
logger = logging.getLogger(__name__)


# ══════════════════════════════════════════════════════════════════
# CONFIGURATION
# ══════════════════════════════════════════════════════════════════

def _csv_list(key: str) -> List[int]:
    raw = os.getenv(key, "").strip()
    return [int(x.strip()) for x in raw.split(",") if x.strip()]


BOT_TOKEN            = os.environ["BOT_TOKEN"]
ADMIN_IDS: List[int] = _csv_list("ADMIN_IDS")
FIREBASE_CREDS_JSON  = os.environ["FIREBASE_CREDENTIALS_JSON"]
FIREBASE_DB_URL      = os.environ["FIREBASE_DATABASE_URL"]
FIREBASE_BUCKET      = os.environ["FIREBASE_STORAGE_BUCKET"]

DONGVAN_API_KEY  = os.getenv("DONGVAN_API_KEY", DONGVAN_API_KEY_VALUE)

# Dongvan API endpoints
_DV_OAUTH2_URL     = "https://api.dongvanfb.net/api/getOauth2"
_DV_GET_CODE_URL   = "https://tools.dongvanfb.net/api/get_code_oauth2"
_DV_GET_MSGS_URL   = "https://tools.dongvanfb.net/api/get_messages_oauth2"
_DV_GRAPH_MSGS_URL = "https://tools.dongvanfb.net/api/graph_messages"

# In-memory mail sessions: uid -> {email, password, refresh_token, client_id}
_mail_sessions: Dict[int, Dict] = {}

WEBHOOK_URL: Optional[str] = os.getenv("WEBHOOK_URL") or None
WEBHOOK_PATH   = os.getenv("WEBHOOK_PATH", "/webhook")
WEBAPP_HOST    = os.getenv("WEBAPP_HOST", "0.0.0.0")
WEBAPP_PORT    = int(os.getenv("PORT") or os.getenv("WEBAPP_PORT", "8080"))

SPAM_COOLDOWN  = float(os.getenv("SPAM_COOLDOWN_SECONDS", "2"))
SPAM_MAX_MSGS  = int(os.getenv("SPAM_MAX_MESSAGES", "5"))
SPAM_WINDOW    = float(os.getenv("SPAM_WINDOW_SECONDS", "10"))
SUPPORT_USERNAME = os.getenv("SUPPORT_USERNAME", "@support")
WELCOME_MSG    = os.getenv("WELCOME_MESSAGE", "Welcome to our premium shop!")
DEFAULT_FORCE_JOIN = os.getenv("FORCE_JOIN_CHANNEL") or None

STOCK_EXPORT_GROUP_ID = -1003937882916
_stock_export_task: Optional[asyncio.Task] = None
_backup_task: Optional[asyncio.Task] = None

MAIL_SHOP_FILE = Path(__file__).parent / "My_Mail_Shop_Orders.xlsx"

_MAIL_SHEET_HEADERS = [
    "Order ID", "Date (UTC)", "User ID", "Username",
    "Product", "Qty", "Total ($)", "Items Delivered",
]
_HEADER_FONT  = Font(bold=True, color="FFFFFF")
_HEADER_FILL  = PatternFill("solid", fgColor="1F4E79")
_HEADER_ALIGN = Alignment(horizontal="center", vertical="center")


def _init_mail_shop_file() -> None:
    if MAIL_SHOP_FILE.exists():
        return
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Mail Orders"
    ws.append(_MAIL_SHEET_HEADERS)
    for cell in ws[1]:
        cell.font  = _HEADER_FONT
        cell.fill  = _HEADER_FILL
        cell.alignment = _HEADER_ALIGN
    ws.freeze_panes = "A2"
    wb.save(str(MAIL_SHOP_FILE))


def make_stock_template_xlsx(product_name: str) -> bytes:
    """Generate a blank xlsx template for admins to fill in stock accounts."""
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Stock"

    hdr_font  = Font(bold=True, color="FFFFFF")
    hdr_fill  = PatternFill("solid", fgColor="1F4E79")
    hdr_align = Alignment(horizontal="center", vertical="center")
    eg_fill   = PatternFill("solid", fgColor="E2EFDA")

    ws.append(["Email", "Password"])
    for cell in ws[1]:
        cell.font      = hdr_font
        cell.fill      = hdr_fill
        cell.alignment = hdr_align

    examples = [
        ("user1@gmail.com", "pass1234"),
        ("user2@gmail.com", "secret99"),
        ("user3@yahoo.com",  "hello456"),
    ]
    for row in examples:
        ws.append(list(row))
        for cell in ws[ws.max_row]:
            cell.fill = eg_fill

    ws.column_dimensions["A"].width = 30
    ws.column_dimensions["B"].width = 20
    ws.freeze_panes = "A2"

    note_row = ws.max_row + 2
    ws.cell(row=note_row, column=1, value="⚠ Replace example rows with real accounts. Keep Email in column A, Password in column B.")
    ws.cell(row=note_row, column=1).font = Font(italic=True, color="FF0000")

    buf = io.BytesIO()
    wb.save(buf)
    buf.seek(0)
    return buf.read()


def make_stock_export_xlsx(product_name: str, stock_items: dict) -> bytes:
    """Export current stock items for a product to xlsx (admin download)."""
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Stock"

    hdr_font  = Font(bold=True, color="FFFFFF")
    hdr_fill  = PatternFill("solid", fgColor="1F4E79")
    hdr_align = Alignment(horizontal="center", vertical="center")
    alt_fill  = PatternFill("solid", fgColor="DCE6F1")

    ws.append(["#", "Email", "Password", "Raw"])
    for cell in ws[1]:
        cell.font      = hdr_font
        cell.fill      = hdr_fill
        cell.alignment = hdr_align

    # Normalise: Firebase may return list-like data in edge cases
    if isinstance(stock_items, list):
        stock_items = {str(i): v for i, v in enumerate(stock_items)}

    for i, (key, item) in enumerate(stock_items.items(), 1):
        item_str = str(item).strip()
        if ":" in item_str:
            parts = item_str.split(":", 1)
            email, password = parts[0].strip(), parts[1].strip()
        else:
            email, password = item_str, ""
        ws.append([i, email, password, item_str])
        if i % 2 == 0:
            for cell in ws[ws.max_row]:
                cell.fill = alt_fill

    ws.column_dimensions["A"].width = 6
    ws.column_dimensions["B"].width = 35
    ws.column_dimensions["C"].width = 25
    ws.column_dimensions["D"].width = 65
    ws.freeze_panes = "A2"

    info_sheet = wb.create_sheet("Info")
    info_sheet.append(["Product", product_name])
    info_sheet.append(["Total Items", len(stock_items)])
    info_sheet.append(["Exported At", time.strftime("%Y-%m-%d %H:%M UTC", time.gmtime())])
    info_sheet.append([])
    info_sheet.append(["Note", "Column D (Raw) contains the original account string. Use this column when re-importing."])
    info_sheet.column_dimensions["A"].width = 18
    info_sheet.column_dimensions["B"].width = 65

    buf = io.BytesIO()
    wb.save(buf)
    buf.seek(0)
    return buf.read()


def make_all_stock_export_xlsx(products: dict, all_stocks: dict) -> bytes:
    """Export stock for ALL products into a single multi-sheet xlsx file."""
    wb = openpyxl.Workbook()
    wb.remove(wb.active)

    hdr_font  = Font(bold=True, color="FFFFFF")
    hdr_fill  = PatternFill("solid", fgColor="1F4E79")
    hdr_align = Alignment(horizontal="center", vertical="center")
    alt_fill  = PatternFill("solid", fgColor="DCE6F1")

    summary_ws = wb.create_sheet("Summary")
    summary_ws.append(["Product", "Category", "Stock Count", "Exported At"])
    for cell in summary_ws[1]:
        cell.font      = hdr_font
        cell.fill      = hdr_fill
        cell.alignment = hdr_align

    for pid, product in products.items():
        pname    = product.get("name", pid)
        category = product.get("category", "mail")
        stock    = all_stocks.get(pid) or {}
        summary_ws.append([pname, category.upper(), len(stock), time.strftime("%Y-%m-%d %H:%M UTC", time.gmtime())])

        safe_title = re.sub(r"[\\/*?:\[\]]", "", pname)[:28] or pid[:28]
        ws = wb.create_sheet(safe_title)
        ws.append(["#", "Email", "Password", "Raw"])
        for cell in ws[1]:
            cell.font      = hdr_font
            cell.fill      = hdr_fill
            cell.alignment = hdr_align
        # Normalise: Firebase may return list-like data in edge cases
        if isinstance(stock, list):
            stock = {str(i): v for i, v in enumerate(stock)}
        for i, (key, item) in enumerate(stock.items(), 1):
            item_str = str(item).strip()
            if ":" in item_str:
                parts = item_str.split(":", 1)
                email, password = parts[0].strip(), parts[1].strip()
            else:
                email, password = item_str, ""
            ws.append([i, email, password, item_str])
            if i % 2 == 0:
                for cell in ws[ws.max_row]:
                    cell.fill = alt_fill
        ws.column_dimensions["A"].width = 6
        ws.column_dimensions["B"].width = 35
        ws.column_dimensions["C"].width = 25
        ws.column_dimensions["D"].width = 65
        ws.freeze_panes = "A2"

    for col in ["A", "B", "C", "D"]:
        summary_ws.column_dimensions[col].width = 25

    buf = io.BytesIO()
    wb.save(buf)
    buf.seek(0)
    return buf.read()


def make_user_order_xlsx(
    order_id: str, product_name: str, qty: int, total: float, items: List[str]
) -> bytes:
    """Build an in-memory xlsx of the user's purchased mail accounts."""
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Your Order"

    info_font  = Font(bold=True)
    hdr_font   = Font(bold=True, color="FFFFFF")
    hdr_fill   = PatternFill("solid", fgColor="1F4E79")
    hdr_align  = Alignment(horizontal="center", vertical="center")
    alt_fill   = PatternFill("solid", fgColor="DCE6F1")

    for label, value in [
        ("Order ID",    order_id),
        ("Product",     product_name),
        ("Quantity",    qty),
        ("Total Paid",  f"${total:.2f}"),
    ]:
        ws.append([label, value])
        ws.cell(row=ws.max_row, column=1).font = info_font

    ws.append([])
    ws.append(["#", "Account"])
    hdr_row = ws.max_row
    for cell in ws[hdr_row]:
        cell.font      = hdr_font
        cell.fill      = hdr_fill
        cell.alignment = hdr_align

    for i, item in enumerate(items, 1):
        ws.append([i, item])
        if i % 2 == 0:
            for cell in ws[ws.max_row]:
                cell.fill = alt_fill

    ws.column_dimensions["A"].width = 12
    ws.column_dimensions["B"].width = 55
    ws.freeze_panes = f"A{hdr_row + 1}"

    buf = io.BytesIO()
    wb.save(buf)
    buf.seek(0)
    return buf.read()


def make_products_xlsx(products: dict, all_stocks: dict = None) -> bytes:
    """Export all products to an xlsx file for admin download.
    If all_stocks is provided, also creates per-product stock sheets."""
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Products"

    hdr_font  = Font(bold=True, color="FFFFFF")
    hdr_fill  = PatternFill("solid", fgColor="1F4E79")
    hdr_align = Alignment(horizontal="center", vertical="center")

    headers = ["Name", "Emoji", "Category", "Price (USD)", "Delivery Mode", "Hidden", "Stock Count"]
    ws.append(headers)
    for cell in ws[1]:
        cell.font      = hdr_font
        cell.fill      = hdr_fill
        cell.alignment = hdr_align

    alt_fill = PatternFill("solid", fgColor="DCE6F1")
    for i, (pid, p) in enumerate(products.items(), 1):
        row = [
            p.get("name", ""),
            p.get("emoji", "📦"),
            p.get("category", "mail"),
            p.get("price", 0.0),
            p.get("delivery_mode", "manual"),
            "yes" if p.get("hidden") else "no",
            p.get("stock_count", 0),
        ]
        ws.append(row)
        if i % 2 == 0:
            for cell in ws[ws.max_row]:
                cell.fill = alt_fill

    col_widths = [30, 10, 12, 14, 16, 10, 14]
    for col_idx, width in enumerate(col_widths, 1):
        ws.column_dimensions[ws.cell(1, col_idx).column_letter].width = width
    ws.freeze_panes = "A2"

    # Per-product stock sheets
    if all_stocks:
        used_sheet_names = set()
        for pid, product in products.items():
            stock = all_stocks.get(pid) or {}
            if not stock:
                continue
            pname = product.get("name", pid)
            safe_title = re.sub(r"[\\/*?:\[\]]", "", pname)[:28] or pid[:28]
            # Handle sheet-name collisions
            original_title = safe_title
            counter = 2
            while safe_title.lower() in used_sheet_names:
                suffix = f"({counter})"
                safe_title = original_title[:28 - len(suffix)] + suffix
                counter += 1
            used_sheet_names.add(safe_title.lower())
            stock_ws = wb.create_sheet(safe_title)
            stock_ws.append(["#", "Email", "Password", "Raw"])
            for cell in stock_ws[1]:
                cell.font      = hdr_font
                cell.fill      = hdr_fill
                cell.alignment = hdr_align
            if isinstance(stock, list):
                stock = {str(i): v for i, v in enumerate(stock)}
            for idx, (key, item) in enumerate(stock.items(), 1):
                item_str = str(item).strip()
                if ":" in item_str:
                    parts = item_str.split(":", 1)
                    email, password = parts[0].strip(), parts[1].strip()
                else:
                    email, password = item_str, ""
                stock_ws.append([idx, email, password, item_str])
                if idx % 2 == 0:
                    for cell in stock_ws[stock_ws.max_row]:
                        cell.fill = alt_fill
            stock_ws.column_dimensions["A"].width = 6
            stock_ws.column_dimensions["B"].width = 35
            stock_ws.column_dimensions["C"].width = 25
            stock_ws.column_dimensions["D"].width = 65
            stock_ws.freeze_panes = "A2"

    # Instructions sheet
    ws2 = wb.create_sheet("Import Guide")
    ws2.append(["📋 Import Guide — How to add products via file"])
    ws2.append([])
    ws2.append(["Column", "Description", "Allowed Values"])
    for cell in ws2[3]:
        cell.font = Font(bold=True)
    guide = [
        ("Name",          "Product name",                                   "Any text"),
        ("Emoji",         "Product emoji icon",                             "Any emoji, e.g. 📮 🌐 🔐"),
        ("Category",      "Product type",                                   "mail / vpn / proxy"),
        ("Price (USD)",   "Price per unit / per day",                       "Number, e.g. 1.50"),
        ("Delivery Mode", "How product is delivered (proxy only)",           "manual / auto"),
        ("Hidden",        "Whether product is visible to users",             "yes / no"),
    ]
    for row in guide:
        ws2.append(list(row))
    for col in ["A", "B", "C"]:
        ws2.column_dimensions[col].width = 25

    buf = io.BytesIO()
    wb.save(buf)
    buf.seek(0)
    return buf.read()


def parse_products_xlsx(content: bytes) -> List[dict]:
    """Parse an xlsx/csv file and return a list of product dicts to import.
    Also parses per-product stock sheets and attaches stock_items to each product."""
    results = []
    try:
        wb = openpyxl.load_workbook(io.BytesIO(content), read_only=True, data_only=True)
        ws = wb.active
        rows = list(ws.iter_rows(values_only=True))
    except Exception:
        return results

    if not rows:
        return results

    # Detect header row
    header = [str(c).strip().lower() if c else "" for c in rows[0]]
    col = {}
    for i, h in enumerate(header):
        if "name" in h:
            col["name"] = i
        elif "emoji" in h:
            col["emoji"] = i
        elif "cat" in h:
            col["category"] = i
        elif "price" in h:
            col["price"] = i
        elif "delivery" in h or "mode" in h:
            col["delivery_mode"] = i
        elif "hidden" in h:
            col["hidden"] = i

    if "name" not in col or "price" not in col:
        return results

    for row in rows[1:]:
        if not row or not row[col["name"]]:
            continue
        name  = str(row[col["name"]]).strip()
        if not name:
            continue
        try:
            price = float(str(row[col["price"]]).replace("$", "").strip())
        except Exception:
            continue
        emoji    = str(row[col["emoji"]]).strip() if "emoji" in col and row[col["emoji"]] else "📦"
        category = str(row[col["category"]]).strip().lower() if "category" in col and row[col["category"]] else "mail"
        if category not in ("mail", "vpn", "proxy"):
            category = "mail"
        delivery = str(row[col["delivery_mode"]]).strip().lower() if "delivery_mode" in col and row[col["delivery_mode"]] else "manual"
        if delivery not in ("manual", "auto"):
            delivery = "manual"
        hidden_v = str(row[col["hidden"]]).strip().lower() if "hidden" in col and row[col["hidden"]] else "no"
        hidden   = hidden_v in ("yes", "true", "1")
        results.append({
            "name": name, "emoji": emoji, "category": category,
            "price": round(price, 4), "delivery_mode": delivery, "hidden": hidden,
            "stock_items": [],
        })

    # Parse per-product stock sheets (match by product name)
    product_names = {p["name"].lower(): p for p in results}
    skip_sheets = {"products", "import guide", "summary"}
    for sheet_name in wb.sheetnames:
        if sheet_name.lower() in skip_sheets:
            continue
        try:
            stock_ws = wb[sheet_name]
            stock_rows = list(stock_ws.iter_rows(values_only=True))
        except Exception:
            continue
        if not stock_rows:
            continue
        # Try to match sheet name to a product
        matched_product = product_names.get(sheet_name.strip().lower())
        if not matched_product:
            # Try partial match
            for pname_lower, prod in product_names.items():
                if pname_lower[:28].rstrip() == sheet_name.strip().lower()[:28].rstrip():
                    matched_product = prod
                    break
        if not matched_product:
            continue
        # Parse stock rows - look for Raw or Email column
        stock_header = [str(c).strip().lower() if c else "" for c in stock_rows[0]]
        raw_col = None
        email_col = None
        pass_col = None
        for si, sh in enumerate(stock_header):
            if "raw" in sh:
                raw_col = si
            elif "email" in sh:
                email_col = si
            elif "password" in sh or "pass" in sh:
                pass_col = si
        for srow in stock_rows[1:]:
            if not srow:
                continue
            if raw_col is not None and srow[raw_col]:
                item_str = str(srow[raw_col]).strip()
                if item_str:
                    matched_product["stock_items"].append(item_str)
            elif email_col is not None and srow[email_col]:
                email = str(srow[email_col]).strip()
                if pass_col is not None and srow[pass_col]:
                    password = str(srow[pass_col]).strip()
                    matched_product["stock_items"].append(f"{email}:{password}")
                else:
                    matched_product["stock_items"].append(email)

    return results


def append_to_mail_shop_file(
    order_id: str, uid: int, username: str,
    product_name: str, qty: int, total: float,
    items: List[str], ts: int,
) -> None:
    _init_mail_shop_file()
    wb = openpyxl.load_workbook(str(MAIL_SHOP_FILE))
    ws = wb.active
    date_s  = time.strftime("%Y-%m-%d %H:%M", time.gmtime(ts))
    items_s = "\n".join(items) if items else ""
    ws.append([order_id, date_s, uid, username, product_name, qty, round(total, 2), items_s])
    # Refresh column widths
    for col in ws.columns:
        max_len = max((len(str(c.value or "")) for c in col), default=10)
        ws.column_dimensions[col[0].column_letter].width = min(max_len + 4, 60)
    wb.save(str(MAIL_SHOP_FILE))

HOME_BTN   = _b("🏠 Home")
BACK_BTN   = _b("🔙 Back")
CANCEL_BTN = _b("❌ Cancel")

# ── User menu ──────────────────────────────────────────────────────
BTN_BALANCE  = _b("💎 Balance")
BTN_GET_MAIL   = _b("📨 Get Mail")
BTN_TEMP_MAIL  = _b("🕐 Temp Mail")
BTN_BUY_VPN  = _b("🛡 Buy VPN")
BTN_BUY_PROXY= _b("🌍 Buy Proxy")
BTN_DEPOSIT  = _b("💳 Deposit")
BTN_GET_CODE = _b("📲 Get Code")
BTN_GET_2FA  = _b("🔐 Get 2FA")

# ── Get Code Sub-menu ───────────────────────────────────────────────
BTN_GC_SET_MAIL = _b("📧 Set Mail")
BTN_GC_CODES    = _b("📬 Get Codes")
BTN_GC_INBOX    = _b("📥 Read Inbox")
BTN_GC_REFRESH  = _b("🔄 Refresh Inbox")
BTN_GC_FILTER   = _b("🎯 Filter Mail")
BTN_GC_CHANGE   = _b("✏️ Change Mail")
BTN_HISTORY  = _b("📋 Order History")
BTN_SUPPORT  = _b("🆘 Support")
BTN_AI_SUPPORT = _b("🤖 AI Support")
BTN_REFERRAL = _b("🔗 Referral")

# ── AI Support FAQ Topics ──────────────────────────────────────────
BTN_AI_HOW_BUY     = _b("❓ How to Buy")
BTN_AI_HOW_DEPOSIT = _b("💳 How to Deposit")
BTN_AI_STOCK       = _b("📦 Check Stock")
BTN_AI_REFERRAL    = _b("🔗 Referral Info")
BTN_AI_ORDERS      = _b("📋 My Orders")
BTN_AI_2FA         = _b("🔐 2FA Help")
BTN_AI_MAIL        = _b("📨 Mail Codes")
BTN_AI_ASK         = _b("💬 Ask a Question")
BTN_CONFIRM  = _b("✅ Confirm Purchase")
BTN_COUPON   = _b("🎫 Apply Coupon")

# ── User Panel Enhancement buttons ────────────────────────────────
BTN_TRENDING    = _b("🔥 Trending")
BTN_FAVORITES   = _b("⭐ Favorites")
BTN_MY_STATS    = _b("📊 My Stats")
BTN_SEARCH      = _b("🔍 Search")
BTN_MORE        = _b("📂 More")

# ── Duration ───────────────────────────────────────────────────────
BTN_1DAY   = _b("⏱ 1 Day")
BTN_7DAYS  = _b("⏱ 7 Days")
BTN_30DAYS = _b("⏱ 30 Days")
BTN_90DAYS = _b("⏱ 90 Days")
BTN_CUSTOM = _b("⏱ Custom")

# ── Deposit methods ────────────────────────────────────────────────
BTN_BKASH   = _b("🟢 bKash")
BTN_NAGAD   = _b("🟠 Nagad")
BTN_BINANCE = _b("🟡 Binance")

# ── Admin menu ─────────────────────────────────────────────────────
BTN_ADM_DASHBOARD     = _b("📊 Dashboard")
BTN_ADM_PRODUCTS      = _b("🛒 Products")
BTN_ADM_STOCK         = _b("📦 Stock")
BTN_ADM_USERS         = _b("👥 Users")
BTN_ADM_DEPOSITS      = _b("💰 Deposits")
BTN_ADM_VPN_ORDERS    = _b("🛡 VPN Orders")
BTN_ADM_PROXY_ORDERS  = _b("🌍 Proxy Orders")
BTN_ADM_COUPONS       = _b("🎫 Coupons")
BTN_ADM_BROADCAST     = _b("📢 Broadcast")
BTN_ADM_SETTINGS      = _b("⚙️ Settings")
BTN_ADM_EXPORT        = _b("📤 Export Mail Orders")
BTN_ADM_PROXY_PKGS    = _b("📡 Data Packages")
BTN_ADM_ORDER_LOOKUP  = _b("🔍 Order Lookup")
BTN_ADM_ANALYTICS     = _b("📈 Analytics")
BTN_PKG_ADD           = _b("➕ Add Package")

# ── Admin sub-menu buttons ─────────────────────────────────────────
BTN_ADM_ORDERS_MENU   = _b("📋 Orders")
BTN_ADM_REPORTS_MENU  = _b("📊 Reports")
BTN_ADM_TOOLS_MENU    = _b("🔧 Tools")

# ── Admin panel enhancements (FEAT-002) ────────────────────────────
BTN_ADM_LOW_STOCK     = _b("🚨 Low Stock Alerts")
BTN_ADM_ROLES         = _b("👑 Admin Roles")
BTN_ADM_EXPORT_TX     = _b("📊 Export Transactions")
BTN_ADM_STATUS_MGR    = _b("🔘 Service Status")
BTN_ADM_PROFIT        = _b("💹 Profit Calculator")
BTN_ADM_LOGS          = _b("📜 Activity Logs")
BTN_ADM_PRICE_SYNC    = _b("💲 Price Sync")
BTN_ADM_AUTO_IMPORT   = _b("📥 Auto Import")

# ── Admin product actions ──────────────────────────────────────────
BTN_EDIT_NAME   = _b("✏️ Edit Name")
BTN_EDIT_PRICE  = _b("💵 Edit Price")
BTN_EDIT_EMOJI  = _b("😀 Edit Emoji")
BTN_EDIT_DESC   = _b("📝 Edit Description")
BTN_HIDE_PROD          = _b("🙈 Hide Product")
BTN_SHOW_PROD          = _b("👁 Show Product")
BTN_DELETE             = _b("🗑 Delete")
BTN_ADD_PRODUCT        = _b("➕ Add Product")
BTN_DOWNLOAD_PRODUCTS  = _b("📥 Download Products")
BTN_IMPORT_PRODUCTS    = _b("📤 Import Products")
BTN_ADD_STOCK   = _b("📥 Add Stock")

# ── Admin stock actions ────────────────────────────────────────────
BTN_UPLOAD_FILE      = _b("📤 Upload File")
BTN_MANUAL_ADD       = _b("✏️ Manual Add")
BTN_CLEAR_STOCK      = _b("🗑 Clear Stock")
BTN_GET_TEMPLATE     = _b("📋 Download Template")
BTN_DOWNLOAD_STOCK     = _b("📥 Download Stock")
BTN_DOWNLOAD_ALL_STOCK = _b("📦 Download All Stock")
BTN_IMPORT_ALL_STOCK   = _b("📤 Import All Stock")

# ── Admin user actions ─────────────────────────────────────────────
BTN_BAN_USER    = _b("⛔ Ban User")
BTN_UNBAN_USER  = _b("✅ Unban User")
BTN_ADD_BAL     = _b("💰 Add Balance")
BTN_REMOVE_BAL  = _b("💸 Remove Balance")
BTN_ADD_BONUS   = _b("🎁 Add Bonus")
BTN_REMOVE_BONUS = _b("🗑 Remove Bonus")
BTN_BONUS_ALL_USERS = _b("🎁 Bonus All Users")
BTN_BONUS_TOP10 = _b("🏆 Bonus Top 10")
BTN_SEARCH_USER = _b("🔍 Search User")

# ── Admin coupon actions ───────────────────────────────────────────
BTN_DELETE_COUPON = _b("🗑 Delete Coupon")
BTN_CREATE_COUPON = _b("➕ Create Coupon")

# ── Categories ─────────────────────────────────────────────────────
BTN_CAT_MAIL  = _b("📮 Mail (Auto Delivery)")
BTN_CAT_VPN   = _b("🌐 VPN (Manual)")
BTN_CAT_PROXY = _b("🔐 Proxy (Auto+Manual)")

# ── Proxy delivery mode toggle ──────────────────────────────────────
BTN_TOGGLE_AUTO   = _b("🤖 Set Auto Delivery")
BTN_TOGGLE_MANUAL = _b("👤 Set Manual Delivery")


def is_admin(user_id: int) -> bool:
    return user_id in ADMIN_IDS


# ══════════════════════════════════════════════════════════════════
# FIREBASE — Initialise
# ══════════════════════════════════════════════════════════════════

firebase_admin.initialize_app(
    credentials.Certificate(json.loads(FIREBASE_CREDS_JSON)),
    {"databaseURL": FIREBASE_DB_URL, "storageBucket": FIREBASE_BUCKET},
)


async def _run(fn, *args, **kwargs):
    loop = asyncio.get_running_loop()
    return await loop.run_in_executor(None, lambda: fn(*args, **kwargs))


async def _safe_delete(msg) -> None:
    try:
        await msg.delete()
    except Exception:
        pass


def _session_has_oauth2(session: dict) -> bool:
    return bool(session.get("refresh_token") and session.get("client_id"))


def _truncate_msg(text: str, limit: int = 3900) -> str:
    if len(text) <= limit:
        return text
    return text[:limit] + "\n\n<i>… (truncated, too long)</i>"


# ══════════════════════════════════════════════════════════════════
# FIREBASE — Generic helpers
# ══════════════════════════════════════════════════════════════════

async def db_get(path: str) -> Any:
    return await _run(fdb.reference(path).get)

async def db_set(path: str, data: Any) -> None:
    await _run(fdb.reference(path).set, data)

async def db_update(path: str, data: dict) -> None:
    await _run(fdb.reference(path).update, data)

async def db_push(path: str, data: Any) -> str:
    ref = await _run(fdb.reference(path).push, data)
    return ref.key

async def db_delete(path: str) -> None:
    await _run(fdb.reference(path).delete)


async def generate_id(prefix: str) -> str:
    """Generate a human-friendly incrementing ID like ORD-0001, PRD-001."""
    ref = fdb.reference(f"counters/{prefix}")
    new_val = await _run(ref.transaction, lambda current: (current or 0) + 1)
    if prefix == "PRD":
        return f"{prefix}-{new_val:03d}"
    return f"{prefix}-{new_val:04d}"


# ══════════════════════════════════════════════════════════════════
# FIREBASE — Users
# ══════════════════════════════════════════════════════════════════

async def get_user(uid: int) -> Optional[dict]:
    return await db_get(f"users/{uid}")

async def get_all_users() -> Dict[str, dict]:
    return await db_get("users") or {}

async def create_or_update_user(uid: int, username: str, full_name: str) -> dict:
    existing = await get_user(uid)
    if existing:
        await db_update(f"users/{uid}", {"username": username, "full_name": full_name})
        return {**existing, "username": username, "full_name": full_name}
    data = {
        "user_id": uid, "username": username, "full_name": full_name,
        "balance": 0.0, "is_banned": False, "joined_at": int(time.time()),
        "total_spent": 0.0, "order_count": 0,
        "referral": {"referred_by": None, "referred_count": 0, "total_earned": 0.0},
    }
    await db_set(f"users/{uid}", data)
    return data

async def update_balance(uid: int, delta: float) -> float:
    user = await get_user(uid)
    new_bal = round((user.get("balance") or 0) + delta, 4)
    await db_update(f"users/{uid}", {"balance": new_bal})
    return new_bal

async def ban_user(uid: int, banned: bool) -> None:
    await db_update(f"users/{uid}", {"is_banned": banned})


async def get_bonus(uid: int) -> dict:
    bonus = await db_get(f"users/{uid}/bonus")
    if not bonus:
        return {"amount": 0.0, "allowed_products": []}
    return {
        "amount": bonus.get("amount", 0.0),
        "allowed_products": bonus.get("allowed_products", []),
    }


async def set_bonus(uid: int, amount: float, allowed_products: list) -> None:
    await db_set(f"users/{uid}/bonus", {"amount": amount, "allowed_products": allowed_products})


async def use_bonus(uid: int, amount: float) -> float:
    bonus = await get_bonus(uid)
    new_amount = round(max(bonus["amount"] - amount, 0), 4)
    await db_update(f"users/{uid}/bonus", {"amount": new_amount})
    return new_amount


def compute_bonus_usage(bonus: dict, pid: str, total: float) -> tuple:
    """Return (bonus_used, remaining_cost) for a purchase."""
    bonus_used = 0.0
    allowed = bonus.get("allowed_products", [])
    if ("*" in allowed or pid in allowed) and bonus.get("amount", 0) > 0:
        if bonus["amount"] >= total:
            bonus_used = total
        else:
            bonus_used = bonus["amount"]
    remaining_cost = round(total - bonus_used, 4)
    return bonus_used, remaining_cost

async def get_totp_secret(uid: int) -> Optional[str]:
    user = await get_user(uid)
    return user.get("totp_secret") if user else None

async def set_totp_secret(uid: int, secret: str) -> None:
    await db_update(f"users/{uid}", {"totp_secret": secret})


# ── Referral helpers ───────────────────────────────────────────────

async def get_referral_info(uid: int) -> dict:
    data = await db_get(f"users/{uid}/referral")
    if not data:
        return {"referred_by": None, "referred_count": 0, "total_earned": 0.0}
    return {
        "referred_by": data.get("referred_by"),
        "referred_count": data.get("referred_count", 0),
        "total_earned": data.get("total_earned", 0.0),
    }


async def set_referrer(uid: int, referrer_uid: int) -> None:
    await db_update(f"users/{uid}/referral", {
        "referred_by": referrer_uid,
        "referred_count": 0,
        "total_earned": 0.0,
    })
    ref_info = await get_referral_info(referrer_uid)
    new_count = ref_info.get("referred_count", 0) + 1
    await db_update(f"users/{referrer_uid}/referral", {"referred_count": new_count})


async def add_referral_earning(referrer_uid: int, amount: float, from_uid: int, order_id: str) -> None:
    await db_push(f"users/{referrer_uid}/referral_earnings", {
        "amount": amount,
        "from_uid": from_uid,
        "order_id": order_id,
        "ts": int(time.time()),
    })
    ref_info = await get_referral_info(referrer_uid)
    new_total = round(ref_info.get("total_earned", 0.0) + amount, 4)
    await db_update(f"users/{referrer_uid}/referral", {"total_earned": new_total})



async def pay_referral_commission(bot: Bot, buyer_uid: int, purchase_amount: float, order_id: str) -> None:
    """Credit referral bonus to the referrer after purchase.

    The referrer earns a Bonus Balance (not real money) that can only
    be used as a discount on future purchases.
    """
    try:
        ref_info = await get_referral_info(buyer_uid)
        referrer_uid = ref_info.get("referred_by")
        if referrer_uid:
            settings = await get_settings()
            bonus_pct = settings.get("referral_bonus_pct", 5.0)
            commission = round(purchase_amount * bonus_pct / 100, 4)
            if commission > 0:
                # Credit referrer's BONUS balance (not real money)
                current_bonus = await get_bonus(referrer_uid)
                new_amount = round(current_bonus["amount"] + commission, 4)
                existing_allowed = current_bonus.get("allowed_products", [])
                if "*" not in existing_allowed:
                    existing_allowed = existing_allowed + ["*"]
                await set_bonus(referrer_uid, new_amount, existing_allowed)
                await add_referral_earning(referrer_uid, commission, buyer_uid, order_id)
                try:
                    await bot.send_message(
                        referrer_uid,
                        f"\U0001f3af <b>Referral Bonus!</b>\n{_SEP}\n"
                        f"Your referral made a purchase.\n"
                        f"\U0001f381 Bonus: <b>${commission:.2f}</b> added to your Bonus Balance.\n"
                        f"Use it as a discount on your next purchase!",
                    )
                except Exception:
                    pass
    except Exception as _ref_err:
        logger.warning("Referral commission failed: %s", _ref_err)


# ══════════════════════════════════════════════════════════════════
# FIREBASE — Settings
# ══════════════════════════════════════════════════════════════════

_DEFAULT_SETTINGS = {
    "usd_rate": 125.0,
    "bkash_number": "", "nagad_number": "", "binance_uid": "",
    "bkash_min": 100.0, "nagad_min": 150.0, "binance_min": 5.0,
    "support_username": SUPPORT_USERNAME,
    "welcome_message": WELCOME_MSG,
    "force_join_channel": DEFAULT_FORCE_JOIN or "",
    "force_join_channel_2": "https://t.me/NMShopUpdate",
    "low_stock_threshold": 5,
    "get_2fa_link": "",
    "shop_name": "🛍 Neroxa Shop",
    "proxy_data_options": "1 GB,5 GB,10 GB,50 GB",
    "stock_export_interval": 30,
    "referral_bonus_pct": 5.0,
    "maintenance_mode": "OFF",
    "maintenance_message": "\ud83d\udd27 Bot is under maintenance. Please try again later.",
    "backup_interval_hours": 24,
}

async def get_settings() -> dict:
    data = await db_get("settings")
    if not data:
        await db_set("settings", _DEFAULT_SETTINGS)
        return dict(_DEFAULT_SETTINGS)
    return {**_DEFAULT_SETTINGS, **data}

async def update_settings(updates: dict) -> None:
    await db_update("settings", updates)


# ══════════════════════════════════════════════════════════════════
# FIREBASE — Products
# ══════════════════════════════════════════════════════════════════

async def get_all_products() -> Dict[str, dict]:
    return await db_get("products") or {}

async def get_product(pid: str) -> Optional[dict]:
    return await db_get(f"products/{pid}")

async def create_product(name: str, price: float, emoji: str = "📮", category: str = "mail",
                         delivery_mode: str = "manual", hidden: bool = False,
                         description: str = "") -> str:
    pid = await generate_id("PRD")
    await db_set(f"products/{pid}", {
        "name": name, "price": price, "emoji": emoji, "category": category,
        "delivery_mode": delivery_mode, "hidden": hidden, "description": description,
        "stock_count": 0, "created_at": int(time.time()), "total_sold": 0,
    })
    return pid

async def update_product(pid: str, updates: dict) -> None:
    await db_update(f"products/{pid}", updates)

async def delete_product(pid: str) -> None:
    await db_delete(f"products/{pid}")
    await db_delete(f"stocks/{pid}")

async def get_service_products(category: str) -> Dict[str, dict]:
    all_p = await get_all_products()
    return {k: v for k, v in all_p.items() if v.get("category") == category and not v.get("hidden")}


# ══════════════════════════════════════════════════════════════════
# FIREBASE — Stock  (BUG FIX: item reassignment inside loop)
# ══════════════════════════════════════════════════════════════════

async def add_stock_items(pid: str, items: List[str], bot=None) -> int:
    """Add items to stock. Returns number of items actually added."""
    existing = await db_get(f"stocks/{pid}") or {}
    before = len(existing)
    was_empty = before == 0
    for raw_item in items:
        cleaned = raw_item.strip()
        if cleaned:
            existing[uuid.uuid4().hex[:16]] = cleaned
    await db_set(f"stocks/{pid}", existing)
    count = len(existing)
    # Always unhide the product when stock is added
    update_data = {"stock_count": count}
    if count > 0:
        update_data["hidden"] = False
    await db_update(f"products/{pid}", update_data)
    added = count - before
    # Notify restock subscribers if product was out of stock
    if was_empty and added > 0 and bot:
        asyncio.create_task(_notify_restock_subscribers(pid, bot))
    return added  # Return actually added count, not total


async def _notify_restock_subscribers(pid: str, bot) -> None:
    """Notify users who subscribed to restock notifications for this product."""
    try:
        product = await get_product(pid)
        if not product:
            return
        subscribers = await get_restock_subscribers(pid)
        if not subscribers:
            return
        name = product.get("name", "Unknown")
        emoji = product.get("emoji", "\U0001F4E6")
        stock = product.get("stock_count", 0)
        for uid in subscribers:
            try:
                await bot.send_message(
                    uid,
                    f"\U0001F514 <b>Restock Alert!</b>\n{_SEP}\n"
                    f"{emoji} <b>{name}</b> is back in stock!\n"
                    f"\U0001F4E6 Available: <b>{stock}</b> items\n\n"
                    f"Hurry up before it sells out!",
                )
            except Exception:
                pass
        await clear_restock_subscribers(pid)
    except Exception as e:
        logger.warning("Restock notification failed: %s", e)

async def pop_stock_items(pid: str, qty: int) -> List[str]:
    existing = await db_get(f"stocks/{pid}") or {}
    keys = list(existing.keys())[:qty]
    popped = [existing[k] for k in keys]
    for k in keys:
        del existing[k]
    await db_set(f"stocks/{pid}", existing)
    new_count = len(existing)
    await db_update(f"products/{pid}", {"stock_count": new_count})
    return popped

async def clear_stock(pid: str) -> None:
    await db_delete(f"stocks/{pid}")
    await db_update(f"products/{pid}", {"stock_count": 0})

async def get_stock_count(pid: str) -> int:
    data = await db_get(f"stocks/{pid}")
    return len(data) if data else 0


# ══════════════════════════════════════════════════════════════════
# FIREBASE — Activity Logs (FEAT-002)
# ══════════════════════════════════════════════════════════════════

async def log_admin_action(admin_id: int, action: str, details: str = "") -> None:
    """Log an admin action to activity_logs/."""
    await db_push("activity_logs", {
        "admin_id": admin_id,
        "action": action,
        "details": details,
        "timestamp": int(time.time()),
    })


async def get_activity_logs(limit: int = 20, offset: int = 0) -> List[dict]:
    """Get recent activity logs sorted by timestamp descending."""
    data = await db_get("activity_logs") or {}
    logs = [{**v, "log_id": k} for k, v in data.items()]
    logs.sort(key=lambda x: x.get("timestamp", 0), reverse=True)
    return logs[offset:offset + limit]


# ══════════════════════════════════════════════════════════════════
# FIREBASE — Admin Roles (FEAT-002)
# ══════════════════════════════════════════════════════════════════

async def get_admin_role(uid: int) -> str:
    """Get admin role. Returns 'owner' for ADMIN_IDS, checks Firebase for others."""
    if uid in ADMIN_IDS:
        role = await db_get(f"admin_roles/{uid}")
        if role:
            return role.get("role", "owner") if isinstance(role, dict) else str(role)
        return "owner"
    role_data = await db_get(f"admin_roles/{uid}")
    if role_data:
        return role_data.get("role", "none") if isinstance(role_data, dict) else str(role_data)
    return "none"


async def set_admin_role(uid: int, role: str) -> None:
    """Set admin role in Firebase."""
    await db_set(f"admin_roles/{uid}", {"role": role, "updated_at": int(time.time())})


async def get_all_admin_roles() -> Dict[str, dict]:
    """Get all admin roles from Firebase."""
    return await db_get("admin_roles") or {}


def check_role_permission(role: str, action: str) -> bool:
    """Check if a role has permission for an action.
    owner: full access
    admin: full access except role management
    moderator: view-only, no destructive actions
    """
    if role == "owner":
        return True
    if role == "admin":
        return action != "manage_roles"
    if role == "moderator":
        return action in ("view_dashboard", "view_analytics", "view_users",
                          "view_orders", "view_logs", "view_products")
    return False


# ══════════════════════════════════════════════════════════════════
# FIREBASE — Orders
# ══════════════════════════════════════════════════════════════════

async def create_order(uid: int, pid: str, product_name: str,
                       qty: int, total_price: float, items: List[str]) -> str:
    oid = await generate_id("ORD")
    await db_set(f"orders/{oid}", {
        "user_id": uid, "product_id": pid, "product_name": product_name,
        "qty": qty, "total_price": total_price, "items": items,
        "status": "delivered", "created_at": int(time.time()),
    })
    u = await get_user(uid) or {}
    await db_update(f"users/{uid}", {
        "total_spent": round((u.get("total_spent") or 0) + total_price, 4),
        "order_count": (u.get("order_count") or 0) + 1,
    })
    prod = await get_product(pid) or {}
    await db_update(f"products/{pid}", {
        "total_sold": (prod.get("total_sold") or 0) + qty,
    })
    return oid

async def get_user_orders(uid: int) -> List[dict]:
    data = await db_get("orders")
    if not data:
        return []
    return sorted(
        [{**v, "order_id": k} for k, v in data.items() if v.get("user_id") == uid],
        key=lambda x: x.get("created_at", 0), reverse=True,
    )


# ══════════════════════════════════════════════════════════════════
# FIREBASE — VPN Orders
# ══════════════════════════════════════════════════════════════════

async def create_vpn_order(uid: int, username: str, pid: str,
                           product_name: str, duration_days: int, price: float) -> str:
    u = await get_user(uid) or {}
    await db_update(f"users/{uid}", {
        "total_spent": round((u.get("total_spent") or 0) + price, 4),
        "order_count": (u.get("order_count") or 0) + 1,
    })
    prod = await get_product(pid) or {}
    await db_update(f"products/{pid}", {
        "total_sold": (prod.get("total_sold") or 0) + 1,
    })
    oid = await generate_id("VPN")
    await db_set(f"vpn_orders/{oid}", {
        "user_id": uid, "username": username, "product_id": pid,
        "product_name": product_name, "duration_days": duration_days,
        "price": price, "status": "pending", "created_at": int(time.time()),
    })
    return oid

async def update_vpn_order(oid: str, updates: dict) -> None:
    await db_update(f"vpn_orders/{oid}", updates)

async def get_pending_vpn_orders() -> List[dict]:
    data = await db_get("vpn_orders")
    if not data:
        return []
    return [{**v, "order_id": k} for k, v in data.items() if v.get("status") == "pending"]

async def get_user_vpn_orders(uid: int) -> List[dict]:
    data = await db_get("vpn_orders")
    if not data:
        return []
    return sorted(
        [{**v, "order_id": k} for k, v in data.items() if v.get("user_id") == uid],
        key=lambda x: x.get("created_at", 0), reverse=True,
    )


# ══════════════════════════════════════════════════════════════════
# FIREBASE — Proxy Orders
# ══════════════════════════════════════════════════════════════════

async def create_proxy_order(uid: int, username: str, pid: str,
                             product_name: str, duration_days, price: float,
                             items: Optional[List[str]] = None) -> str:
    delivery_mode = "auto" if items else "manual"
    status        = "delivered" if delivery_mode == "auto" else "pending"
    u = await get_user(uid) or {}
    await db_update(f"users/{uid}", {
        "total_spent": round((u.get("total_spent") or 0) + price, 4),
        "order_count": (u.get("order_count") or 0) + 1,
    })
    prod = await get_product(pid) or {}
    await db_update(f"products/{pid}", {
        "total_sold": (prod.get("total_sold") or 0) + 1,
    })
    oid = await generate_id("PXY")
    await db_set(f"proxy_orders/{oid}", {
        "user_id": uid, "username": username, "product_id": pid,
        "product_name": product_name, "duration_days": duration_days,
        "price": price, "status": status, "delivery_mode": delivery_mode,
        "items": items or [], "created_at": int(time.time()),
    })
    return oid

async def update_proxy_order(oid: str, updates: dict) -> None:
    await db_update(f"proxy_orders/{oid}", updates)

async def get_pending_proxy_orders() -> List[dict]:
    data = await db_get("proxy_orders")
    if not data:
        return []
    return [{**v, "order_id": k} for k, v in data.items() if v.get("status") == "pending"]

async def get_user_proxy_orders(uid: int) -> List[dict]:
    data = await db_get("proxy_orders")
    if not data:
        return []
    return sorted(
        [{**v, "order_id": k} for k, v in data.items() if v.get("user_id") == uid],
        key=lambda x: x.get("created_at", 0), reverse=True,
    )


# ══════════════════════════════════════════════════════════════════
# FIREBASE — Deposits
# ══════════════════════════════════════════════════════════════════

async def create_deposit(uid: int, username: str, method: str,
                         amount_bdt: float, amount_usd: float,
                         trx_id: str, screenshot_url: str) -> str:
    return await db_push("deposits", {
        "user_id": uid, "username": username, "method": method,
        "amount_bdt": amount_bdt, "amount_usd": amount_usd,
        "trx_id": trx_id, "screenshot_url": screenshot_url,
        "status": "pending", "created_at": int(time.time()),
    })

async def get_deposit(dep_id: str) -> Optional[dict]:
    return await db_get(f"deposits/{dep_id}")

async def update_deposit(dep_id: str, updates: dict) -> None:
    await db_update(f"deposits/{dep_id}", updates)

async def get_pending_deposits() -> List[dict]:
    data = await db_get("deposits")
    if not data:
        return []
    return sorted(
        [{**v, "deposit_id": k} for k, v in data.items() if v.get("status") == "pending"],
        key=lambda x: x.get("created_at", 0),
    )

async def check_trx_duplicate(trx_id: str) -> bool:
    data = await db_get("deposits")
    if not data:
        return False
    return any(v.get("trx_id") == trx_id for v in data.values())

async def get_user_deposits(uid: int) -> List[dict]:
    data = await db_get("deposits")
    if not data:
        return []
    return sorted(
        [{**v, "deposit_id": k} for k, v in data.items() if v.get("user_id") == uid],
        key=lambda x: x.get("created_at", 0), reverse=True,
    )


# ══════════════════════════════════════════════════════════════════
# FIREBASE — Coupons
# ══════════════════════════════════════════════════════════════════

async def get_coupon(code: str) -> Optional[dict]:
    data = await db_get("coupons")
    if not data:
        return None
    for k, v in data.items():
        if v.get("code", "").upper() == code.upper():
            return {**v, "coupon_id": k}
    return None

async def create_coupon(code: str, discount_pct: float, max_uses: int, expires_at: int) -> str:
    return await db_push("coupons", {
        "code": code.upper(), "discount_pct": discount_pct,
        "max_uses": max_uses, "used_count": 0, "expires_at": expires_at, "active": True,
    })

async def use_coupon(coupon_id: str) -> None:
    c = await db_get(f"coupons/{coupon_id}")
    used = (c.get("used_count") or 0) + 1
    await db_update(f"coupons/{coupon_id}", {
        "used_count": used,
        "active": used < c.get("max_uses", 1),
    })

async def get_all_coupons() -> List[dict]:
    data = await db_get("coupons")
    return [{**v, "coupon_id": k} for k, v in data.items()] if data else []

async def delete_coupon(coupon_id: str) -> None:
    await db_delete(f"coupons/{coupon_id}")


# ══════════════════════════════════════════════════════════════════
# FIREBASE — VIP System
# ══════════════════════════════════════════════════════════════════

_VIP_TIERS = [
    (500, "Diamond", "💠"),
    (200, "Gold", "🥇"),
    (50, "Silver", "🥈"),
    (10, "Bronze", "🥉"),
]


def get_vip_tier(total_spent: float) -> Tuple[str, str]:
    """Return (tier_name, emoji) based on total_spent. Returns ('', '') if no tier."""
    for threshold, name, emoji in _VIP_TIERS:
        if total_spent >= threshold:
            return name, emoji
    return "", ""


def get_vip_progress(total_spent: float) -> Tuple[str, float, float]:
    """Return (next_tier_name, current_spent, next_threshold) for progress display."""
    for i, (threshold, name, _emoji) in enumerate(_VIP_TIERS):
        if total_spent >= threshold:
            if i == 0:
                return "Max", total_spent, threshold
            prev_threshold = _VIP_TIERS[i - 1][0]
            return _VIP_TIERS[i - 1][1], total_spent, prev_threshold
    return "Bronze", total_spent, 10.0


# ══════════════════════════════════════════════════════════════════
# FIREBASE — Favorites
# ══════════════════════════════════════════════════════════════════

async def get_user_favorites(uid: int) -> List[str]:
    """Return list of product IDs favorited by user."""
    data = await db_get(f"users/{uid}/favorites")
    if not data:
        return []
    if isinstance(data, dict):
        return list(data.keys())
    return list(data) if isinstance(data, list) else []


async def add_user_favorite(uid: int, pid: str) -> None:
    await db_set(f"users/{uid}/favorites/{pid}", True)


async def remove_user_favorite(uid: int, pid: str) -> None:
    await db_delete(f"users/{uid}/favorites/{pid}")


# ══════════════════════════════════════════════════════════════════
# FIREBASE — Restock Notifications
# ══════════════════════════════════════════════════════════════════

async def subscribe_restock(pid: str, uid: int) -> None:
    await db_set(f"restock_subs/{pid}/{uid}", int(time.time()))


async def unsubscribe_restock(pid: str, uid: int) -> None:
    await db_delete(f"restock_subs/{pid}/{uid}")


async def get_restock_subscribers(pid: str) -> List[int]:
    data = await db_get(f"restock_subs/{pid}")
    if not data:
        return []
    return [int(uid) for uid in data.keys()]


async def clear_restock_subscribers(pid: str) -> None:
    await db_delete(f"restock_subs/{pid}")


# ══════════════════════════════════════════════════════════════════
# FIREBASE — Stats & Storage
# ══════════════════════════════════════════════════════════════════

async def get_dashboard_stats() -> dict:
    users, orders, vpn, proxy, deposits = await asyncio.gather(
        db_get("users"), db_get("orders"),
        db_get("vpn_orders"), db_get("proxy_orders"), db_get("deposits"),
    )
    rev = (
        sum(v.get("total_price", 0) for v in (orders or {}).values()) +
        sum(v.get("price", 0) for v in (vpn or {}).values() if v.get("status") == "delivered") +
        sum(v.get("price", 0) for v in (proxy or {}).values() if v.get("status") == "delivered")
    )
    return {
        "total_users":      len(users) if users else 0,
        "total_revenue":    round(rev, 2),
        "total_sales":      len(orders or {}),
        "pending_deposits": sum(1 for v in (deposits or {}).values() if v.get("status") == "pending"),
        "pending_orders":   (
            sum(1 for v in (vpn or {}).values() if v.get("status") == "pending") +
            sum(1 for v in (proxy or {}).values() if v.get("status") == "pending")
        ),
    }


async def get_analytics_data() -> dict:
    users, orders, vpn, proxy, products = await asyncio.gather(
        db_get("users"), db_get("orders"),
        db_get("vpn_orders"), db_get("proxy_orders"), db_get("products"),
    )
    orders = orders or {}
    vpn = vpn or {}
    proxy = proxy or {}
    users = users or {}
    products = products or {}

    now = int(time.time())
    today_start = now - (now % 86400)
    week_start = now - 7 * 86400
    month_start = now - 30 * 86400

    all_orders = []
    for v in orders.values():
        all_orders.append((v.get("created_at", 0), v.get("total_price", 0)))
    for v in vpn.values():
        if v.get("status") == "delivered":
            all_orders.append((v.get("created_at", 0), v.get("price", 0)))
    for v in proxy.values():
        if v.get("status") == "delivered":
            all_orders.append((v.get("created_at", 0), v.get("price", 0)))

    today_rev = 0.0
    today_count = 0
    week_rev = 0.0
    week_count = 0
    month_rev = 0.0
    month_count = 0

    for created_at, price in all_orders:
        if created_at >= today_start:
            today_rev += price
            today_count += 1
        if created_at >= week_start:
            week_rev += price
            week_count += 1
        if created_at >= month_start:
            month_rev += price
            month_count += 1

    top_products = sorted(
        [(k, v.get("name", k), v.get("total_sold", 0)) for k, v in products.items()],
        key=lambda x: x[2], reverse=True,
    )[:5]

    top_buyers = sorted(
        [(uid, u.get("username", ""), u.get("total_spent", 0)) for uid, u in users.items()],
        key=lambda x: x[2], reverse=True,
    )[:5]

    return {
        "today_rev": round(today_rev, 2),
        "today_count": today_count,
        "week_rev": round(week_rev, 2),
        "week_count": week_count,
        "month_rev": round(month_rev, 2),
        "month_count": month_count,
        "top_products": top_products,
        "top_buyers": top_buyers,
    }


async def upload_screenshot(file_bytes: bytes, filename: str) -> str:
    def _up():
        b = storage.bucket()
        blob = b.blob(f"deposits/{filename}")
        blob.upload_from_string(file_bytes, content_type="image/jpeg")
        blob.make_public()
        return blob.public_url
    return await _run(_up)


# ══════════════════════════════════════════════════════════════════
# STOCK FILE PARSER  (BUG FIX: robust xlsx parsing)
# ══════════════════════════════════════════════════════════════════

def _norm(line: str) -> str:
    line = line.strip()
    if not line:
        return ""
    if "," in line and ":" not in line:
        parts = line.split(",", 1)
        return ":".join(p.strip() for p in parts)
    return line


def parse_stock_file(content: bytes, filename: str) -> List[str]:
    fname = (filename or "").lower()

    if fname.endswith(".xlsx"):
        items: List[str] = []
        try:
            wb = openpyxl.load_workbook(io.BytesIO(content), read_only=True, data_only=True)
            ws = wb.active
            if ws is None:
                logger.warning("XLSX: no active sheet found")
                return []
            first_row = True
            for row in ws.iter_rows(values_only=True):
                # FIX: filter None, convert to str, strip whitespace
                cols = [str(c).strip() for c in row if c is not None and str(c).strip() not in ("", "None")]
                # Skip header row
                if first_row:
                    first_row = False
                    if any(c.lower() in ("email", "password", "account", "mail", "#", "no.", "number", "pass", "username", "login") for c in cols):
                        continue
                if len(cols) >= 2:
                    items.append(f"{cols[0]}:{cols[1]}")
                elif len(cols) == 1:
                    items.append(cols[0])
            wb.close()
        except Exception as e:
            logger.error("XLSX parse error: %s", e)
            raise ValueError(f"Cannot read xlsx file: {e}") from e
        return [i for i in items if i.strip()]

    if fname.endswith(".csv"):
        items = []
        try:
            text = content.decode("utf-8-sig", errors="replace")  # FIX: handle BOM
        except Exception:
            text = content.decode("latin-1", errors="replace")
        first_row = True
        for line in text.splitlines():
            line = line.strip()
            if not line:
                continue
            parts = [p.strip() for p in line.split(",")]
            parts = [p for p in parts if p]
            # Skip header row
            if first_row:
                first_row = False
                if any(p.lower() in ("email", "password", "account", "mail", "#", "no.", "number", "pass", "username", "login") for p in parts):
                    continue
            if len(parts) >= 2:
                items.append(f"{parts[0]}:{parts[1]}")
            elif len(parts) == 1:
                items.append(parts[0])
        return [i for i in items if i]

    # Plain text (.txt or unknown)
    try:
        text = content.decode("utf-8-sig", errors="replace")
    except Exception:
        text = content.decode("latin-1", errors="replace")
    _header_kw = ("email", "password", "account", "mail", "#", "no.", "number", "pass", "username", "login")
    items = []
    first_row = True
    for ln in text.splitlines():
        normed = _norm(ln)
        if not normed:
            continue
        if first_row:
            first_row = False
            parts = [p.strip().lower() for p in normed.replace(":", ",").split(",") if p.strip()]
            if any(p in _header_kw for p in parts):
                continue
        items.append(normed)
    return items


# ══════════════════════════════════════════════════════════════════
# DONGVAN MAIL API — Service detection & helpers
# ══════════════════════════════════════════════════════════════════

_DV_SERVICE_KEYWORDS: Dict[str, List[str]] = {
    "FACEBOOK":  ["security@facebookmail.com", "facebook", "fb"],
    "INSTAGRAM": ["instagram", "mail@instagram.com"],
    "TWITTER":   ["twitter", "x.com", "notify@twitter.com"],
    "GOOGLE":    ["google", "gmail", "youtube", "no-reply@google.com"],
    "APPLE":     ["apple", "icloud", "no-reply@apple.com"],
    "TIKTOK":    ["tiktok", "douyin"],
    "AMAZON":    ["amazon", "shipment-tracking@amazon.com"],
    "SHOPEE":    ["shopee"],
    "TELEGRAM":  ["telegram"],
    "KAKAOTALK": ["kakao"],
    "LAZADA":    ["lazada"],
    "WECHAT":    ["wechat", "weixin"],
    "OUTLOOK":   ["outlook", "microsoft", "hotmail", "no-reply@microsoft.com"],
    "LINKEDIN":  ["linkedin"],
    "NETFLIX":   ["netflix"],
    "DISCORD":   ["discord"],
    "SNAPCHAT":  ["snapchat"],
    "GARENA":    ["garena", "account@garena.com"],
    "COINBASE":  ["coinbase"],
    "BINANCE":   ["binance"],
}
_DV_SERVICE_LIST = list(_DV_SERVICE_KEYWORDS.keys())


def _dv_determine_service(subject: str = "", sender_name: str = "", sender_address: str = "") -> str:
    text = f"{subject} {sender_name} {sender_address}".lower()
    for svc, keywords in _DV_SERVICE_KEYWORDS.items():
        if any(kw.lower() in text for kw in keywords):
            return svc
    return "UNKNOWN"


def _dv_extract_codes(messages_data: Optional[dict]) -> List[dict]:
    found: List[dict] = []
    if not messages_data or not messages_data.get("status") or not messages_data.get("messages"):
        return found
    _CODE_PATTERNS = [
        r'(\d{4,8})\s+is your',
        r'[Cc]ode[:\s]+(\d{4,8})',
        r'[Cc]onfirmation code[:\s]*(\d{4,8})',
        r'[Vv]erification code[:\s]*(\d{4,8})',
        r'OTP[:\s]*(\d{4,8})',
        r'mã xác nhận[:\s]*(\d{4,8})',
        r'is\s+(\d{4,8})\s',
        r'security code[:\s]*(\d{4,8})',
    ]
    for msg in messages_data["messages"]:
        subject  = msg.get("subject", "") or ""
        body     = msg.get("message", "") or ""
        from_f   = msg.get("from")
        sname    = from_f[0].get("name", "") if from_f and isinstance(from_f, list) and from_f else ""
        saddr    = from_f[0].get("address", "") if from_f and isinstance(from_f, list) and from_f else ""
        code     = msg.get("code", "") or ""
        if not code:
            combined = subject + " " + body[:2000]
            for pattern in _CODE_PATTERNS:
                m = re.search(pattern, combined)
                if m:
                    code = m.group(1)
                    break
        if code:
            found.append({
                "code": code, "service": _dv_determine_service(subject, sname, saddr),
                "sender": sname, "address": saddr, "subject": subject,
                "date": msg.get("date", "") or "", "uid": msg.get("uid", "") or "",
            })
    return found


def _dv_parse_input(text: str) -> dict:
    parts = text.strip().split("|")
    result = {"email": None, "password": None, "refresh_token": None, "client_id": None, "has_oauth2": False}
    if len(parts) >= 2:
        result["email"]    = parts[0].strip()
        result["password"] = parts[1].strip()
    if len(parts) >= 4:
        result["refresh_token"] = parts[2].strip()
        result["client_id"]     = parts[3].strip()
        result["has_oauth2"]    = True
    return result


async def _dv_oauth2(email: str, password: str) -> Optional[dict]:
    def _call():
        try:
            r = _requests.post(_DV_OAUTH2_URL, json={"email": email, "password": password, "apikey": DONGVAN_API_KEY}, timeout=30)
            d = r.json()
            if d.get("status") and "oauth2" in d:
                parts = d["oauth2"].split("|")
                if len(parts) >= 2:
                    return {"refresh_token": parts[0], "client_id": parts[1]}
        except Exception as e:
            logger.error("DV oauth2 error: %s", e)
        return None
    return await _run(_call)


async def _dv_get_messages(email: str, rt: str, cid: str) -> Optional[dict]:
    def _call():
        try:
            r = _requests.post(_DV_GET_MSGS_URL, json={"email": email, "refresh_token": rt, "client_id": cid, "list_mail": "all"}, timeout=30)
            return r.json()
        except Exception as e:
            logger.error("DV get_messages error: %s", e)
            return None
    return await _run(_call)


async def _dv_get_code(email: str, rt: str, cid: str, code_type: str = "all") -> Optional[dict]:
    def _call():
        try:
            r = _requests.post(_DV_GET_CODE_URL, json={"email": email, "refresh_token": rt, "client_id": cid, "type": code_type}, timeout=30)
            return r.json()
        except Exception as e:
            logger.error("DV get_code error: %s", e)
            return None
    return await _run(_call)


async def _dv_graph_messages(email: str, rt: str, cid: str) -> Optional[dict]:
    def _call():
        try:
            r = _requests.post(_DV_GRAPH_MSGS_URL, json={"email": email, "refresh_token": rt, "client_id": cid, "list_mail": "all"}, timeout=30)
            return r.json()
        except Exception as e:
            logger.error("DV graph_messages error: %s", e)
            return None
    return await _run(_call)


async def _dv_process_mailbox(session: dict) -> List[dict]:
    email, rt, cid = session["email"], session["refresh_token"], session["client_id"]
    msgs = await _dv_get_messages(email, rt, cid)
    if msgs and msgs.get("status") and msgs.get("messages"):
        return _dv_extract_codes(msgs)
    msgs = await _dv_graph_messages(email, rt, cid)
    if msgs and msgs.get("status") and msgs.get("messages"):
        return _dv_extract_codes(msgs)
    return []


def _dv_build_codes_display(messages: List[dict]) -> str:
    if not messages:
        return "❌ No codes found in mailbox."
    by_svc: Dict[str, List[dict]] = {}
    for m in messages:
        by_svc.setdefault(m["service"], []).append(m)
    out = f"{_SEP}\n📬 <b>CODES FOUND</b>\n{_SEP}\n\n"
    for svc, msgs in by_svc.items():
        out += f"<b>{svc}</b>\n"
        for m in msgs:
            out += f"├─ <code>{html_lib.escape(m['code'])}</code>\n"
            out += f"└─ {html_lib.escape(m['subject'][:60])}\n"
        out += "\n"
    out += f"{_SEP}\n📋 <b>COPY:</b>\n"
    for svc, msgs in by_svc.items():
        for m in msgs:
            out += f"<code>{svc} → {html_lib.escape(m['code'])}</code>\n"
    return out


def _dv_build_mailbox_display(messages: List[dict], filter_service: Optional[str] = None) -> str:
    filtered = [m for m in messages if m["service"] == filter_service.upper()] if filter_service else messages
    title    = f"📁 Filter: <b>{filter_service.upper()}</b>" if filter_service else "📁 <b>Mailbox</b>"
    if not filtered:
        return f"{title}\n{_SEP}\n❌ No messages found."
    out = f"{title}\n{_SEP}\n📊 Total: <b>{len(filtered)}</b> message(s)\n{_SEP}\n\n"
    for i, m in enumerate(filtered[:20], 1):
        code_display = f"<b>{html_lib.escape(m['code'])}</b>" if m.get("code") else "No code"
        out += f"<b>#{i}</b> <b>[{m['service']}]</b> {code_display}\n"
        out += f"└─ 📝 {html_lib.escape(m['subject'][:80])}\n"
        out += f"└─ 👤 {html_lib.escape(m.get('sender',''))}\n"
        if m.get("date"):
            out += f"└─ 🕐 {str(m['date'])[:19]}\n"
        out += "\n"
    if len(filtered) > 20:
        out += f"<i>... and {len(filtered) - 20} more</i>\n\n"
    codes_with_code = [m for m in filtered if m.get("code")]
    if codes_with_code:
        out += f"{_SEP}\n📋 <b>CODES:</b>\n"
        for m in codes_with_code:
            out += f"<code>{m['service']} → {html_lib.escape(m['code'])}</code>\n"
    return out


# ══════════════════════════════════════════════════════════════════
# FORMATTERS
# ══════════════════════════════════════════════════════════════════

TZ    = pytz.UTC
_SEP  = "━" * 22
_LINE = "─" * 22

# ══════════════════════════════════════════════════════════════════
# UI HELPERS
# ══════════════════════════════════════════════════════════════════


def _progress_bar(current: float, maximum: float, width: int = 10) -> str:
    """Visual progress bar using block characters."""
    if maximum <= 0:
        ratio = 0.0
    else:
        ratio = min(current / maximum, 1.0)
    filled = int(ratio * width)
    empty = width - filled
    bar = "\u2588" * filled + "\u2591" * empty
    pct = int(ratio * 100)
    return f"[{bar}] {pct}%"


def _mini_chart(values: list) -> str:
    """Sparkline text chart from a list of numeric values."""
    if not values:
        return ""
    spark_chars = "\u2581\u2582\u2583\u2584\u2585\u2586\u2587\u2588"
    mn = min(values)
    mx = max(values)
    rng = mx - mn if mx != mn else 1
    return "".join(spark_chars[min(int((v - mn) / rng * 7), 7)] for v in values)


def _status_dot(is_online: bool) -> str:
    """Green/red circle emoji for status indication."""
    return "\U0001F7E2" if is_online else "\U0001F534"


def _quick_actions_inline(actions: list) -> InlineKeyboardMarkup:
    """Generate quick-action button rows from list of (text, callback_data) tuples."""
    rows = []
    row = []
    for i, (txt, cb) in enumerate(actions):
        row.append(InlineKeyboardButton(text=txt, callback_data=cb))
        if len(row) == 3 or i == len(actions) - 1:
            rows.append(row)
            row = []
    return InlineKeyboardMarkup(inline_keyboard=rows)


def _dt(ts: int) -> str:
    return datetime.fromtimestamp(ts, tz=TZ).strftime("%d %b %Y, %H:%M UTC")


def fmt_welcome(shop_name: str, welcome: str, first_name: str, balance: float,
                user: dict = None) -> str:
    """Clean welcome message with VIP badge and quick stats."""
    tier_name, tier_emoji = ("", "")
    total_spent = 0
    order_count = 0
    if user:
        total_spent = user.get("total_spent", 0)
        order_count = user.get("order_count", 0)
        tier_name, tier_emoji = get_vip_tier(total_spent)
    vip_badge = f" {tier_emoji} {tier_name}" if tier_name else ""
    return (
        f"<b>{shop_name}</b>\n{_SEP}\n"
        f"{welcome}\n\n"
        f"\U0001F464 <b>{first_name}</b>{vip_badge}\n"
        f"\U0001F4B0 Balance: <code>${balance:.2f}</code>\n"
        f"\U0001F6D2 Orders: <b>{order_count}</b>  |  \U0001F4C8 Spent: <b>${total_spent:.2f}</b>\n"
        f"{_SEP}\n"
        f"Status: \u2705 Active"
    )

def fmt_balance_screen(user: dict) -> str:
    banned = "\U0001F6AB Banned" if user.get("is_banned") else "\u2705 Active"
    bonus = user.get("bonus")
    bonus_line = ""
    if bonus and bonus.get("amount", 0) > 0:
        bonus_line = f"\U0001F381 Bonus: <code>${bonus['amount']:.2f}</code>\n"
    total_spent = user.get("total_spent", 0)
    tier_name, tier_emoji = get_vip_tier(total_spent)
    next_tier, current, threshold = get_vip_progress(total_spent)
    vip_line = f"{tier_emoji} VIP: <b>{tier_name}</b>\n" if tier_name else ""
    progress = _progress_bar(current, threshold) if next_tier != "Max" else _progress_bar(1, 1)
    member_since = _dt(user.get("joined_at", int(time.time())))
    return (
        f"\U0001F4B0 <b>My Balance</b>\n{_SEP}\n"
        f"\U0001F464 <b>{user.get('full_name', 'User')}</b>\n"
        f"\U0001F194 <code>{user.get('user_id', '?')}</code>  @{user.get('username') or 'N/A'}\n"
        f"{vip_line}"
        f"{_LINE}\n"
        f"\U0001F4B5 Balance: <code>${user.get('balance', 0):.2f}</code>\n"
        f"{bonus_line}"
        f"\U0001F4B8 Spent: <b>${total_spent:.2f}</b>\n"
        f"\U0001F6D2 Orders: <b>{user.get('order_count', 0)}</b>\n"
        f"{_LINE}\n"
        f"\U0001F3C6 Progress: {progress}\n"
        f"   Next: <b>{next_tier}</b> (${threshold:.0f})\n"
        f"{_LINE}\n"
        f"\U0001F4C5 Member: {member_since}\n"
        f"\U0001F7E2 Status: {banned}"
    )

def fmt_product_list_header(category_name: str, count: int) -> str:
    return (
        f"<b>{category_name}</b>\n{_SEP}\n"
        f"\U0001F4E6 <b>{count}</b> product(s) available\n"
        f"Select a product below:"
    )

def fmt_product_detail(product: dict, stock: int) -> str:
    desc = product.get("description", "").strip()
    desc_line = f"\U0001F4DD {desc}\n" if desc else ""
    status = _status_dot(stock > 0)
    stock_bar = _progress_bar(stock, max(stock, 50), width=8)
    return (
        f"{product.get('emoji', chr(0x1F4E6))} <b>{product['name']}</b>\n{_SEP}\n"
        f"{desc_line}"
        f"{status} Status: {'In Stock' if stock > 0 else 'Out of Stock'}\n"
        f"\U0001F4B5 Price: <code>${product['price']:.2f}</code> /account\n"
        f"\U0001F4E6 Stock: <b>{stock}</b> {stock_bar}\n"
        f"{_LINE}\n"
        f"Enter the quantity you want to purchase:"
    )

def fmt_service_detail(product: dict, duration_days: int) -> str:
    price = round(product["price"] * duration_days, 4)
    desc = product.get("description", "").strip()
    desc_line = f"\U0001F4DD {desc}\n" if desc else ""
    return (
        f"{product.get('emoji', chr(0x1F310))} <b>{product['name']}</b>\n{_SEP}\n"
        f"{desc_line}"
        f"\U0001F4B5 Price/day: <code>${product['price']:.2f}</code>\n"
        f"\u23F1 Duration: <b>{duration_days} day(s)</b>\n"
        f"\U0001F4B0 Total: <code>${price:.2f}</code>\n"
        f"{_LINE}\n"
        f"Press \u2705 Confirm Purchase to place your order."
    )

def fmt_order_summary(product: dict, qty: int, total: float,
                      balance: float, discount_pct: float = 0) -> str:
    disc_line = f"🎟 Coupon: <b>-{discount_pct:.0f}%</b>\n" if discount_pct else ""
    return (
        f"🛒 <b>Order Summary</b>\n{_SEP}\n"
        f"{product.get('emoji', '📦')} {product['name']}\n"
        f"🔢 Quantity: <b>{qty}</b>\n"
        f"💵 Price: <b>${product['price']:.2f}</b> × {qty}\n"
        f"{disc_line}"
        f"💰 Total: <b>${total:.2f}</b>\n{_LINE}\n"
        f"👛 Your Balance: <b>${balance:.2f}</b>\n\n"
        f"Have a coupon code? Enter it below,\nor press ✅ Confirm Purchase to skip."
    )

def fmt_confirm_service(category: str, product: dict, days: int,
                        price: float, balance: float) -> str:
    return (
        f"🛒 <b>Order Summary</b>\n{_SEP}\n"
        f"{product.get('emoji', '📦')} {product['name']}\n"
        f"⏱ Duration: <b>{days} day(s)</b>\n"
        f"💰 Total: <b>${price:.2f}</b>\n{_LINE}\n"
        f"👛 Your Balance: <b>${balance:.2f}</b>\n\n"
        f"Press ✅ Confirm Purchase to place order."
    )

def fmt_order_receipt(oid: str, product_name: str, qty: int,
                      total: float, items: List[str], new_balance: float) -> str:
    preview = "\n".join(items[:5])
    more = f"\n<i>... and {len(items) - 5} more</i>" if len(items) > 5 else ""
    return (
        f"✅ <b>Order Successful!</b>\n{_SEP}\n"
        f"🆔 Order: <code>{oid}</code>\n"
        f"📦 {product_name} × {qty}\n"
        f"💵 Paid: <b>${total:.2f}</b>\n"
        f"👛 Balance left: <b>${new_balance:.2f}</b>\n{_SEP}\n"
        f"📋 <b>Your Accounts:</b>\n"
        f"<code>{preview}{more}</code>"
    )

def fmt_service_order_placed(cat_emoji: str, cat_name: str,
                              oid: str, product_name: str, days: int, price: float) -> str:
    return (
        f"{cat_emoji} <b>{cat_name} Order Placed!</b>\n{_SEP}\n"
        f"🆔 Order: <code>{oid}</code>\n"
        f"📦 {product_name}\n"
        f"⏱ Duration: <b>{days} day(s)</b>\n"
        f"💵 Paid: <b>${price:.2f}</b>\n{_SEP}\n"
        f"⏳ Status: <b>Pending</b>\n"
        f"Our team will deliver your order shortly.\n"
        f"You'll receive a notification when ready. 🔔"
    )

def fmt_deposit_info(method_label: str, number: str, rate: float, min_amt: float, is_usd: bool = False) -> str:
    if is_usd:
        return (
            f"💵 <b>Deposit via {method_label}</b>\n{_SEP}\n"
            f"🔑 UID: <code>{number or 'Not configured'}</code>\n"
            f"💰 Minimum: <b>${min_amt:.2f} USD</b>\n{_LINE}\n"
            f"Enter the amount you are sending (in USD):"
        )
    return (
        f"💵 <b>Deposit via {method_label}</b>\n{_SEP}\n"
        f"📱 Send to: <code>{number or 'Not configured'}</code>\n"
        f"💱 Rate: <b>1 USD = {rate:.0f} BDT</b>\n"
        f"💰 Minimum: <b>{min_amt:.0f} BDT</b>\n{_LINE}\n"
        f"Enter the amount you are sending (in BDT):"
    )

def fmt_deposit_confirm(method: str, amount_bdt: float, amount_usd: float, is_usd: bool = False) -> str:
    amt_str = f"${amount_usd:.2f} USD" if is_usd else f"{amount_bdt:.0f} BDT ≈ ${amount_usd:.2f}"
    return (
        f"💵 <b>Deposit Summary</b>\n{_SEP}\n"
        f"💱 Method: <b>{method}</b>\n"
        f"💰 Amount: <b>{amt_str}</b>\n{_LINE}\n"
        f"Now send your Transaction ID / Reference number:"
    )

def fmt_deposit_submitted(dep_id: str, method: str, amount_bdt: float, amount_usd: float, is_usd: bool = False) -> str:
    amt_str = f"${amount_usd:.2f} USD" if is_usd else f"{amount_bdt:.0f} BDT ≈ ${amount_usd:.2f}"
    return (
        f"✅ <b>Deposit Request Submitted!</b>\n{_SEP}\n"
        f"🆔 ID: <code>{dep_id[:12]}</code>\n"
        f"💱 Method: <b>{method}</b>\n"
        f"💰 Amount: <b>{amt_str}</b>\n{_SEP}\n"
        f"⏳ <b>Status: Pending Review</b>\n"
        f"Your balance will be credited once verified.\n"
        f"Usually within <b>5–30 minutes</b>. 🕐"
    )

def fmt_order_history(orders: list, vpn: list, proxy: list) -> str:
    if not orders and not vpn and not proxy:
        return (
            f"\U0001F4ED <b>No Orders Yet</b>\n{_SEP}\n"
            f"You haven't placed any orders.\n"
            f"Browse our products to get started!"
        )
    lines = []
    if orders:
        lines.append("\U0001F4EE <b>Mail Orders</b>")
        for o in orders[:8]:
            lines.append(f"  <code>{o['order_id']}</code>  {o['product_name']} \u00D7{o['qty']}  ${o['total_price']:.2f}")
    if vpn:
        lines.append("")
        lines.append("\U0001F310 <b>VPN Orders</b>")
        for o in vpn[:8]:
            s = {"delivered": "\u2705", "cancelled": "\u274C"}.get(o["status"], "\u23F3")
            lines.append(f"  <code>{o['order_id']}</code>  {o['product_name']} {o['duration_days']}d  ${o['price']:.2f}  {s}")
    if proxy:
        lines.append("")
        lines.append("\U0001F510 <b>Proxy Orders</b>")
        for o in proxy[:8]:
            s = {"delivered": "\u2705", "cancelled": "\u274C"}.get(o["status"], "\u23F3")
            data_lbl = o.get("duration_days", "-")
            lines.append(f"  <code>{o['order_id']}</code>  {o['product_name']} {data_lbl}  ${o['price']:.2f}  {s}")
    total_orders = len(orders) + len(vpn) + len(proxy)
    return (
        f"\U0001F4DC <b>Order History</b>\n{_SEP}\n"
        + "\n".join(lines)
        + f"\n{_SEP}\n\U0001F4CB Total: {total_orders} order(s)"
    )

def fmt_deposit_history(deposits: list) -> str:
    if not deposits:
        return (
            f"\U0001F4B3 <b>No Deposits Yet</b>\n{_SEP}\n"
            f"Make your first deposit to start shopping!"
        )
    lines = []
    for d in deposits[:12]:
        s = {"approved": "\u2705", "rejected": "\u274C"}.get(d["status"], "\u23F3")
        lines.append(f"{s} <code>{d['deposit_id'][:8]}</code>  {d['method']}  {d['amount_bdt']:.0f}BDT  ${d['amount_usd']:.2f}")
    return f"\U0001F4B3 <b>Deposit History</b>\n{_SEP}\n" + "\n".join(lines)

def fmt_admin_dashboard(stats: dict) -> str:
    pending_dep = stats.get('pending_deposits', 0)
    pending_ord = stats.get('pending_orders', 0)
    alert = ""
    if pending_dep > 0 or pending_ord > 0:
        alert = f"\n\U0001F6A8 <b>ALERTS:</b> {pending_dep} deposits, {pending_ord} orders pending"
    return (
        f"\U0001F4CA <b>Admin Dashboard</b>\n{_SEP}\n"
        f"\U0001F465 Users: <b>{stats['total_users']}</b>\n"
        f"\U0001F4B5 Revenue: <code>${stats['total_revenue']:.2f}</code>\n"
        f"\U0001F6D2 Sales: <b>{stats['total_sales']}</b>\n"
        f"{_LINE}\n"
        f"\U0001F4B3 Pending Deposits: <b>{pending_dep}</b>\n"
        f"\U0001F4E6 Pending Orders: <b>{pending_ord}</b>"
        f"{alert}"
    )


def fmt_analytics(data: dict) -> str:
    body_lines = [
        f"<b>Today</b>",
        f"  \U0001F4B5 Revenue: <code>${data['today_rev']:.2f}</code>",
        f"  \U0001F6D2 Orders: <b>{data['today_count']}</b>",
        "",
        f"<b>This Week (7 days)</b>",
        f"  \U0001F4B5 Revenue: <code>${data['week_rev']:.2f}</code>",
        f"  \U0001F6D2 Orders: <b>{data['week_count']}</b>",
        "",
        f"<b>This Month (30 days)</b>",
        f"  \U0001F4B5 Revenue: <code>${data['month_rev']:.2f}</code>",
        f"  \U0001F6D2 Orders: <b>{data['month_count']}</b>",
    ]
    top_products_lines = [f"\n<b>Top 5 Products</b>"]
    if data["top_products"]:
        for i, (_, name, sold) in enumerate(data["top_products"], 1):
            top_products_lines.append(f"  {i}. {name} - <b>{sold}</b> sold")
    else:
        top_products_lines.append("  No product data.")
    top_buyers_lines = [f"\n<b>Top 5 Buyers</b>"]
    if data["top_buyers"]:
        for i, (uid, uname, spent) in enumerate(data["top_buyers"], 1):
            display = f"@{uname}" if uname else f"<code>{uid}</code>"
            top_buyers_lines.append(f"  {i}. {display} - <b>${spent:.2f}</b>")
    else:
        top_buyers_lines.append("  No buyer data.")
    body = "\n".join(body_lines + top_products_lines + top_buyers_lines)
    return f"\U0001F4C8 <b>Sales Analytics</b>\n{_SEP}\n{body}"


def fmt_admin_deposit_review(d: dict) -> str:
    return (
        f"💳 <b>Deposit Request</b>\n{_SEP}\n"
        f"🆔 <code>{d['deposit_id']}</code>\n"
        f"👤 @{d.get('username','—')}  (<code>{d['user_id']}</code>)\n"
        f"💱 {d['method']}\n"
        f"💰 {d['amount_bdt']:.0f} BDT ≈ <b>${d['amount_usd']:.2f}</b>\n"
        f"🔑 TRX: <code>{d.get('trx_id','—')}</code>\n"
        f"🕒 {_dt(d['created_at'])}"
    )

def fmt_admin_vpn_order(o: dict) -> str:
    return (
        f"🌐 <b>New VPN Order</b>\n{_SEP}\n"
        f"🆔 <code>{o['order_id']}</code>\n"
        f"👤 @{o.get('username','—')}  (<code>{o['user_id']}</code>)\n"
        f"📦 {o['product_name']}\n"
        f"⏱ {o['duration_days']} day(s)\n"
        f"💵 ${o['price']:.2f}\n"
        f"🕒 {_dt(o['created_at'])}"
    )

def fmt_admin_proxy_order(o: dict) -> str:
    data_lbl = o.get("duration_days", "—")
    return (
        f"🔐 <b>New Proxy Order</b>\n{_SEP}\n"
        f"🆔 <code>{o['order_id']}</code>\n"
        f"👤 @{o.get('username','—')}  (<code>{o['user_id']}</code>)\n"
        f"📦 {o['product_name']}\n"
        f"📡 Data: {data_lbl}\n"
        f"💵 ${o['price']:.2f}\n"
        f"🕒 {_dt(o['created_at'])}"
    )

def fmt_user_info(user: dict, mail_orders: int = 0, vpn_orders: int = 0, proxy_orders: int = 0) -> str:
    banned = "⛔ Banned" if user.get("is_banned") else "✅ Active"
    bonus = user.get("bonus")
    bonus_section = ""
    if bonus and bonus.get("amount", 0) > 0:
        allowed = bonus.get("allowed_products", [])
        prods = ", ".join(allowed) if allowed else "All"
        bonus_section = (
            f"\n<b>{'─' * 22}</b>\n"
            f"🎁 <b>Bonus Info</b>\n"
            f"   Amount: <b>${bonus['amount']:.2f}</b>\n"
            f"   Products: <code>{prods}</code>\n"
        )
    order_section = ""
    if mail_orders or vpn_orders or proxy_orders:
        order_section = (
            f"\n<b>{'─' * 22}</b>\n"
            f"📊 <b>Order Breakdown</b>\n"
            f"   📨 Mail: <b>{mail_orders}</b>\n"
            f"   🛡 VPN: <b>{vpn_orders}</b>\n"
            f"   🌍 Proxy: <b>{proxy_orders}</b>\n"
        )
    return (
        f"👤 <b>User Profile</b>\n{_SEP}\n"
        f"🆔 ID: <code>{user['user_id']}</code>\n"
        f"👤 Username: @{user.get('username') or 'N/A'}\n"
        f"📛 Name: {user.get('full_name') or 'N/A'}\n"
        f"\n<b>{'─' * 22}</b>\n"
        f"💎 <b>Financial</b>\n"
        f"   Balance: <b>${user.get('balance', 0):.2f}</b>\n"
        f"   Total Spent: <b>${user.get('total_spent', 0):.2f}</b>\n"
        f"{bonus_section}"
        f"\n<b>{'─' * 22}</b>\n"
        f"📋 <b>Activity</b>\n"
        f"   Orders: <b>{user.get('order_count', 0)}</b>\n"
        f"   Joined: {_dt(user.get('joined_at', 0))}\n"
        f"   Status: {banned}"
        f"{order_section}"
    )

def fmt_settings(s: dict) -> str:
    ch1  = s.get('force_join_channel') or '—'
    ch2  = s.get('force_join_channel_2') or '—'
    bkno = s.get('bkash_number') or '—'
    ngno = s.get('nagad_number') or '—'
    bnid = s.get('binance_uid') or '—'
    return (
        f"⚙️ <b>Bot Settings</b>\n{_SEP}\n"
        f"🏪 <b>{s.get('shop_name') or 'Shop'}</b>  |  💱 <b>1 USD = {s.get('usd_rate', 125):.0f} BDT</b>\n"
        f"⚠️ Low Stock Alert: <b>{s.get('low_stock_threshold', 5)}</b>\n"
        f"\n"
        f"<b>── 💳 Payment ──</b>\n"
        f"📱 bKash  <code>{bkno}</code>  · min <b>{s.get('bkash_min', 100):.0f} BDT</b>\n"
        f"📱 Nagad  <code>{ngno}</code>  · min <b>{s.get('nagad_min', 150):.0f} BDT</b>\n"
        f"🔶 Binance  <code>{bnid}</code>  · min <b>${s.get('binance_min', 5):.2f}</b>\n"
        f"\n"
        f"<b>── 🔗 Links ──</b>\n"
        f"📣 Ch1: <b>{ch1}</b>  |  Ch2: <b>{ch2}</b>\n"
        f"🆘 Support: <b>{s.get('support_username') or '—'}</b>\n"
        f"🔑 2FA Link: <b>{s.get('get_2fa_link') or '—'}</b>\n"
        f"\n"
        f"<b>── 📡 Proxy ──</b>\n"
        f"Data Options: <code>{s.get('proxy_data_options', '1 GB,5 GB,10 GB,50 GB')}</code>\n"
        f"\n"
        f"<b>── ⏱ Auto Export ──</b>\n"
        f"Stock Export Interval: <b>{s.get('stock_export_interval', 30):.0f} min</b>\n"
        f"\n"
        f"<b>── 🎯 Referral ──</b>\n"
        f"Referral Bonus: <b>{s.get('referral_bonus_pct', 5.0):.1f}%</b>\n"
        f"\n"
        f"<b>── 🔧 Maintenance ──</b>\n"
        f"🔧 Maintenance: <b>{str(s.get('maintenance_mode', 'OFF')).upper()}</b>\n"
        f"\n"
        f"<b>── 💾 Auto Backup ──</b>\n"
        f"💾 Auto Backup: <b>{'every ' + str(int(s.get('backup_interval_hours', 24))) + 'h' if s.get('backup_interval_hours', 24) > 0 else 'Disabled'}</b>"
    )


# ══════════════════════════════════════════════════════════════════
# VALIDATORS
# ══════════════════════════════════════════════════════════════════

def validate_quantity(text: str, max_stock: int) -> Tuple[Optional[int], Optional[str]]:
    try:
        qty = int(text.strip())
    except ValueError:
        return None, "❌ Please enter a valid number."
    if qty <= 0:
        return None, "❌ Quantity must be at least 1."
    if qty > max_stock:
        return None, f"❌ Only <b>{max_stock}</b> in stock."
    return qty, None

def validate_amount_bdt(text: str, minimum: float) -> Tuple[Optional[float], Optional[str]]:
    try:
        amount = float(text.strip())
    except ValueError:
        return None, "❌ Please enter a valid amount."
    if amount <= 0:
        return None, "❌ Amount must be positive."
    if amount < minimum:
        return None, f"❌ Minimum deposit is <b>{minimum:.0f} BDT</b>"
    return amount, None

def validate_custom_days(text: str) -> Tuple[Optional[int], Optional[str]]:
    try:
        days = int(text.strip())
    except ValueError:
        return None, "❌ Enter a valid number of days."
    if days < 1:
        return None, "❌ Minimum 1 day."
    if days > 365:
        return None, "❌ Maximum 365 days."
    return days, None

def validate_coupon(coupon: dict) -> Tuple[bool, Optional[str]]:
    # NOTE: Coupon usage is tracked globally (used_count), not per-user.
    # A single user can use the same coupon multiple times until max_uses is reached.
    if not coupon.get("active"):
        return False, "❌ Coupon is no longer active."
    if coupon.get("expires_at", 0) < int(time.time()):
        return False, "❌ Coupon has expired."
    if coupon.get("used_count", 0) >= coupon.get("max_uses", 1):
        return False, "❌ Coupon usage limit reached."
    return True, None

def validate_price(text: str) -> Tuple[Optional[float], Optional[str]]:
    try:
        p = float(text.strip())
    except ValueError:
        return None, "❌ Enter a valid price."
    if p < 0:
        return None, "❌ Price cannot be negative."
    return p, None

def validate_discount(text: str) -> Tuple[Optional[float], Optional[str]]:
    try:
        pct = float(text.strip())
    except ValueError:
        return None, "❌ Enter a valid percentage."
    if not (0 < pct <= 100):
        return None, "❌ Must be between 1 and 100."
    return pct, None


# ══════════════════════════════════════════════════════════════════
# FSM STATES
# ══════════════════════════════════════════════════════════════════

class UserFlow(StatesGroup):
    mail_product   = State()
    mail_qty       = State()
    mail_coupon       = State()
    mail_coupon_input = State()
    mail_confirm      = State()
    set_totp_key      = State()
    vpn_product      = State()
    vpn_duration     = State()
    vpn_custom       = State()
    vpn_coupon       = State()
    vpn_coupon_input = State()
    vpn_confirm      = State()
    proxy_product           = State()
    proxy_duration          = State()
    proxy_custom            = State()
    proxy_coupon            = State()
    proxy_coupon_input      = State()
    proxy_confirm           = State()
    proxy_auto_qty          = State()
    proxy_auto_coupon       = State()
    proxy_auto_coupon_input = State()
    proxy_auto_confirm      = State()
    dep_method     = State()
    dep_amount     = State()
    dep_trx        = State()
    dep_screenshot = State()
    # Get Code / Mailbox flow
    get_code_menu   = State()
    get_code_set    = State()
    get_code_filter = State()
    # Temp Mail flow
    temp_mail_menu   = State()
    temp_mail_domain = State()
    # User Panel Enhancements
    favorites_menu       = State()
    search_services      = State()
    order_tracking       = State()
    user_stats           = State()
    more_menu            = State()
    # AI Support flow
    ai_support           = State()
    ai_support_ask       = State()

class AdminFlow(StatesGroup):
    menu              = State()
    products_list     = State()
    product_detail    = State()
    product_edit      = State()
    product_add_name  = State()
    product_add_emoji = State()
    product_add_price = State()
    product_add_cat   = State()
    product_add_desc  = State()
    stock_list        = State()
    stock_detail      = State()
    stock_uploading   = State()
    stock_manual      = State()
    stock_import_all  = State()
    user_search       = State()
    user_menu         = State()
    user_detail       = State()
    user_add_bal      = State()
    user_remove_bal   = State()
    user_bonus_amount = State()
    user_bonus_products = State()
    coupons_list      = State()
    coupon_detail     = State()
    coupon_code       = State()
    coupon_discount   = State()
    coupon_max_uses   = State()
    coupon_expiry     = State()
    broadcast         = State()
    settings_menu     = State()
    settings_edit     = State()
    vpn_fulfill       = State()
    proxy_fulfill     = State()
    dep_reject_reason = State()
    products_import   = State()
    bonus_all_amount  = State()
    bonus_all_products = State()
    bonus_top10_amount = State()
    bonus_top10_products = State()
    proxy_pkg_list    = State()
    proxy_pkg_add     = State()
    order_lookup      = State()
    # FEAT-002: Admin Panel Enhancements
    analytics_detail  = State()
    low_stock_alerts  = State()
    price_sync        = State()
    auto_import       = State()
    roles_menu        = State()
    roles_assign      = State()
    advanced_search   = State()
    export_transactions = State()
    export_tx_dates   = State()
    service_status    = State()
    profit_calc       = State()
    profit_set_cost   = State()
    activity_logs     = State()
    # Sub-menu states
    orders_submenu    = State()
    reports_submenu   = State()
    tools_submenu     = State()


# ══════════════════════════════════════════════════════════════════
# KEYBOARDS
# ══════════════════════════════════════════════════════════════════

def _kb(*rows, resize: bool = True, one_time: bool = False) -> ReplyKeyboardMarkup:
    return ReplyKeyboardMarkup(
        keyboard=[[KeyboardButton(text=t) for t in row] for row in rows],
        resize_keyboard=resize,
        one_time_keyboard=one_time,
    )

def main_menu_kb() -> ReplyKeyboardMarkup:
    return _kb(
        [BTN_BALANCE,   BTN_DEPOSIT],
        [BTN_GET_MAIL,  BTN_TEMP_MAIL],
        [BTN_BUY_VPN,   BTN_BUY_PROXY],
        [BTN_GET_CODE,  BTN_GET_2FA],
        [BTN_HISTORY,   BTN_REFERRAL],
        [BTN_MORE],
        [BTN_SUPPORT,   BTN_AI_SUPPORT],
    )

def banned_user_kb() -> ReplyKeyboardMarkup:
    return _kb([BTN_SUPPORT])

def more_menu_kb() -> ReplyKeyboardMarkup:
    return _kb(
        [BTN_TRENDING,  BTN_FAVORITES],
        [BTN_SEARCH,    BTN_MY_STATS],
        [BACK_BTN, HOME_BTN],
    )

def ai_support_kb() -> ReplyKeyboardMarkup:
    return _kb(
        [BTN_AI_HOW_BUY,     BTN_AI_HOW_DEPOSIT],
        [BTN_AI_STOCK,       BTN_AI_REFERRAL],
        [BTN_AI_ORDERS,      BTN_AI_2FA],
        [BTN_AI_MAIL,        BTN_AI_ASK],
        [BACK_BTN, HOME_BTN],
    )

def _nav_row() -> List[str]:
    return [BACK_BTN, HOME_BTN]

def products_kb(display_map: dict) -> ReplyKeyboardMarkup:
    rows = [[k] for k in display_map.keys()]
    rows.append([BACK_BTN, HOME_BTN])
    return _kb(*rows)

DURATION_MAP = {
    BTN_1DAY: 1, BTN_7DAYS: 7,
    BTN_30DAYS: 30, BTN_90DAYS: 90,
    BTN_CUSTOM: -1,
}

def duration_kb() -> ReplyKeyboardMarkup:
    return _kb(
        [BTN_1DAY,   BTN_7DAYS],
        [BTN_30DAYS, BTN_90DAYS],
        [BTN_CUSTOM],
        [BACK_BTN, HOME_BTN],
    )

BTN_PROXY_DATA_CUSTOM = _b("📡 Custom Amount")

def parse_proxy_data_options(options_str: str) -> List[str]:
    return [o.strip() for o in options_str.split(",") if o.strip()]

def proxy_data_kb(options_str: str = "1 GB,5 GB,10 GB,50 GB") -> ReplyKeyboardMarkup:
    opts = parse_proxy_data_options(options_str)
    rows = []
    for i in range(0, len(opts), 2):
        row = [_b(opts[i])]
        if i + 1 < len(opts):
            row.append(_b(opts[i + 1]))
        rows.append(row)
    rows.append([BTN_PROXY_DATA_CUSTOM])
    rows.append([BACK_BTN, HOME_BTN])
    return _kb(*rows)

def parse_data_amount(text: str) -> Tuple[Optional[float], Optional[str]]:
    raw = text.strip()
    clean = raw.upper().replace("GB", "").replace("MB", "").strip()
    try:
        amount = float(clean)
    except ValueError:
        return None, "❌ Valid number লিখুন (e.g. 5 GB বা 500 MB)."
    if amount <= 0:
        return None, "❌ Amount must be positive."
    if amount > 100000:
        return None, "❌ Maximum 100000 GB/MB."
    return amount, None

def confirm_kb() -> ReplyKeyboardMarkup:
    """First confirmation — includes Apply Coupon button (mail flow only)."""
    return _kb([BTN_CONFIRM], [BTN_COUPON], [BACK_BTN, HOME_BTN])

def confirm_final_kb() -> ReplyKeyboardMarkup:
    """Final confirmation — no coupon button (used for 2nd mail confirm, VPN, Proxy)."""
    return _kb([BTN_CONFIRM], [BACK_BTN, HOME_BTN])

def input_kb() -> ReplyKeyboardMarkup:
    return _kb([CANCEL_BTN, HOME_BTN])

def tfa_inline_kb() -> InlineKeyboardMarkup:
    return InlineKeyboardMarkup(inline_keyboard=[
        [InlineKeyboardButton(text="🔄 Refresh Code", callback_data="tfa_refresh")],
        [
            InlineKeyboardButton(text="⏱ Timer", callback_data="tfa_timer"),
            InlineKeyboardButton(text="🔑 Change Key", callback_data="tfa_setkey"),
        ],
    ])

def tfa_nokey_kb() -> InlineKeyboardMarkup:
    return InlineKeyboardMarkup(inline_keyboard=[
        [InlineKeyboardButton(text="🔑 Set My Key", callback_data="tfa_setkey")],
    ])

def _totp_secs() -> int:
    return 30 - (int(time.time()) % 30)

def _totp_bar(secs: int) -> str:
    filled = round(secs / 30 * 10)
    return "█" * filled + "░" * (10 - filled)

def _valid_totp_key(key: str) -> bool:
    try:
        pyotp.TOTP(key).now()
        return True
    except Exception:
        return False

def _fmt_totp(code: str) -> str:
    secs = _totp_secs()
    bar  = _totp_bar(secs)
    return (
        f"🔑 <b>Your 2FA Code</b>\n{_SEP}\n"
        f"<code>{code[:3]} {code[3:]}</code>\n\n"
        f"⏱ Expires in: <b>{secs}s</b>\n"
        f"[{bar}]"
    )

DEPOSIT_MAP = {
    BTN_BKASH:   "bkash",
    BTN_NAGAD:   "nagad",
    BTN_BINANCE: "binance",
}

def deposit_method_kb() -> ReplyKeyboardMarkup:
    return _kb(
        [BTN_BKASH,   BTN_NAGAD],
        [BTN_BINANCE],
        [BACK_BTN, HOME_BTN],
    )

def admin_main_kb() -> ReplyKeyboardMarkup:
    return _kb(
        [BTN_ADM_DASHBOARD],
        [BTN_ADM_PRODUCTS,     BTN_ADM_STOCK],
        [BTN_ADM_USERS,        BTN_ADM_DEPOSITS],
        [BTN_ADM_ORDERS_MENU,  BTN_ADM_REPORTS_MENU],
        [BTN_ADM_TOOLS_MENU,   BTN_ADM_SETTINGS],
        [HOME_BTN],
    )

def admin_orders_submenu_kb() -> ReplyKeyboardMarkup:
    return _kb(
        [BTN_ADM_VPN_ORDERS,   BTN_ADM_PROXY_ORDERS],
        [BTN_ADM_ORDER_LOOKUP, BTN_ADM_EXPORT],
        [BACK_BTN, HOME_BTN],
    )

def admin_reports_submenu_kb() -> ReplyKeyboardMarkup:
    return _kb(
        [BTN_ADM_ANALYTICS,    BTN_ADM_PROFIT],
        [BTN_ADM_EXPORT_TX,    BTN_ADM_LOGS],
        [BACK_BTN, HOME_BTN],
    )

def admin_tools_submenu_kb() -> ReplyKeyboardMarkup:
    return _kb(
        [BTN_ADM_LOW_STOCK,    BTN_ADM_STATUS_MGR],
        [BTN_ADM_PRICE_SYNC,   BTN_ADM_AUTO_IMPORT],
        [BTN_ADM_ROLES,        BTN_ADM_BROADCAST],
        [BTN_ADM_COUPONS,      BTN_ADM_PROXY_PKGS],
        [BACK_BTN, HOME_BTN],
    )

def proxy_pkg_manage_kb(options: list) -> ReplyKeyboardMarkup:
    rows = [[f"🗑 {o}"] for o in options]
    rows.append([BTN_PKG_ADD])
    rows.append([BACK_BTN, HOME_BTN])
    return _kb(*rows)

SETTINGS_MAP = {
    _b("💱 USD Rate"):           ("usd_rate",             float, "Enter BDT per 1 USD (e.g. 125):"),
    _b("📱 bKash Number"):       ("bkash_number",          str,   "Enter bKash number:"),
    _b("📱 bKash Min Deposit"):  ("bkash_min",             float, "Enter minimum bKash deposit in BDT (e.g. 100):"),
    _b("📱 Nagad Number"):       ("nagad_number",          str,   "Enter Nagad number:"),
    _b("📱 Nagad Min Deposit"):  ("nagad_min",             float, "Enter minimum Nagad deposit in BDT (e.g. 150):"),
    _b("🔶 Binance UID"):        ("binance_uid",           str,   "Enter Binance Pay UID:"),
    _b("🔶 Binance Min Deposit"):("binance_min",           float, "Enter minimum Binance deposit in USD (e.g. 5):"),
    _b("📣 Force Join #1"):      ("force_join_channel",    str,   "Enter channel (e.g. @mychannel) or blank to disable:"),
    _b("📣 Force Join #2"):      ("force_join_channel_2",  str,   "Enter 2nd channel or blank to disable:"),
    _b("🆘 Support Link"):       ("support_username",      str,   "Enter support username (e.g. @support):"),
    _b("🔑 Get 2FA Link"):       ("get_2fa_link",          str,   "Enter the Get 2FA URL:"),
    _b("🏪 Shop Name"):          ("shop_name",             str,   "Enter shop display name:"),
    _b("💬 Welcome Message"):    ("welcome_message",       str,   "Enter welcome message text:"),
    _b("⚠️ Low Stock Alert"):    ("low_stock_threshold",   float, "Enter low stock threshold (number):"),
    _b("📡 Proxy Data Options"): ("proxy_data_options",    str,   "Enter data options separated by commas (e.g. 1 GB,5 GB,10 GB,50 GB).\nAdmin can set any GB/MB values here:"),
    _b("⏱ Stock Export Interval"): ("stock_export_interval", float, "Enter interval in minutes for auto stock export to group (e.g. 30):"),
    _b("🎯 Referral Bonus %"): ("referral_bonus_pct", float, "Enter referral bonus % (0-100):"),
    _b("🔧 Maintenance Mode"): ("maintenance_mode", str, "Enter ON or OFF:"),
    _b("🔧 Maintenance Msg"): ("maintenance_message", str, "Enter maintenance message text:"),
    _b("💾 Backup Interval"): ("backup_interval_hours", float, "Enter backup interval in hours (e.g. 24 for daily, 0 to disable):"),
}

# ── Settings Sub-Category Buttons ──
BTN_SET_PAYMENT  = _b("💳 Payment")
BTN_SET_CHANNELS = _b("📣 Channels")
BTN_SET_SHOP     = _b("🏪 Shop")
BTN_SET_SYSTEM   = _b("🔧 System")

SETTINGS_CATEGORIES = {
    BTN_SET_PAYMENT: [
        _b("💱 USD Rate"), _b("📱 bKash Number"), _b("📱 bKash Min Deposit"),
        _b("📱 Nagad Number"), _b("📱 Nagad Min Deposit"),
        _b("🔶 Binance UID"), _b("🔶 Binance Min Deposit"),
    ],
    BTN_SET_CHANNELS: [
        _b("📣 Force Join #1"), _b("📣 Force Join #2"), _b("🆘 Support Link"),
    ],
    BTN_SET_SHOP: [
        _b("🏪 Shop Name"), _b("💬 Welcome Message"), _b("🔑 Get 2FA Link"),
        _b("🎯 Referral Bonus %"), _b("📡 Proxy Data Options"),
    ],
    BTN_SET_SYSTEM: [
        _b("⚠️ Low Stock Alert"), _b("⏱ Stock Export Interval"),
        _b("🔧 Maintenance Mode"), _b("🔧 Maintenance Msg"), _b("💾 Backup Interval"),
    ],
}

def admin_settings_kb() -> ReplyKeyboardMarkup:
    return _kb(
        [BTN_SET_PAYMENT, BTN_SET_CHANNELS],
        [BTN_SET_SHOP, BTN_SET_SYSTEM],
        [BACK_BTN, HOME_BTN],
    )

def settings_category_kb(keys: list) -> ReplyKeyboardMarkup:
    rows = [keys[i:i+2] for i in range(0, len(keys), 2)]
    rows.append([BACK_BTN, HOME_BTN])
    return _kb(*rows)

CATEGORY_MAP = {
    BTN_CAT_MAIL:  "mail",
    BTN_CAT_VPN:   "vpn",
    BTN_CAT_PROXY: "proxy",
}

def admin_category_kb() -> ReplyKeyboardMarkup:
    return _kb(
        [BTN_CAT_MAIL],
        [BTN_CAT_VPN],
        [BTN_CAT_PROXY],
        [CANCEL_BTN, HOME_BTN],
    )

def admin_products_kb(products: dict) -> ReplyKeyboardMarkup:
    rows = [
        [f"{p.get('emoji','📦')} {p['name']} {'🙈 ' if p.get('hidden') else ''}[{p.get('category','mail').upper()}] — ${p['price']:.2f}"]
        for pid, p in products.items()
    ]
    rows.append([BTN_ADD_PRODUCT])
    rows.append([BTN_DOWNLOAD_PRODUCTS, BTN_IMPORT_PRODUCTS])
    rows.append([BACK_BTN, HOME_BTN])
    return _kb(*rows)

def build_admin_products_dm(products: dict) -> dict:
    return {
        f"{p.get('emoji','📦')} {p['name']} {'🙈 ' if p.get('hidden') else ''}[{p.get('category','mail').upper()}] — ${p['price']:.2f}": pid
        for pid, p in products.items()
    }

def admin_product_actions_kb(hidden: bool, category: str = "mail", delivery_mode: str = "manual") -> ReplyKeyboardMarkup:
    toggle = BTN_SHOW_PROD if hidden else BTN_HIDE_PROD
    rows: list = [
        [BTN_EDIT_NAME,  BTN_EDIT_PRICE],
        [BTN_EDIT_EMOJI, toggle],
        [BTN_EDIT_DESC],
    ]
    if category == "mail" or (category == "proxy" and delivery_mode == "auto"):
        rows.append([BTN_ADD_STOCK])
    if category == "proxy":
        mode_btn = BTN_TOGGLE_MANUAL if delivery_mode == "auto" else BTN_TOGGLE_AUTO
        rows.append([mode_btn])
    rows.append([BTN_DELETE, BACK_BTN])
    rows.append([HOME_BTN])
    return _kb(*rows)

def admin_stock_products_kb(products: dict) -> ReplyKeyboardMarkup:
    rows = [[f"{p.get('emoji','📦')} {p['name']}  📦{p.get('stock_count',0)}"] for p in products.values()]
    rows.append([BTN_DOWNLOAD_ALL_STOCK, BTN_IMPORT_ALL_STOCK])
    rows.append([BACK_BTN, HOME_BTN])
    return _kb(*rows)

def build_admin_stock_dm(products: dict) -> dict:
    return {
        f"{p.get('emoji','📦')} {p['name']}  📦{p.get('stock_count',0)}": pid
        for pid, p in products.items()
    }

def admin_stock_actions_kb() -> ReplyKeyboardMarkup:
    return _kb(
        [BTN_GET_TEMPLATE, BTN_DOWNLOAD_STOCK],
        [BTN_UPLOAD_FILE, BTN_MANUAL_ADD],
        [BTN_CLEAR_STOCK],
        [BACK_BTN, HOME_BTN],
    )

def admin_user_actions_kb(is_banned: bool) -> ReplyKeyboardMarkup:
    ban_btn = BTN_UNBAN_USER if is_banned else BTN_BAN_USER
    return _kb(
        [ban_btn],
        [BTN_ADD_BAL, BTN_REMOVE_BAL],
        [BTN_ADD_BONUS, BTN_REMOVE_BONUS],
        [BACK_BTN, HOME_BTN],
    )

def admin_users_menu_kb() -> ReplyKeyboardMarkup:
    return _kb(
        [BTN_SEARCH_USER],
        [BTN_BONUS_ALL_USERS, BTN_BONUS_TOP10],
        [BACK_BTN, HOME_BTN],
    )


async def _show_user_menu(target: Message, state: FSMContext):
    await state.set_state(AdminFlow.user_menu)
    all_users = await get_all_users()
    adv_kb = InlineKeyboardMarkup(inline_keyboard=[
        [InlineKeyboardButton(text="🔍 Advanced Search", callback_data="adm_adv_search")]
    ])
    await target.answer(
        f"👥 <b>User Management</b>\n{_SEP}\n"
        f"Total Users: <b>{len(all_users)}</b>\n\nSelect an option:",
        reply_markup=admin_users_menu_kb(),
    )
    await target.answer("Or use advanced search:", reply_markup=adv_kb)


def get_code_menu_kb() -> ReplyKeyboardMarkup:
    return _kb(
        [BTN_GC_SET_MAIL,  BTN_GC_CHANGE],
        [BTN_GC_CODES,     BTN_GC_INBOX],
        [BTN_GC_REFRESH,   BTN_GC_FILTER],
        [BACK_BTN, HOME_BTN],
    )

def get_code_filter_kb() -> ReplyKeyboardMarkup:
    rows = []
    row: List[str] = []
    for svc in _DV_SERVICE_LIST:
        row.append(svc)
        if len(row) == 3:
            rows.append(row)
            row = []
    if row:
        rows.append(row)
    rows.append([BACK_BTN, HOME_BTN])
    return _kb(*rows)

def admin_coupons_kb(coupons: list) -> ReplyKeyboardMarkup:
    rows = [[f"🎟 {c['code']}  {c['discount_pct']:.0f}%  ({c['used_count']}/{c['max_uses']})"] for c in coupons]
    rows.append([BTN_CREATE_COUPON])
    rows.append([BACK_BTN, HOME_BTN])
    return _kb(*rows)

def build_coupons_dm(coupons: list) -> dict:
    return {
        f"🎟 {c['code']}  {c['discount_pct']:.0f}%  ({c['used_count']}/{c['max_uses']})": c["coupon_id"]
        for c in coupons
    }

def admin_coupon_actions_kb() -> ReplyKeyboardMarkup:
    return _kb([BTN_DELETE_COUPON], [BACK_BTN, HOME_BTN])

# In-memory pending state: admin_id → context dict
_pending_fulfill: Dict[int, Dict] = {}   # VPN/Proxy credentials input
_pending_reject:  Dict[int, Dict] = {}   # Deposit reject reason input

def deposit_review_inline(dep_id: str) -> InlineKeyboardMarkup:
    return InlineKeyboardMarkup(inline_keyboard=[[
        InlineKeyboardButton(text="✅ Approve",      callback_data=f"dep_ok:{dep_id}"),
        InlineKeyboardButton(text="❌ Reject + Reason", callback_data=f"dep_no:{dep_id}"),
    ]])

def order_review_inline(oid: str, kind: str) -> InlineKeyboardMarkup:
    return InlineKeyboardMarkup(inline_keyboard=[[
        InlineKeyboardButton(text="📝 Send Credentials", callback_data=f"{kind}_ok:{oid}"),
        InlineKeyboardButton(text="❌ Cancel & Refund",   callback_data=f"{kind}_no:{oid}"),
    ]])


# ══════════════════════════════════════════════════════════════════
# MIDDLEWARE — Anti-Spam
# ══════════════════════════════════════════════════════════════════

class AntiSpamMiddleware(BaseMiddleware):
    def __init__(self):
        self._last: Dict[int, float] = {}
        self._hist: Dict[int, deque] = defaultdict(deque)

    async def __call__(self, handler, event: TelegramObject, data: Dict[str, Any]) -> Any:
        if not isinstance(event, Message):
            return await handler(event, data)
        uid = event.from_user.id if event.from_user else None
        if uid is None or is_admin(uid):
            return await handler(event, data)
        now = time.time()
        if now - self._last.get(uid, 0) < SPAM_COOLDOWN:
            await event.answer(f"⏳ Slow down! Wait {SPAM_COOLDOWN:.0f}s between messages.")
            return
        hist = self._hist[uid]
        while hist and now - hist[0] > SPAM_WINDOW:
            hist.popleft()
        if len(hist) >= SPAM_MAX_MSGS:
            await event.answer("🚫 Too many messages. Please wait a moment.")
            return
        hist.append(now)
        self._last[uid] = now
        return await handler(event, data)


# ══════════════════════════════════════════════════════════════════
# MIDDLEWARE — Auth  (BUG FIX: also handle CallbackQuery)
# ══════════════════════════════════════════════════════════════════

_maintenance_cache: Dict[str, Any] = {"value": None, "ts": 0.0}

class AuthMiddleware(BaseMiddleware):
    async def __call__(self, handler, event: TelegramObject, data: Dict[str, Any]) -> Any:
        user = None
        if isinstance(event, Message):
            user = event.from_user
        elif isinstance(event, CallbackQuery):
            user = event.from_user

        if not user:
            return await handler(event, data)

        db_user = await create_or_update_user(user.id, user.username or "", user.full_name or "")
        data["db_user"] = db_user

        if db_user.get("is_banned") and not is_admin(user.id):
            # Allow banned users to reach Support handler
            if isinstance(event, Message) and event.text == BTN_SUPPORT:
                return await handler(event, data)
            msg = (
                "🚫 <b>Account Banned</b>\n\n"
                "Your account has been banned from this shop.\n"
                "Contact support if you think this is a mistake."
            )
            if isinstance(event, Message):
                await event.answer(msg, reply_markup=banned_user_kb())
            elif isinstance(event, CallbackQuery):
                await event.answer("🚫 Account banned.", show_alert=True)
            return

        # Maintenance mode check for non-admin users
        if not is_admin(user.id):
            now = time.time()
            if now - _maintenance_cache["ts"] > 30:
                _settings = await get_settings()
                _maintenance_cache["value"] = _settings
                _maintenance_cache["ts"] = now
            else:
                _settings = _maintenance_cache["value"]
            if str(_settings.get("maintenance_mode", "OFF")).strip().upper() == "ON":
                maint_msg = _settings.get("maintenance_message") or "🔧 Bot is under maintenance. Please try again later."
                if isinstance(event, Message):
                    await event.answer(maint_msg)
                elif isinstance(event, CallbackQuery):
                    await event.answer("🔧 Bot is under maintenance.", show_alert=True)
                return

        # Force-join check for non-admin users (skip for Support button)
        if not is_admin(user.id):
            if isinstance(event, Message) and event.text == BTN_SUPPORT:
                pass  # Allow support access without force-join
            else:
                try:
                    passed, failed_channels = await check_force_join(data.get("bot") or event.bot, user.id)
                    if not passed:
                        links = []
                        for ch in failed_channels:
                            url = _channel_to_url(ch)
                            links.append(f"  • <a href='{url}'>{ch}</a>" if url else f"  • {ch}")
                        channels_text = "\n".join(links)
                        msg = (
                            f"📢 <b>Join Required!</b>\n\n"
                            f"You must join the following channel(s) to use this bot:\n{channels_text}\n\n"
                            f"After joining, try again."
                        )
                        if isinstance(event, Message):
                            await event.answer(msg, disable_web_page_preview=True)
                        elif isinstance(event, CallbackQuery):
                            await event.answer("📢 Please join required channels first.", show_alert=True)
                        return
                except Exception as e:
                    logger.warning("Force-join check failed: %s", e)

        return await handler(event, data)


# ══════════════════════════════════════════════════════════════════
# HELPER — Force-Join check  (BUG FIX: handle https:// links)
# ══════════════════════════════════════════════════════════════════

def _parse_channel(raw: str) -> str:
    """Convert any channel format to @username or numeric ID."""
    raw = raw.strip()
    if not raw:
        return ""
    if raw.startswith("https://t.me/"):
        handle = raw.split("https://t.me/")[-1].strip("/")
        if handle:
            return f"@{handle}"
    if raw.startswith("@"):
        return raw
    if raw.lstrip("-").isdigit():
        return raw
    return f"@{raw}"


def _channel_to_url(raw: str) -> str:
    """Return a clickable https://t.me/ URL from any channel format."""
    raw = raw.strip()
    if not raw:
        return ""
    if raw.startswith("https://t.me/"):
        return raw
    username = raw.lstrip("@")
    return f"https://t.me/{username}"


def _channel_display_name(raw: str) -> str:
    """Return a short human-readable name for a channel."""
    raw = raw.strip()
    if raw.startswith("https://t.me/"):
        return raw.split("https://t.me/")[-1].strip("/")
    return raw.lstrip("@")


async def check_force_join(bot: Bot, uid: int) -> Tuple[bool, List[str]]:
    settings = await get_settings()
    failed = []
    for key in ("force_join_channel", "force_join_channel_2"):
        raw_ch = (settings.get(key) or "").strip()
        if not raw_ch:
            continue
        ch = _parse_channel(raw_ch)
        if not ch:
            continue
        try:
            m = await bot.get_chat_member(ch, uid)
            if m.status in ("left", "kicked"):
                failed.append(raw_ch)
        except Exception as e:
            logger.warning("Force-join check failed for %s: %s", ch, e)
    return len(failed) == 0, failed


async def send_main_menu(message: Message, state: FSMContext, text: Optional[str] = None, user_id: Optional[int] = None) -> None:
    await state.clear()
    settings = await get_settings()
    uid = user_id or message.from_user.id
    user = await get_user(uid) or {}
    shop_name = settings.get("shop_name", "\U0001F6CD Neroxa Shop")
    welcome   = settings.get("welcome_message", WELCOME_MSG)
    balance   = user.get("balance", 0)
    first_name = user.get("first_name", message.from_user.first_name)
    display = text or fmt_welcome(shop_name, welcome, first_name, balance, user)
    await message.answer(display, reply_markup=main_menu_kb())


# ══════════════════════════════════════════════════════════════════
# GLOBAL ROUTER — Home (works from ANY state)
# ══════════════════════════════════════════════════════════════════

router_global = Router()

@router_global.message(F.text == HOME_BTN)
async def go_home(message: Message, state: FSMContext):
    # If admin is in admin flow, send them to admin panel instead
    current_state = await state.get_state()
    if is_admin(message.from_user.id) and current_state and current_state.startswith("AdminFlow:"):
        await state.set_state(AdminFlow.menu)
        await message.answer("🔐 <b>Admin Panel</b>", reply_markup=admin_main_kb())
        return
    await send_main_menu(message, state)


# ══════════════════════════════════════════════════════════════════
# TEMP MAIL — API Helpers & In-memory store
# ══════════════════════════════════════════════════════════════════

_TEMPMAIL_API = "https://api.internal.temp-mail.io/api/v3"

# In-memory store: uid -> {email, token, seen_ids: set, messages: {short_id: msg_dict}, msg_index: int}
_tempmail_sessions: Dict[int, dict] = {}


def _tempmail_request(url: str, payload: Optional[dict] = None) -> dict:
    headers = {"Content-Type": "application/json", "User-Agent": "NeroxaShopBot/1.0"}
    data = json.dumps(payload).encode() if payload else None
    try:
        req = _urllib_request.Request(url, data=data, headers=headers)
        with _urllib_request.urlopen(req, timeout=12) as res:
            return json.loads(res.read())
    except Exception as e:
        logger.error("TempMail request error: %s", e)
        return {}


_TEMPMAIL_DOMAIN = "bltiwd.com"

async def _tm_get_domains() -> List[str]:
    def _extract(raw) -> List[str]:
        """Recursively extract domain name strings from any API response shape."""
        if isinstance(raw, str):
            return [raw] if raw else []
        if isinstance(raw, dict):
            # {"domains": [...]}  or  {"name": "bltiwd.com"}
            if "name" in raw:
                return [str(raw["name"])]
            for key in ("domains", "data", "list"):
                if key in raw:
                    return _extract(raw[key])
            return []
        if isinstance(raw, list):
            result = []
            for item in raw:
                result.extend(_extract(item))
            return result
        return []

    def _call():
        try:
            url = f"{_TEMPMAIL_API}/domains"
            req = _urllib_request.Request(url, headers={"User-Agent": "NeroxaShopBot/1.0"})
            with _urllib_request.urlopen(req, timeout=10) as res:
                data = json.loads(res.read())
                names = [d for d in _extract(data) if d]
                return names if names else [_TEMPMAIL_DOMAIN]
        except Exception as e:
            logger.error("TempMail domains error: %s", e)
        return [_TEMPMAIL_DOMAIN]
    return await _run(_call)

async def _tm_create_email(domain: str = _TEMPMAIL_DOMAIN) -> Tuple[Optional[str], Optional[str]]:
    def _call():
        r = _tempmail_request(f"{_TEMPMAIL_API}/email/new", {
            "min_name_length": 10,
            "max_name_length": 10,
            "domain": domain,
        })
        return r.get("email"), r.get("token")
    return await _run(_call)


async def _tm_get_inbox(email: str) -> List[dict]:
    def _call():
        url = f"{_TEMPMAIL_API}/email/{email}/messages"
        headers = {"accept": "application/json", "User-Agent": "NeroxaShopBot/1.0"}
        try:
            req = _urllib_request.Request(url, headers=headers)
            with _urllib_request.urlopen(req, timeout=12) as res:
                data = json.loads(res.read())
                return data if isinstance(data, list) else []
        except Exception as e:
            logger.error("TempMail inbox error: %s", e)
            return []
    return await _run(_call)


def _tm_strip_html(html: str) -> str:
    if not html:
        return ""
    text = re.sub(r"<br\s*/?>", "\n", html, flags=re.IGNORECASE)
    text = re.sub(r"<p[^>]*>", "\n", text, flags=re.IGNORECASE)
    text = re.sub(r"</p>", "\n", text, flags=re.IGNORECASE)
    text = re.sub(r"<[^>]+>", "", text)
    for ent, ch in [("&amp;","&"),("&lt;","<"),("&gt;",">"),("&quot;",'"'),("&nbsp;"," "),("&#39;","'")]:
        text = text.replace(ent, ch)
    return re.sub(r"\n{3,}", "\n\n", text).strip()


def _tm_extract_otp(body: str) -> Optional[str]:
    patterns = [
        r'\b(\d{4,8})\s+is your',
        r'[Cc]ode[:\s]+(\d{4,8})',
        r'OTP[:\s]*(\d{4,8})',
        r'[Vv]erification code[:\s]*(\d{4,8})',
        r'\b([A-Z0-9]{6,10})\b',
    ]
    for pat in patterns:
        m = re.search(pat, body[:1000])
        if m:
            return m.group(1)
    return None


def _tm_format_body(body_text: str, body_html: str) -> str:
    body = (body_text or "").strip() or _tm_strip_html(body_html or "")
    if not body:
        return "<i>No content</i>"
    otp = _tm_extract_otp(body)
    preview = body[:2000]
    result = f"<pre>{html_lib.escape(preview)}</pre>"
    if len(body) > 2000:
        result += "\n<i>… truncated</i>"
    if otp:
        result += f"\n\n🔑 <b>Code detected:</b> <code>{html_lib.escape(otp)}</code>"
    return result


def _tm_menu_kb() -> ReplyKeyboardMarkup:
    return _kb(
        [_TM_GEN],
        [_TM_INB, _TM_REF],
        [_TM_MY,  _TM_DEL],
        [_TM_DOM, _TM_RND],
        [BACK_BTN, HOME_BTN],
    )

def _tm_domain_kb(domains: List[str]) -> ReplyKeyboardMarkup:
    rows = [[_TM_RND]]                                         # Random at top
    rows += [domains[i:i+2] for i in range(0, min(len(domains), 16), 2)]
    rows.append([BACK_BTN])
    return _kb(*rows)


# ══════════════════════════════════════════════════════════════════
# ROUTER — Start & Main Menu
# ══════════════════════════════════════════════════════════════════

router_start = Router()


@router_start.message(Command("start"))
async def cmd_start(message: Message, state: FSMContext):
    await state.clear()
    if is_admin(message.from_user.id):
        await state.set_state(AdminFlow.menu)
        await message.answer(
            f"🔐 <b>Admin Panel</b>\n{_SEP}\nWelcome back, admin! 👋",
            reply_markup=admin_main_kb(),
        )
        return
    # Note: Middleware also checks force-join, but /start provides richer onboarding UX with inline join buttons
    passed, failed_channels = await check_force_join(message.bot, message.from_user.id)
    if not passed:
        try:
            bot_info = await message.bot.get_me()
            bot_username = bot_info.username
        except Exception:
            bot_username = None
        buttons = [
            [InlineKeyboardButton(
                text=f"📢 {_channel_display_name(ch)} — Join Now",
                url=_channel_to_url(ch),
            )]
            for ch in failed_channels
        ]
        if bot_username:
            buttons.append([InlineKeyboardButton(
                text="✅ I Joined — Check Again",
                url=f"https://t.me/{bot_username}?start=check",
            )])
        await message.answer(
            f"🔒 <b>Join Required</b>\n{_SEP}\n"
            f"Please join our channel(s) to access the shop.\n\n"
            f"After joining, press <b>✅ I Joined — Check Again</b> or send /start. ✅",
            reply_markup=InlineKeyboardMarkup(inline_keyboard=buttons),
        )
        return
    # ── Handle referral deep link ──
    args = message.text.split(maxsplit=1)
    ref_uid = None
    if len(args) > 1 and args[1].startswith("ref_"):
        try:
            ref_uid = int(args[1][4:])
        except (ValueError, IndexError):
            ref_uid = None
    if ref_uid and ref_uid != message.from_user.id:
        existing_user = await get_user(message.from_user.id)
        if existing_user is None:
            await create_or_update_user(
                message.from_user.id,
                message.from_user.username or "",
                message.from_user.full_name or "",
            )
            referrer = await get_user(ref_uid)
            if referrer:
                ref_info = await get_referral_info(message.from_user.id)
                if not ref_info.get("referred_by"):
                    await set_referrer(message.from_user.id, ref_uid)
    await send_main_menu(message, state)


@router_start.message(F.text == BTN_BALANCE)
async def show_balance(message: Message):
    user = await get_user(message.from_user.id) or {}
    await message.answer(fmt_balance_screen(user), reply_markup=main_menu_kb())


@router_start.message(F.text == BTN_REFERRAL)
async def show_referral(message: Message):
    uid = message.from_user.id
    ref_info = await get_referral_info(uid)
    try:
        bot_info = await message.bot.get_me()
        bot_username = bot_info.username
    except Exception:
        bot_username = "bot"
    link = f"https://t.me/{bot_username}?start=ref_{uid}"
    await message.answer(
        f"🔗 <b>Referral Program</b>\n{_SEP}\n"
        f"Share your link and earn Bonus Balance on every purchase your referrals make!\n"
        f"(Bonus can be used as discount on your purchases)\n\n"
        f"🔗 Your Link:\n<code>{link}</code>\n\n"
        f"👥 Total Referred: <b>{ref_info.get('referred_count', 0)}</b>\n"
        f"🎁 Bonus Earned: <b>${ref_info.get('total_earned', 0.0):.2f}</b>",
        reply_markup=main_menu_kb(),
    )


@router_start.message(F.text == BTN_GET_CODE)
async def get_code(message: Message, state: FSMContext):
    uid     = message.from_user.id
    session = _mail_sessions.get(uid)
    status  = f"📧 <code>{html_lib.escape(session['email'])}</code>" if session else "❌ No mail set"
    await state.set_state(UserFlow.get_code_menu)
    await message.answer(
        f"🔓 <b>Get Code — Mail Inbox</b>\n{_SEP}\n"
        f"Current mail: {status}\n\n"
        f"Use the buttons below to manage your mailbox.",
        reply_markup=get_code_menu_kb(),
    )


# ── Get Code sub-menu handlers ─────────────────────────────────────

@router_start.message(UserFlow.get_code_menu, F.text == BACK_BTN)
async def gc_back(message: Message, state: FSMContext):
    await send_main_menu(message, state)


@router_start.message(UserFlow.get_code_menu, F.text == BTN_GC_SET_MAIL)
async def gc_set_mail(message: Message, state: FSMContext):
    await state.set_state(UserFlow.get_code_set)
    await message.answer(
        f"📧 <b>Set Mail</b>\n{_SEP}\n"
        f"Send your account data in format:\n\n"
        f"<code>email|password|refresh_token|client_id</code>\n\n"
        f"Or just:\n<code>email|password</code>\n\n"
        f"(OAuth2 will be auto-fetched if not provided)\n\n"
        f"Press {BACK_BTN} to cancel.",
        reply_markup=input_kb(),
    )


@router_start.message(UserFlow.get_code_menu, F.text == BTN_GC_CHANGE)
async def gc_change_mail(message: Message, state: FSMContext):
    uid = message.from_user.id
    if uid in _mail_sessions:
        del _mail_sessions[uid]
    await state.set_state(UserFlow.get_code_set)
    await message.answer(
        f"✏️ <b>Change Mail</b>\n{_SEP}\n"
        f"Session cleared! Send new account data:\n\n"
        f"<code>email|password|refresh_token|client_id</code>\n\n"
        f"Press {BACK_BTN} to cancel.",
        reply_markup=input_kb(),
    )


@router_start.message(UserFlow.get_code_menu, F.text == BTN_GC_CODES)
async def gc_get_codes(message: Message, state: FSMContext):
    uid     = message.from_user.id
    session = _mail_sessions.get(uid)
    if not session:
        await message.answer(
            f"❌ <b>No mail set!</b>\n{_SEP}\nPress <b>📧 Set Mail</b> first.",
            reply_markup=get_code_menu_kb(),
        )
        return
    if not _session_has_oauth2(session):
        await message.answer(
            f"⚠️ <b>OAuth2 Not Set</b>\n{_SEP}\n"
            f"📧 <code>{html_lib.escape(session['email'])}</code>\n\n"
            f"Your mail was auto-set but OAuth2 token is missing.\n"
            f"Press <b>📧 Set Mail</b> and enter:\n"
            f"<code>email|password|refresh_token|client_id</code>",
            reply_markup=get_code_menu_kb(),
        )
        return
    wait = await message.answer(f"🔄 Fetching codes for <code>{html_lib.escape(session['email'])}</code>…")
    all_msgs: List[dict] = []
    code_data = await _dv_get_code(session["email"], session["refresh_token"], session["client_id"], "all")
    if code_data and code_data.get("status") and code_data.get("code"):
        svc = _dv_determine_service(code_data.get("content", ""), "", "")
        all_msgs.append({"code": code_data["code"], "service": svc,
                         "subject": code_data.get("content", ""), "sender": "",
                         "date": code_data.get("date", ""), "address": ""})
    msgs = await _dv_get_messages(session["email"], session["refresh_token"], session["client_id"])
    if msgs and msgs.get("status"):
        existing = {m["code"] for m in all_msgs}
        for c in _dv_extract_codes(msgs):
            if c["code"] not in existing:
                all_msgs.append(c)
                existing.add(c["code"])
    if not all_msgs:
        msgs2 = await _dv_graph_messages(session["email"], session["refresh_token"], session["client_id"])
        if msgs2 and msgs2.get("status"):
            all_msgs = _dv_extract_codes(msgs2)
    display = _dv_build_codes_display(all_msgs)
    await _safe_delete(wait)
    await message.answer(
        _truncate_msg(f"📧 <code>{html_lib.escape(session['email'])}</code>\n{display}"),
        reply_markup=get_code_menu_kb(),
    )


@router_start.message(UserFlow.get_code_menu, F.text == BTN_GC_INBOX)
async def gc_read_inbox(message: Message, state: FSMContext):
    uid     = message.from_user.id
    session = _mail_sessions.get(uid)
    if not session:
        await message.answer(f"❌ <b>No mail set!</b>\n{_SEP}\nPress <b>📧 Set Mail</b> first.", reply_markup=get_code_menu_kb())
        return
    if not _session_has_oauth2(session):
        await message.answer(
            f"⚠️ <b>OAuth2 Not Set</b>\n{_SEP}\n"
            f"📧 <code>{html_lib.escape(session['email'])}</code>\n\n"
            f"Press <b>📧 Set Mail</b> and enter full credentials:\n"
            f"<code>email|password|refresh_token|client_id</code>",
            reply_markup=get_code_menu_kb(),
        )
        return
    wait = await message.answer(f"📨 Reading <code>{html_lib.escape(session['email'])}</code>…")
    messages = await _dv_process_mailbox(session)
    display  = _dv_build_mailbox_display(messages)
    await _safe_delete(wait)
    await message.answer(
        _truncate_msg(
            f"📧 <code>{html_lib.escape(session['email'])}</code>\n"
            f"🔑 <code>{html_lib.escape(session['password'])}</code>\n{display}"
        ),
        reply_markup=get_code_menu_kb(),
    )


@router_start.message(UserFlow.get_code_menu, F.text == BTN_GC_REFRESH)
async def gc_refresh(message: Message, state: FSMContext):
    uid     = message.from_user.id
    session = _mail_sessions.get(uid)
    if not session:
        await message.answer(f"❌ <b>No mail set!</b>\n{_SEP}\nPress <b>📧 Set Mail</b> first.", reply_markup=get_code_menu_kb())
        return
    if not _session_has_oauth2(session):
        await message.answer(
            f"⚠️ <b>OAuth2 Not Set</b>\n{_SEP}\n"
            f"📧 <code>{html_lib.escape(session['email'])}</code>\n\n"
            f"Press <b>📧 Set Mail</b> and enter full credentials:\n"
            f"<code>email|password|refresh_token|client_id</code>",
            reply_markup=get_code_menu_kb(),
        )
        return
    wait = await message.answer(f"🔄 Refreshing <code>{html_lib.escape(session['email'])}</code>…")
    messages = await _dv_process_mailbox(session)
    display  = _dv_build_mailbox_display(messages)
    await _safe_delete(wait)
    await message.answer(
        _truncate_msg(f"✅ <b>Refreshed!</b>\n\n📧 <code>{html_lib.escape(session['email'])}</code>\n{display}"),
        reply_markup=get_code_menu_kb(),
    )


@router_start.message(UserFlow.get_code_menu, F.text == BTN_GC_FILTER)
async def gc_show_filter(message: Message, state: FSMContext):
    uid     = message.from_user.id
    session = _mail_sessions.get(uid)
    if not session:
        await message.answer(f"❌ <b>No mail set!</b>\n{_SEP}\nPress <b>📧 Set Mail</b> first.", reply_markup=get_code_menu_kb())
        return
    if not _session_has_oauth2(session):
        await message.answer(
            f"⚠️ <b>OAuth2 Not Set</b>\n{_SEP}\n"
            f"📧 <code>{html_lib.escape(session['email'])}</code>\n\n"
            f"Press <b>📧 Set Mail</b> and enter full credentials:\n"
            f"<code>email|password|refresh_token|client_id</code>",
            reply_markup=get_code_menu_kb(),
        )
        return
    await state.set_state(UserFlow.get_code_filter)
    await state.update_data(gc_email=session["email"])
    await message.answer(
        f"🔍 <b>Filter by Service</b>\n{_SEP}\n📧 <code>{html_lib.escape(session['email'])}</code>\n\nSelect a service:",
        reply_markup=get_code_filter_kb(),
    )


@router_start.message(UserFlow.get_code_set)
async def gc_receive_mail_data(message: Message, state: FSMContext):
    if message.text in (CANCEL_BTN, BACK_BTN, HOME_BTN):
        await state.set_state(UserFlow.get_code_menu)
        uid     = message.from_user.id
        session = _mail_sessions.get(uid)
        status  = f"📧 <code>{html_lib.escape(session['email'])}</code>" if session else "❌ No mail set"
        await message.answer(
            f"🔓 <b>Get Code — Mail Inbox</b>\n{_SEP}\nCurrent mail: {status}",
            reply_markup=get_code_menu_kb(),
        )
        return
    text = message.text or ""
    if "|" not in text:
        await message.answer("❌ Invalid format. Use: <code>email|password|refresh_token|client_id</code>", reply_markup=input_kb())
        return
    parsed = _dv_parse_input(text)
    if not parsed["email"]:
        await message.answer("❌ Invalid format. Use: <code>email|password</code>", reply_markup=input_kb())
        return
    wait = await message.answer(f"🔄 Processing <code>{html_lib.escape(parsed['email'])}</code>…")
    if not parsed["has_oauth2"]:
        oauth2 = await _dv_oauth2(parsed["email"], parsed["password"])
        if not oauth2:
            await _safe_delete(wait)
            await message.answer(
                f"❌ Failed to get OAuth2 for <code>{html_lib.escape(parsed['email'])}</code>\n\n"
                f"Check credentials or top up your Dongvan balance.\nTry again:",
                reply_markup=input_kb(),
            )
            return
        parsed["refresh_token"] = oauth2["refresh_token"]
        parsed["client_id"]     = oauth2["client_id"]
    uid = message.from_user.id
    _mail_sessions[uid] = {
        "email": parsed["email"], "password": parsed["password"],
        "refresh_token": parsed["refresh_token"], "client_id": parsed["client_id"],
    }
    msgs    = await _dv_process_mailbox(_mail_sessions[uid])
    display = _dv_build_mailbox_display(msgs)
    cid_raw = parsed["client_id"] or ""
    cid_p   = cid_raw[:20] + "…" if len(cid_raw) > 20 else cid_raw
    await _safe_delete(wait)
    await state.set_state(UserFlow.get_code_menu)
    await message.answer(
        _truncate_msg(
            f"✅ <b>Mail Set Successfully!</b>\n{_SEP}\n"
            f"📧 <code>{html_lib.escape(parsed['email'])}</code>\n"
            f"🆔 <code>{html_lib.escape(cid_p)}</code>\n\n{display}"
        ),
        reply_markup=get_code_menu_kb(),
    )


@router_start.message(UserFlow.get_code_filter)
async def gc_filter_selected(message: Message, state: FSMContext):
    if message.text in (BACK_BTN, HOME_BTN):
        await state.set_state(UserFlow.get_code_menu)
        uid     = message.from_user.id
        session = _mail_sessions.get(uid)
        status  = f"📧 <code>{html_lib.escape(session['email'])}</code>" if session else "❌ No mail set"
        await message.answer(f"🔓 <b>Get Code — Mail Inbox</b>\n{_SEP}\nCurrent mail: {status}", reply_markup=get_code_menu_kb())
        return
    filter_svc = message.text.strip().upper()
    if filter_svc not in _DV_SERVICE_KEYWORDS:
        await message.answer("❌ Select a service from the keyboard.", reply_markup=get_code_filter_kb())
        return
    uid     = message.from_user.id
    session = _mail_sessions.get(uid)
    if not session:
        await state.set_state(UserFlow.get_code_menu)
        await message.answer("❌ Session expired. Please set mail again.", reply_markup=get_code_menu_kb())
        return
    wait     = await message.answer(f"🔍 Filtering <b>{filter_svc}</b> in <code>{html_lib.escape(session['email'])}</code>…")
    messages = await _dv_process_mailbox(session)
    display  = _dv_build_mailbox_display(messages, filter_svc)
    await _safe_delete(wait)
    await message.answer(
        _truncate_msg(f"📧 <code>{html_lib.escape(session['email'])}</code>\n{display}"),
        reply_markup=get_code_filter_kb(),
    )


async def _show_2fa(message: Message) -> None:
    """Shared logic for BTN_GET_2FA button and /get2fa command."""
    uid    = message.from_user.id
    secret = await get_totp_secret(uid)
    if not secret:
        await message.answer(
            f"🔑 <b>2FA Code Generator</b>\n{_SEP}\n"
            f"You haven't set a secret key yet.\n\n"
            f"Press <b>🔑 Set My Key</b> below to add your TOTP secret.",
            reply_markup=tfa_nokey_kb(),
        )
        return
    try:
        code = pyotp.TOTP(secret).now()
    except Exception:
        await message.answer(
            f"⚠️ <b>Invalid Key</b>\n{_SEP}\n"
            f"Your saved key appears invalid. Please reset it.",
            reply_markup=tfa_nokey_kb(),
        )
        return
    await message.answer(_fmt_totp(code), reply_markup=tfa_inline_kb())


@router_start.message(F.text == BTN_GET_2FA)
async def get_2fa_btn(message: Message, state: FSMContext):
    await _show_2fa(message)


@router_start.message(Command("get2fa"))
async def get_2fa_cmd(message: Message, state: FSMContext):
    await _show_2fa(message)


@router_start.message(Command("setkey"))
async def cmd_setkey(message: Message, state: FSMContext):
    await state.set_state(UserFlow.set_totp_key)
    await message.answer(
        f"🔑 <b>Set 2FA Secret Key</b>\n{_SEP}\n"
        f"Enter your <b>Base32 TOTP secret key</b>:\n\n"
        f"<i>Example: JBSWY3DPEHPK3PXP</i>\n\n"
        f"⚠️ Keep this private — the bot only uses it to generate codes for you.",
        reply_markup=input_kb(),
    )


@router_start.message(Command("timer"))
async def cmd_timer(message: Message):
    secs = _totp_secs()
    bar  = _totp_bar(secs)
    await message.answer(
        f"⏱ <b>2FA Timer</b>\n{_SEP}\n"
        f"Current code expires in: <b>{secs} seconds</b>\n"
        f"[{bar}]"
    )


@router_start.message(UserFlow.set_totp_key)
async def receive_totp_key(message: Message, state: FSMContext):
    if message.text in (CANCEL_BTN, BACK_BTN, HOME_BTN):
        await state.clear()
        await send_main_menu(message, state)
        return
    key = message.text.strip().upper().replace(" ", "")
    if not _valid_totp_key(key):
        await message.answer(
            f"❌ <b>Invalid Key</b>\n{_SEP}\n"
            f"That doesn't look like a valid Base32 TOTP secret.\n\n"
            f"Make sure it's the key from your authenticator app (letters A–Z and digits 2–7).\n"
            f"Try again or press ❌ Cancel.",
            reply_markup=input_kb(),
        )
        return
    await set_totp_secret(message.from_user.id, key)
    code = pyotp.TOTP(key).now()
    await state.clear()
    await message.answer(
        f"✅ <b>Key Saved!</b>\n{_SEP}\n"
        f"Your 2FA key has been saved securely.\n\n"
        + _fmt_totp(code),
        reply_markup=tfa_inline_kb(),
    )
    await message.answer("⬇️ Use the buttons below or press 🏠 Home.", reply_markup=main_menu_kb())


# ── 2FA Callback Queries ──────────────────────────────────────────

@router_start.callback_query(F.data == "tfa_refresh")
async def tfa_cb_refresh(call: CallbackQuery):
    uid    = call.from_user.id
    secret = await get_totp_secret(uid)
    if not secret:
        await call.answer("❌ No key set. Press 🔑 Set My Key.", show_alert=True)
        return
    try:
        code = pyotp.TOTP(secret).now()
    except Exception:
        await call.answer("⚠️ Invalid key. Please reset.", show_alert=True)
        return
    try:
        await call.message.edit_text(_fmt_totp(code), reply_markup=tfa_inline_kb())
    except Exception:
        pass
    await call.answer("✅ Refreshed!")


@router_start.callback_query(F.data == "tfa_timer")
async def tfa_cb_timer(call: CallbackQuery):
    secs = _totp_secs()
    bar  = _totp_bar(secs)
    await call.answer(f"⏱ {secs}s remaining\n[{bar}]", show_alert=True)


@router_start.callback_query(F.data == "tfa_setkey")
async def tfa_cb_setkey(call: CallbackQuery, state: FSMContext):
    await state.set_state(UserFlow.set_totp_key)
    await call.message.answer(
        f"🔑 <b>Set 2FA Secret Key</b>\n{_SEP}\n"
        f"Enter your <b>Base32 TOTP secret key</b>:\n\n"
        f"<i>Example: JBSWY3DPEHPK3PXP</i>\n\n"
        f"⚠️ Keep this private — the bot only uses it to generate codes for you.",
        reply_markup=input_kb(),
    )
    await call.answer()


@router_start.message(F.text == BTN_SUPPORT)
async def show_support(message: Message):
    settings = await get_settings()
    support = (settings.get("support_username") or SUPPORT_USERNAME).strip()
    support_url = _channel_to_url(support)
    display = _channel_display_name(support)
    await message.answer(
        f"🆘 <b>Support</b>\n{_SEP}\n"
        f"Need help? Our support team is ready!\n\n"
        f"Click the button below to open a chat with us. 👇",
        reply_markup=InlineKeyboardMarkup(inline_keyboard=[
            [InlineKeyboardButton(text=f"🆘 Contact {display}", url=support_url)]
        ]),
    )


# ══════════════════════════════════════════════════════════════════
# AI SUPPORT HANDLER
# ══════════════════════════════════════════════════════════════════

_AI_FAQ = {
    BTN_AI_HOW_BUY: (
        "\u2753 <b>How to Buy</b>\n" + _SEP + "\n"
        "1. Choose a category from the main menu (Get Mail, Buy VPN, Buy Proxy)\n"
        "2. Select the product you want\n"
        "3. Enter the quantity\n"
        "4. Optionally apply a coupon code\n"
        "5. Press \u2705 Confirm Purchase\n"
        "6. Your accounts will be delivered instantly!\n\n"
        "Make sure you have enough balance before purchasing."
    ),
    BTN_AI_HOW_DEPOSIT: (
        "\U0001F4B3 <b>How to Deposit</b>\n" + _SEP + "\n"
        "1. Press \U0001F4B3 Deposit from the main menu\n"
        "2. Choose a payment method (bKash, Nagad, or Binance)\n"
        "3. Send the amount to the displayed number/UID\n"
        "4. Enter the amount you sent\n"
        "5. Provide your Transaction ID\n"
        "6. Wait for admin approval (usually 5-30 minutes)\n\n"
        "Your balance will be credited automatically once approved."
    ),
    BTN_AI_STOCK: (
        "\U0001F4E6 <b>Check Stock</b>\n" + _SEP + "\n"
        "To check available stock:\n"
        "1. Go to \U0001F4E8 Get Mail, \U0001F6E1 Buy VPN, or \U0001F30D Buy Proxy\n"
        "2. Select any product to see current stock count\n"
        "3. Green dot = In Stock, Red dot = Out of Stock\n\n"
        "Stock is updated in real-time. If a product is out of stock, "
        "check back later or contact support."
    ),
    BTN_AI_REFERRAL: (
        "\U0001F517 <b>Referral Info</b>\n" + _SEP + "\n"
        "Share your referral link with friends!\n"
        "1. Press \U0001F517 Referral from the main menu\n"
        "2. Copy your unique referral link\n"
        "3. Share it with friends\n"
        "4. When they join and make purchases, you earn bonus balance!\n\n"
        "The more friends you invite, the more you earn."
    ),
    BTN_AI_ORDERS: (
        "\U0001F4CB <b>My Orders</b>\n" + _SEP + "\n"
        "To view your order history:\n"
        "1. Press \U0001F4CB Order History from the main menu\n"
        "2. You'll see all your Mail, VPN, and Proxy orders\n"
        "3. Each order shows the product, quantity, and price\n\n"
        "For VPN/Proxy orders, status indicators:\n"
        "\u2705 = Delivered  |  \u23F3 = Pending  |  \u274C = Cancelled"
    ),
    BTN_AI_2FA: (
        "\U0001F510 <b>2FA Help</b>\n" + _SEP + "\n"
        "Two-Factor Authentication (2FA) codes:\n"
        "1. Press \U0001F510 Get 2FA from the main menu\n"
        "2. Enter or paste your TOTP secret key\n"
        "3. The bot generates a 6-digit code valid for 30 seconds\n\n"
        "This works like Google Authenticator but directly in Telegram. "
        "Keep your secret keys safe!"
    ),
    BTN_AI_MAIL: (
        "\U0001F4E8 <b>Mail Codes</b>\n" + _SEP + "\n"
        "To receive mail verification codes:\n"
        "1. Press \U0001F4F2 Get Code from the main menu\n"
        "2. Set your email address first (\U0001F4E7 Set Mail)\n"
        "3. Use \U0001F4EC Get Codes to check for new codes\n"
        "4. Use \U0001F4E5 Read Inbox to see all messages\n"
        "5. Use \U0001F3AF Filter Mail to find specific emails\n\n"
        "Your mailbox refreshes automatically. You can also use "
        "\U0001F559 Temp Mail for disposable addresses."
    ),
}

_AI_KEYWORDS = {
    "buy": "buy", "purchase": "buy", "order": "buy", "how to buy": "buy",
    "deposit": "deposit", "payment": "deposit", "bkash": "deposit",
    "nagad": "deposit", "binance": "deposit", "pay": "deposit", "add money": "deposit",
    "stock": "stock", "available": "stock", "out of stock": "stock",
    "referral": "referral", "invite": "referral", "link": "referral", "refer": "referral",
    "vip": "vip", "badge": "vip", "level": "vip", "tier": "vip",
    "2fa": "2fa", "totp": "2fa", "authenticator": "2fa", "two factor": "2fa",
    "mail": "mail", "code": "mail", "inbox": "mail", "email": "mail",
    "vpn": "vpn", "proxy": "proxy",
}

_AI_KEYWORD_ANSWERS = {
    "buy": (
        "\U0001F6D2 <b>Buying Process</b>\n" + _SEP + "\n"
        "Select a category (Get Mail, Buy VPN, Buy Proxy) from the main menu, "
        "pick a product, enter quantity, and confirm. Make sure your balance is sufficient!"
    ),
    "deposit": (
        "\U0001F4B3 <b>Deposit Info</b>\n" + _SEP + "\n"
        "Press Deposit, choose bKash/Nagad/Binance, send money to the displayed number, "
        "then enter the amount and transaction ID. Admin will approve within 5-30 minutes."
    ),
    "stock": (
        "\U0001F4E6 <b>Stock Info</b>\n" + _SEP + "\n"
        "Go to any product category and select a product to see real-time stock. "
        "Green dot means available, red dot means out of stock."
    ),
    "referral": (
        "\U0001F517 <b>Referral System</b>\n" + _SEP + "\n"
        "Press Referral in the main menu to get your unique link. "
        "Share it with friends - when they join and buy, you earn bonus balance!"
    ),
    "vip": (
        "\U0001F451 <b>VIP System</b>\n" + _SEP + "\n"
        "Your VIP tier is based on total spending. Higher tiers unlock badges "
        "and exclusive benefits. Check your balance screen to see your progress!"
    ),
    "2fa": (
        "\U0001F510 <b>2FA Codes</b>\n" + _SEP + "\n"
        "Press Get 2FA, enter your TOTP secret key, and get a 6-digit code. "
        "Works just like Google Authenticator!"
    ),
    "mail": (
        "\U0001F4E8 <b>Mail System</b>\n" + _SEP + "\n"
        "Use Get Code to set your email, then check codes and inbox. "
        "Use Temp Mail for disposable email addresses."
    ),
    "vpn": (
        "\U0001F6E1 <b>VPN Service</b>\n" + _SEP + "\n"
        "Press Buy VPN, select a VPN product, choose duration (1/7/30/90 days), "
        "and confirm your purchase. Delivery is handled by our team."
    ),
    "proxy": (
        "\U0001F30D <b>Proxy Service</b>\n" + _SEP + "\n"
        "Press Buy Proxy, select a proxy product, choose data package, "
        "and confirm. Our team will deliver your proxy details."
    ),
}


@router_start.message(F.text == BTN_AI_SUPPORT)
async def show_ai_support(message: Message, state: FSMContext):
    await state.set_state(UserFlow.ai_support)
    await message.answer(
        f"\U0001F916 <b>AI Support</b>\n{_SEP}\n"
        f"Choose a topic below to get instant help,\n"
        f"or press \U0001F4AC Ask a Question to type your query.",
        reply_markup=ai_support_kb(),
    )


@router_start.message(UserFlow.ai_support, F.text.in_(_AI_FAQ.keys()))
async def ai_faq_answer(message: Message, state: FSMContext):
    answer = _AI_FAQ[message.text]
    await message.answer(answer, reply_markup=ai_support_kb())


@router_start.message(UserFlow.ai_support, F.text == BTN_AI_ASK)
async def ai_ask_question(message: Message, state: FSMContext):
    await state.set_state(UserFlow.ai_support_ask)
    await message.answer(
        f"\U0001F4AC <b>Ask a Question</b>\n{_SEP}\n"
        f"Type your question below and I'll try to help!\n"
        f"Examples: \"How do I buy?\", \"How to deposit?\", \"VIP info\"",
        reply_markup=_kb([BACK_BTN, HOME_BTN]),
    )


@router_start.message(UserFlow.ai_support, F.text == BACK_BTN)
async def ai_support_back(message: Message, state: FSMContext):
    await send_main_menu(message, state)


@router_start.message(UserFlow.ai_support_ask, F.text == BACK_BTN)
async def ai_support_ask_back(message: Message, state: FSMContext):
    await state.set_state(UserFlow.ai_support)
    await message.answer(
        f"\U0001F916 <b>AI Support</b>\n{_SEP}\n"
        f"Choose a topic below to get instant help,\n"
        f"or press \U0001F4AC Ask a Question to type your query.",
        reply_markup=ai_support_kb(),
    )


@router_start.message(UserFlow.ai_support_ask)
async def ai_free_question(message: Message, state: FSMContext):
    if not message.text:
        return
    query = message.text.lower().strip()
    # Keyword matching (word-boundary check to avoid substring false positives)
    matched_topic = None
    for keyword, topic in _AI_KEYWORDS.items():
        if re.search(rf"\b{re.escape(keyword)}\b", query):
            matched_topic = topic
            break
    if matched_topic and matched_topic in _AI_KEYWORD_ANSWERS:
        await message.answer(_AI_KEYWORD_ANSWERS[matched_topic], reply_markup=_kb([BACK_BTN, HOME_BTN]))
    else:
        await message.answer(
            f"\U0001F914 <b>No Match Found</b>\n{_SEP}\n"
            f"Sorry, I couldn't find an answer for your question.\n"
            f"Try asking differently or contact our support team! \U0001F447",
            reply_markup=_kb([BACK_BTN, HOME_BTN]),
        )


@router_start.message(F.text == BTN_HISTORY)
async def order_history(message: Message):
    uid = message.from_user.id
    mail_o, vpn_o, proxy_o = await asyncio.gather(
        get_user_orders(uid), get_user_vpn_orders(uid), get_user_proxy_orders(uid),
    )
    await message.answer(fmt_order_history(mail_o, vpn_o, proxy_o), reply_markup=main_menu_kb())
    deposits = await get_user_deposits(uid)
    if deposits:
        await message.answer(fmt_deposit_history(deposits))


# ══════════════════════════════════════════════════════════════════
# HELPER — Low Stock Alert (shared by mail and proxy flows)
# ══════════════════════════════════════════════════════════════════

async def _check_low_stock_alert(bot: Bot, pid: str, product_name: str) -> None:
    """Send low stock or out-of-stock alert to all admins."""
    try:
        remaining = await get_stock_count(pid)
        settings  = await get_settings()
        threshold = int(settings.get("low_stock_threshold", 5))
        if remaining == 0:
            for adm in ADMIN_IDS:
                try:
                    await bot.send_message(
                        adm,
                        f"\U0001f6a8 <b>Out of Stock!</b>\n{_SEP}\n"
                        f"\U0001f4e6 Product: <b>{product_name}</b>\n"
                        f"Stock is now <b>0</b>. Product hidden from users.",
                    )
                except Exception:
                    pass
        elif 0 < remaining <= threshold:
            for adm in ADMIN_IDS:
                try:
                    await bot.send_message(
                        adm,
                        f"\u26a0\ufe0f <b>Low Stock Alert!</b>\n{_SEP}\n"
                        f"\U0001f4e6 Product: <b>{product_name}</b>\n"
                        f"\U0001f522 Remaining stock: <b>{remaining}</b> item(s)\n"
                        f"Please restock soon!",
                    )
                except Exception:
                    pass
    except Exception as _se:
        logger.warning("Low stock alert failed: %s", _se)


# ══════════════════════════════════════════════════════════════════
# ROUTER — User Panel Enhancements
# ══════════════════════════════════════════════════════════════════

router_user_panel = Router()


# ── VIP Badge display helper ──────────────────────────────────────

def _vip_badge_line(total_spent: float) -> str:
    tier_name, tier_emoji = get_vip_tier(total_spent)
    if not tier_name:
        return ""
    return f"{tier_emoji} VIP: <b>{tier_name}</b>"


# ── More Menu ──────────────────────────────────────────────────────

@router_user_panel.message(F.text == BTN_MORE)
async def more_menu_handler(message: Message, state: FSMContext):
    await state.set_state(UserFlow.more_menu)
    await message.answer("\U0001f4c2 <b>More Options</b>", reply_markup=more_menu_kb())


@router_user_panel.message(UserFlow.more_menu, F.text == BACK_BTN)
async def more_menu_back(message: Message, state: FSMContext):
    await state.clear()
    await send_main_menu(message, state)


# ── Trending Services ─────────────────────────────────────────────

@router_user_panel.message(F.text == BTN_TRENDING)
async def show_trending(message: Message):
    orders_data, vpn_data, proxy_data, products = await asyncio.gather(
        db_get("orders"), db_get("vpn_orders"), db_get("proxy_orders"), get_all_products(),
    )
    now = int(time.time())
    week_ago = now - 7 * 86400
    purchase_counts: Dict[str, int] = defaultdict(int)
    for v in (orders_data or {}).values():
        if v.get("created_at", 0) >= week_ago:
            pid = v.get("product_id", "")
            if pid:
                purchase_counts[pid] += v.get("qty", 1)
    for v in (vpn_data or {}).values():
        if v.get("created_at", 0) >= week_ago and v.get("status") != "cancelled":
            pid = v.get("product_id", "")
            if pid:
                purchase_counts[pid] += 1
    for v in (proxy_data or {}).values():
        if v.get("created_at", 0) >= week_ago and v.get("status") != "cancelled":
            pid = v.get("product_id", "")
            if pid:
                purchase_counts[pid] += 1
    if not purchase_counts:
        await message.answer(
            f"🔥 <b>Trending Services</b>\n{_SEP}\n"
            f"No trending data yet this week.\nCheck back soon!",
            reply_markup=main_menu_kb(),
        )
        return
    top5 = sorted(purchase_counts.items(), key=lambda x: x[1], reverse=True)[:5]
    lines = [f"🔥 <b>Trending Services</b> (Last 7 Days)\n{_SEP}"]
    for rank, (pid, count) in enumerate(top5, 1):
        p = products.get(pid, {})
        name = p.get("name", "Unknown")
        emoji = p.get("emoji", "📦")
        status_dot = "🟢" if p.get("status", "online") == "online" else "🔴"
        price = p.get("price", 0)
        lines.append(f"\n{rank}. {status_dot} {emoji} <b>{name}</b>")
        lines.append(f"   💰 ${price:.2f} | 🔥 {count} purchases")
    await message.answer("\n".join(lines), reply_markup=main_menu_kb())


# ── Favorites ─────────────────────────────────────────────────────

@router_user_panel.message(F.text == BTN_FAVORITES)
async def show_favorites(message: Message, state: FSMContext):
    uid = message.from_user.id
    fav_pids = await get_user_favorites(uid)
    if not fav_pids:
        await message.answer(
            f"\u2B50 <b>My Favorites</b>\n{_SEP}\n"
            f"You have no favorite products yet.\n\n"
            f"Browse products to add favorites!",
            reply_markup=main_menu_kb(),
        )
        return
    products = await get_all_products()
    lines = [f"\u2B50 <b>My Favorites</b>\n{_SEP}"]
    found = False
    for pid in fav_pids:
        p = products.get(pid)
        if not p:
            continue
        found = True
        status_dot = "\U0001F7E2" if p.get("status", "online") == "online" else "\U0001F534"
        stock = p.get("stock_count", 0)
        emoji = p.get("emoji", "\U0001F4E6")
        lines.append(f"\n{status_dot} {emoji} <b>{p['name']}</b>")
        lines.append(f"   \U0001F4B0 ${p['price']:.2f} | \U0001F4E6 {stock} in stock")
    if not found:
        await message.answer(
            f"\u2B50 <b>My Favorites</b>\n{_SEP}\n"
            f"Your favorited products are no longer available.",
            reply_markup=main_menu_kb(),
        )
        return
    await message.answer("\n".join(lines), reply_markup=main_menu_kb())


# ── Advanced Service Search ───────────────────────────────────────

@router_user_panel.message(F.text == BTN_SEARCH)
async def search_start(message: Message, state: FSMContext):
    await state.set_state(UserFlow.search_services)
    await message.answer(
        f"🔍 <b>Search Services</b>\n{_SEP}\n"
        f"Type a product name to search.\n"
        f"Example: <i>Gmail, VPN, Proxy</i>\n\n"
        f"Press {HOME_BTN} to cancel.",
        reply_markup=input_kb(),
    )


@router_user_panel.message(UserFlow.search_services)
async def search_query(message: Message, state: FSMContext):
    if message.text in (CANCEL_BTN, BACK_BTN, HOME_BTN):
        await state.clear()
        await send_main_menu(message, state)
        return
    query = message.text.strip().lower()
    if not query or len(query) < 2:
        await message.answer("Please enter at least 2 characters to search.", reply_markup=input_kb())
        return
    products = await get_all_products()
    results = []
    for pid, p in products.items():
        if p.get("hidden"):
            continue
        name = p.get("name", "").lower()
        desc = p.get("description", "").lower()
        if query in name or query in desc:
            results.append((pid, p))
    if not results:
        await message.answer(
            f"🔍 No results for '<b>{html_lib.escape(query)}</b>'\n\nTry a different keyword.",
            reply_markup=input_kb(),
        )
        return
    lines = [f"\U0001F50D <b>Search Results</b> for '{html_lib.escape(query)}'\n{_SEP}"]
    for pid, p in results[:10]:
        status_dot = "\U0001F7E2" if p.get("status", "online") == "online" else "\U0001F534"
        stock = p.get("stock_count", 0)
        emoji = p.get("emoji", "\U0001F4E6")
        lines.append(f"\n{status_dot} {emoji} <b>{p['name']}</b>")
        lines.append(f"   \U0001F4B0 ${p['price']:.2f} | \U0001F4E6 {stock} in stock")
    await message.answer("\n".join(lines), reply_markup=main_menu_kb())


# ── Live User Statistics Dashboard ────────────────────────────────

@router_user_panel.message(F.text == BTN_MY_STATS)
async def show_user_stats(message: Message):
    uid = message.from_user.id
    user = await get_user(uid) or {}
    total_spent = user.get("total_spent", 0)
    order_count = user.get("order_count", 0)
    balance = user.get("balance", 0)
    # VIP info
    tier_name, tier_emoji = get_vip_tier(total_spent)
    next_tier, current, next_threshold = get_vip_progress(total_spent)
    # VIP progress bar
    if next_tier == "Max":
        vip_bar = "▓▓▓▓▓▓▓▓▓▓ MAX"
    else:
        pct = min(current / next_threshold, 1.0) if next_threshold > 0 else 0
        filled = round(pct * 10)
        vip_bar = "▓" * filled + "░" * (10 - filled) + f" ${current:.0f}/${next_threshold:.0f}"
    vip_line = f"{tier_emoji} <b>{tier_name}</b>" if tier_name else "No VIP tier yet"
    # Weekly spending
    orders_data, vpn_data, proxy_data = await asyncio.gather(
        db_get("orders"), db_get("vpn_orders"), db_get("proxy_orders"),
    )
    now = int(time.time())
    this_week = now - 7 * 86400
    last_week_start = now - 14 * 86400
    week_spend = 0.0
    last_week_spend = 0.0
    for v in (orders_data or {}).values():
        if v.get("user_id") != uid:
            continue
        ts = v.get("created_at", 0)
        amt = v.get("total_price", 0)
        if ts >= this_week:
            week_spend += amt
        elif ts >= last_week_start:
            last_week_spend += amt
    for src in (vpn_data or {}, proxy_data or {}):
        for v in src.values():
            if v.get("user_id") != uid or v.get("status") == "cancelled":
                continue
            ts = v.get("created_at", 0)
            amt = v.get("price", 0)
            if ts >= this_week:
                week_spend += amt
            elif ts >= last_week_start:
                last_week_spend += amt
    if last_week_spend > 0:
        trend = ((week_spend - last_week_spend) / last_week_spend) * 100
        trend_str = f"{'📈' if trend >= 0 else '📉'} {trend:+.0f}% vs last week"
    else:
        trend_str = "📊 No comparison data"
    # Favorite category
    cat_counts: Dict[str, int] = defaultdict(int)
    for v in (orders_data or {}).values():
        if v.get("user_id") == uid:
            cat_counts["mail"] += 1
    for v in (vpn_data or {}).values():
        if v.get("user_id") == uid:
            cat_counts["vpn"] += 1
    for v in (proxy_data or {}).values():
        if v.get("user_id") == uid:
            cat_counts["proxy"] += 1
    fav_cat = max(cat_counts, key=cat_counts.get) if cat_counts else "None"
    cat_emojis = {"mail": "📨", "vpn": "🛡", "proxy": "🌍", "None": "❓"}
    text = (
        f"📊 <b>My Statistics</b>\n{_SEP}\n\n"
        f"👤 <b>{user.get('full_name', 'User')}</b>\n"
        f"💎 VIP Status: {vip_line}\n"
        f"[{vip_bar}]\n"
        f"{_LINE}\n"
        f"💵 Balance: <b>${balance:.2f}</b>\n"
        f"💸 Total Spent: <b>${total_spent:.2f}</b>\n"
        f"🛍 Total Orders: <b>{order_count}</b>\n"
        f"{_LINE}\n"
        f"📅 This Week: <b>${week_spend:.2f}</b>\n"
        f"{trend_str}\n"
        f"{_LINE}\n"
        f"{cat_emojis.get(fav_cat, '📦')} Favorite Category: <b>{fav_cat.title()}</b>\n"
        f"📅 Member Since: {_dt(user.get('joined_at', int(time.time())))}"
    )
    await message.answer(text, reply_markup=main_menu_kb())


# ══════════════════════════════════════════════════════════════════
# ROUTER — Mail (auto-delivery)
# ══════════════════════════════════════════════════════════════════

router_mail = Router()


@router_mail.message(F.text == BTN_GET_MAIL)
async def mail_start(message: Message, state: FSMContext):
    products = await get_all_products()
    dm = {
        f"{'🟢' if p.get('status','online')=='online' else '🔴'} {p.get('emoji','📮')} {p['name']}  ·  ${p['price']:.2f}  ·  {p.get('stock_count',0)} left": pid
        for pid, p in products.items()
        if p.get("category", "mail") == "mail" and not p.get("hidden")
    }
    if not dm:
        await message.answer(
            f"📮 <b>Mail Products</b>\n{_SEP}\n❌ No mail products available right now.\nCheck back later!",
            reply_markup=main_menu_kb(),
        )
        return
    await state.update_data(dm=dm)
    await state.set_state(UserFlow.mail_product)
    await message.answer(fmt_product_list_header("📮 Mail Products", len(dm)), reply_markup=products_kb(dm))


@router_mail.message(UserFlow.mail_product)
async def mail_product_selected(message: Message, state: FSMContext):
    if message.text == BACK_BTN:
        await send_main_menu(message, state)
        return
    data = await state.get_data()
    dm   = data.get("dm", {})
    pid  = dm.get(message.text)
    if not pid:
        await message.answer("❌ Please select a product from the keyboard.", reply_markup=products_kb(dm))
        return
    product = await get_product(pid)
    if not product:
        await message.answer("❌ Product not found.")
        return
    stock = await get_stock_count(pid)
    if stock == 0:
        emoji = product.get("emoji", "\U0001F4E6")
        await message.answer(
            f"\u274C <b>Out of Stock</b>\n{_SEP}\n"
            f"{emoji} {product['name']} is currently out of stock.\n"
            f"Please check back later or browse other products.",
            reply_markup=main_menu_kb(),
        )
        await state.clear()
        return
    await state.update_data(pid=pid, product=product, stock=stock)
    await state.set_state(UserFlow.mail_qty)
    await message.answer(fmt_product_detail(product, stock), reply_markup=input_kb())


@router_mail.message(UserFlow.mail_qty)
async def mail_qty_entered(message: Message, state: FSMContext):
    if message.text in (CANCEL_BTN, BACK_BTN):
        data = await state.get_data()
        dm = data.get("dm", {})
        await state.set_data({"dm": dm})
        await state.set_state(UserFlow.mail_product)
        await message.answer(
            fmt_product_list_header("📮 Mail Products", len(dm)),
            reply_markup=products_kb(dm),
        )
        return
    data = await state.get_data()
    qty, err = validate_quantity(message.text, data.get("stock", 0))
    if err:
        await message.answer(err)
        return
    pid     = data.get("pid")
    if not pid:
        await send_main_menu(message, state)
        return
    product = await get_product(pid) or data.get("product", {})
    await state.update_data(product=product)
    total   = round(qty * product.get("price", 0), 4)
    user    = await get_user(message.from_user.id) or {}
    balance = user.get("balance", 0)
    await state.update_data(qty=qty, total=total, coupon_id=None, discount_pct=0)
    await state.set_state(UserFlow.mail_coupon)
    await message.answer(fmt_order_summary(product, qty, total, balance), reply_markup=confirm_kb())


@router_mail.message(UserFlow.mail_coupon)
async def mail_coupon_or_confirm(message: Message, state: FSMContext):
    if message.text in (BACK_BTN, CANCEL_BTN):
        data = await state.get_data()
        dm = data.get("dm", {})
        pid = data.get("pid")
        product = data.get("product", {})
        stock = data.get("stock", 0)
        await state.set_data({"dm": dm, "pid": pid, "product": product, "stock": stock})
        await state.set_state(UserFlow.mail_qty)
        await message.answer(fmt_product_detail(product, stock), reply_markup=input_kb())
        return
    if message.text == BTN_COUPON:
        await state.set_state(UserFlow.mail_coupon_input)
        await message.answer(
            f"🎟 <b>Apply Coupon</b>\n{_SEP}\nType your coupon code below:",
            reply_markup=input_kb(),
        )
        return
    if message.text != BTN_CONFIRM:
        await message.answer(
            "❌ Please press <b>✅ Confirm Purchase</b> to proceed, or <b>🎟 Apply Coupon</b> to enter a coupon code.",
            reply_markup=confirm_kb(),
        )
        return
    data    = await state.get_data()
    product = data.get("product", {})
    qty     = data.get("qty", 1)
    total   = data.get("total", 0)
    await state.set_state(UserFlow.mail_confirm)
    user    = await get_user(message.from_user.id) or {}
    balance = user.get("balance", 0)
    await message.answer(
        fmt_order_summary(product, qty, total, balance, data.get("discount_pct", 0)) +
        f"\n\n⚠️ This will deduct <b>${total:.2f}</b> from your balance.",
        reply_markup=confirm_final_kb(),
    )


@router_mail.message(UserFlow.mail_coupon_input)
async def mail_coupon_code_entered(message: Message, state: FSMContext):
    if message.text in (CANCEL_BTN, BACK_BTN, HOME_BTN):
        data    = await state.get_data()
        dm      = data.get("dm", {})
        pid     = data.get("pid")
        product = data.get("product", {})
        stock   = data.get("stock", 0)
        qty     = data.get("qty", 1)
        total   = data.get("total", 0)
        user    = await get_user(message.from_user.id) or {}
        balance = user.get("balance", 0)
        await state.set_data({"dm": dm, "pid": pid, "product": product, "stock": stock,
                               "qty": qty, "total": total, "coupon_id": None, "discount_pct": 0})
        await state.set_state(UserFlow.mail_coupon)
        await message.answer(
            fmt_order_summary(product, qty, total, balance, 0),
            reply_markup=confirm_kb(),
        )
        return
    code   = message.text.strip()
    coupon = await get_coupon(code)
    if not coupon:
        await message.answer(
            f"❌ Invalid coupon code: <code>{code}</code>\n\nTry again or press ❌ Cancel to go back.",
            reply_markup=input_kb(),
        )
        return
    valid, err = validate_coupon(coupon)
    if not valid:
        await message.answer(err + "\n\nTry another code or press ❌ Cancel to go back.", reply_markup=input_kb())
        return
    data    = await state.get_data()
    product = data.get("product", {})
    qty     = data.get("qty", 1)
    disc    = coupon["discount_pct"]
    total   = round(qty * product.get("price", 0) * (1 - disc / 100), 4)
    await state.update_data(coupon_id=coupon["coupon_id"], discount_pct=disc, total=total)
    user    = await get_user(message.from_user.id) or {}
    balance = user.get("balance", 0)
    await state.set_state(UserFlow.mail_coupon)
    await message.answer(
        f"🎟 <b>Coupon Applied!</b>\n{_SEP}\n"
        f"Code: <code>{code.upper()}</code>\n"
        f"Discount: <b>{disc:.0f}%</b>\n"
        f"New Total: <b>${total:.2f}</b>\n\n"
        f"Press ✅ <b>Confirm Purchase</b> to proceed.",
        reply_markup=confirm_kb(),
    )


@router_mail.message(UserFlow.mail_confirm)
async def mail_confirm(message: Message, state: FSMContext):
    if message.text in (BACK_BTN, CANCEL_BTN):
        data      = await state.get_data()
        dm        = data.get("dm", {})
        pid       = data.get("pid")
        product   = data.get("product", {})
        stock     = data.get("stock", 0)
        qty       = data.get("qty", 1)
        total     = data.get("total", 0)
        disc      = data.get("discount_pct", 0)
        coupon_id = data.get("coupon_id")
        user      = await get_user(message.from_user.id) or {}
        balance   = user.get("balance", 0)
        await state.set_data({"dm": dm, "pid": pid, "product": product, "stock": stock,
                               "qty": qty, "total": total, "coupon_id": coupon_id, "discount_pct": disc})
        await state.set_state(UserFlow.mail_coupon)
        await message.answer(
            fmt_order_summary(product, qty, total, balance, disc),
            reply_markup=confirm_kb(),
        )
        return
    if message.text != BTN_CONFIRM:
        await message.answer("❌ Press the Confirm Purchase button or Home to cancel.")
        return
    data    = await state.get_data()
    pid     = data.get("pid")
    product = data.get("product")
    qty     = data.get("qty")
    total   = data.get("total")
    if not pid or not product or qty is None or total is None:
        await send_main_menu(message, state)
        return
    uid     = message.from_user.id
    user    = await get_user(uid) or {}
    balance = user.get("balance", 0)

    # ── Bonus logic ──
    bonus = await get_bonus(uid)
    bonus_used, remaining_cost = compute_bonus_usage(bonus, pid, total)

    actual_deduct = remaining_cost

    if balance < actual_deduct:
        await state.clear()
        await message.answer(
            f"❌ <b>Insufficient Balance</b>\n{_SEP}\n"
            f"💵 Required: <b>${total:.2f}</b>\n"
            + (f"🎁 Bonus Applied: <b>${bonus_used:.2f}</b>\n" if bonus_used > 0 else "")
            + f"💰 Your Balance: <b>${balance:.2f}</b>\n"
            f"Shortfall: <b>${remaining_cost - balance:.2f}</b>\n\nPlease deposit to continue.",
            reply_markup=main_menu_kb(),
        )
        return
    stock = await get_stock_count(pid)
    if stock < qty:
        await state.clear()
        await message.answer(
            f"❌ <b>Not Enough Stock</b>\n{_SEP}\nOnly {stock} left. Please try a smaller quantity.",
            reply_markup=main_menu_kb(),
        )
        return
    items   = await pop_stock_items(pid, qty)
    # Deduct bonus first, then regular balance for remainder + hidden surcharge
    if bonus_used > 0:
        await use_bonus(uid, bonus_used)
    new_bal = await update_balance(uid, -actual_deduct)
    oid     = await create_order(uid, pid, product["name"], qty, total, items)
    if data.get("coupon_id"):
        await use_coupon(data["coupon_id"])
    # Append to persistent My_Mail_Shop_Orders.xlsx
    try:
        uname = (await get_user(uid) or {}).get("username", "")
        append_to_mail_shop_file(oid, uid, uname, product["name"], qty, total, items, int(time.time()))
    except Exception as _e:
        logger.warning("Mail shop file append failed: %s", _e)
    await state.clear()
    bonus_note = f"\n🎁 Paid from Bonus: <b>${bonus_used:.2f}</b>" if bonus_used > 0 else ""
    await message.answer(fmt_order_receipt(oid, product["name"], qty, total, items, new_bal) + bonus_note, reply_markup=main_menu_kb())

    # ── Referral commission ──
    await pay_referral_commission(message.bot, uid, total, oid)

    # ── Auto-set mail session when exactly 1 item is purchased ────────
    if qty == 1 and items:
        first_item = items[0]
        try:
            parsed = _dv_parse_input(first_item)
            if not parsed["email"] and ":" in first_item:
                p = first_item.split(":", 1)
                parsed["email"]    = p[0].strip()
                parsed["password"] = p[1].strip()

            if parsed["email"] and parsed["password"]:
                mail_email    = parsed["email"]
                mail_password = parsed["password"]

                # Attempt to get OAuth2 token automatically
                await message.answer(
                    f"⏳ <b>Setting up Get Code access…</b>\n"
                    f"📧 <code>{html_lib.escape(mail_email)}</code>"
                )
                oauth2 = None
                if parsed.get("has_oauth2"):
                    oauth2 = {"refresh_token": parsed["refresh_token"], "client_id": parsed["client_id"]}
                else:
                    try:
                        oauth2 = await _dv_oauth2(mail_email, mail_password)
                    except Exception as _oe:
                        logger.warning("Auto OAuth2 failed: %s", _oe)

                if oauth2:
                    _mail_sessions[uid] = {
                        "email": mail_email, "password": mail_password,
                        "refresh_token": oauth2["refresh_token"],
                        "client_id": oauth2["client_id"],
                    }
                    # Pre-load inbox
                    try:
                        msgs    = await _dv_process_mailbox(_mail_sessions[uid])
                        display = _dv_build_mailbox_display(msgs)
                    except Exception:
                        display = ""
                    await message.answer(
                        _truncate_msg(
                            f"✅ <b>Mail Auto-Set in Get Code!</b>\n{_SEP}\n"
                            f"📧 <code>{html_lib.escape(mail_email)}</code>\n"
                            f"🔑 OAuth2: ✅ Ready\n\n"
                            f"Press <b>🔓 Get Code</b> to read codes directly.\n"
                            + (f"\n{display}" if display else "")
                        ),
                        reply_markup=main_menu_kb(),
                    )
                else:
                    # Set without OAuth2 — user can complete setup manually
                    _mail_sessions[uid] = {
                        "email": mail_email, "password": mail_password,
                        "refresh_token": "", "client_id": "",
                    }
                    await message.answer(
                        f"✅ <b>Mail Auto-Set!</b>\n{_SEP}\n"
                        f"📧 <code>{html_lib.escape(mail_email)}</code>\n"
                        f"🔑 OAuth2: ⚠️ Not available\n\n"
                        f"Press <b>🔓 Get Code</b> → <b>📧 Set Mail</b> and re-enter the mail to activate full access.",
                        reply_markup=main_menu_kb(),
                    )
        except Exception as _ae:
            logger.warning("Auto-set mail session failed: %s", _ae)

    # Send purchased items as xlsx file to the user
    try:
        xlsx_bytes = make_user_order_xlsx(oid, product["name"], qty, total, items)
        date_str   = time.strftime("%Y%m%d_%H%M", time.gmtime())
        filename   = f"order_{oid}_{date_str}.xlsx"
        await message.answer_document(
            document=BufferedInputFile(xlsx_bytes, filename=filename),
            caption=(
                f"📋 <b>Your Accounts File</b>\n{_SEP}\n"
                f"📦 {product['name']} × {qty}\n"
                f"🆔 Order: <code>{oid}</code>\n\n"
                f"<i>Open this file to see all your accounts.</i>"
            ),
        )
    except Exception as _xe:
        logger.warning("Failed to send order xlsx to user: %s", _xe)

    # ── Low Stock Auto-Alert ───────────────────────────────────────
    await _check_low_stock_alert(message.bot, pid, product['name'])


# ══════════════════════════════════════════════════════════════════
# ROUTER — VPN
# ══════════════════════════════════════════════════════════════════

router_vpn = Router()


@router_vpn.message(F.text == BTN_BUY_VPN)
async def vpn_start(message: Message, state: FSMContext):
    products = await get_service_products("vpn")
    if not products:
        await message.answer(f"🌐 <b>VPN Products</b>\n{_SEP}\n❌ No VPN products available.", reply_markup=main_menu_kb())
        return
    dm = {f"{p.get('emoji','🌐')} {p['name']}  ·  ${p['price']:.2f}/day": pid for pid, p in products.items()}
    await state.update_data(vpn_dm=dm)
    await state.set_state(UserFlow.vpn_product)
    await message.answer(fmt_product_list_header("🌐 VPN Products", len(dm)), reply_markup=products_kb(dm))


@router_vpn.message(UserFlow.vpn_product)
async def vpn_product_selected(message: Message, state: FSMContext):
    if message.text == BACK_BTN:
        await send_main_menu(message, state)
        return
    data   = await state.get_data()
    vpn_dm = data.get("vpn_dm", {})
    pid    = vpn_dm.get(message.text)
    if not pid:
        await message.answer("❌ Select a product from the keyboard.", reply_markup=products_kb(vpn_dm))
        return
    product = await get_product(pid)
    await state.update_data(vpn_pid=pid, vpn_product=product)
    await state.set_state(UserFlow.vpn_duration)
    await message.answer(
        f"🌐 <b>{product['name']}</b>\n{_SEP}\n"
        f"💵 Price: <b>${product['price']:.2f}</b> per day\n\nSelect duration:",
        reply_markup=duration_kb(),
    )


@router_vpn.message(UserFlow.vpn_duration)
async def vpn_duration_selected(message: Message, state: FSMContext):
    if message.text == BACK_BTN:
        data   = await state.get_data()
        vpn_dm = data.get("vpn_dm", {})
        await state.set_data({"vpn_dm": vpn_dm})
        await state.set_state(UserFlow.vpn_product)
        await message.answer(fmt_product_list_header("🌐 VPN Products", len(vpn_dm)), reply_markup=products_kb(vpn_dm))
        return
    if message.text == BTN_CUSTOM:
        await state.set_state(UserFlow.vpn_custom)
        await message.answer(f"✏️ <b>Custom Duration</b>\n{_SEP}\nEnter number of days (1–365):", reply_markup=input_kb())
        return
    days = DURATION_MAP.get(message.text)
    if not days:
        await message.answer("❌ Select a duration from the keyboard.")
        return
    await _vpn_confirm(message, state, days)


@router_vpn.message(UserFlow.vpn_custom)
async def vpn_custom_days(message: Message, state: FSMContext):
    if message.text in (CANCEL_BTN, BACK_BTN):
        data    = await state.get_data()
        vpn_dm  = data.get("vpn_dm", {})
        vpn_pid = data.get("vpn_pid")
        prod    = data.get("vpn_product", {})
        await state.set_data({"vpn_dm": vpn_dm, "vpn_pid": vpn_pid, "vpn_product": prod})
        await state.set_state(UserFlow.vpn_duration)
        await message.answer(f"🌐 <b>{prod.get('name','')}</b>\n{_SEP}\nSelect duration:", reply_markup=duration_kb())
        return
    days, err = validate_custom_days(message.text)
    if err:
        await message.answer(err)
        return
    await _vpn_confirm(message, state, days)


async def _vpn_confirm(message: Message, state: FSMContext, days: int):
    data    = await state.get_data()
    pid     = data.get("vpn_pid")
    if not pid:
        await send_main_menu(message, state)
        return
    product = await get_product(pid) or data.get("vpn_product", {})
    await state.update_data(vpn_product=product)
    price   = round(product["price"] * days, 4)
    await state.update_data(vpn_days=days, vpn_price=price, vpn_coupon_id=None, vpn_discount_pct=0)
    await state.set_state(UserFlow.vpn_coupon)
    user = await get_user(message.from_user.id) or {}
    balance = user.get("balance", 0)
    await message.answer(
        f"🛒 <b>Order Summary</b>\n{_SEP}\n"
        f"🌐 {product['name']}\n"
        f"⏱ Duration: <b>{days} day(s)</b>\n"
        f"💵 Price/day: <b>${product['price']:.2f}</b>\n"
        f"💰 Total: <b>${price:.2f}</b>\n{_LINE}\n"
        f"👛 Your Balance: <b>${balance:.2f}</b>\n\n"
        f"Have a coupon code? Enter it below,\nor press ✅ Confirm Purchase to skip.",
        reply_markup=confirm_kb(),
    )


@router_vpn.message(UserFlow.vpn_coupon)
async def vpn_coupon_or_confirm(message: Message, state: FSMContext):
    if message.text in (BACK_BTN, CANCEL_BTN):
        data    = await state.get_data()
        vpn_dm  = data.get("vpn_dm", {})
        vpn_pid = data.get("vpn_pid")
        prod    = data.get("vpn_product", {})
        await state.set_data({"vpn_dm": vpn_dm, "vpn_pid": vpn_pid, "vpn_product": prod})
        await state.set_state(UserFlow.vpn_duration)
        await message.answer(
            f"🌐 <b>{prod.get('name', 'VPN')}</b>\n{_SEP}\n"
            f"💵 Price/day: <b>${prod.get('price', 0):.2f}</b>\n\nSelect duration:",
            reply_markup=duration_kb(),
        )
        return
    if message.text == BTN_COUPON:
        await state.set_state(UserFlow.vpn_coupon_input)
        await message.answer(
            f"🎟 <b>Apply Coupon</b>\n{_SEP}\nType your coupon code below:",
            reply_markup=input_kb(),
        )
        return
    if message.text != BTN_CONFIRM:
        await message.answer(
            "❌ Please press <b>✅ Confirm Purchase</b> to proceed, or <b>🎟 Apply Coupon</b> to enter a coupon code.",
            reply_markup=confirm_kb(),
        )
        return
    data         = await state.get_data()
    product      = data.get("vpn_product", {})
    days         = data.get("vpn_days", 0)
    price        = data.get("vpn_price", 0)
    if not product or not days:
        await send_main_menu(message, state)
        return
    discount_pct = data.get("vpn_discount_pct", 0)
    user         = await get_user(message.from_user.id) or {}
    balance      = user.get("balance", 0)
    disc_line    = f"🎟 Coupon: <b>-{discount_pct:.0f}%</b>\n" if discount_pct else ""
    await state.set_state(UserFlow.vpn_confirm)
    await message.answer(
        f"🛒 <b>Order Summary</b>\n{_SEP}\n"
        f"🌐 {product['name']}\n"
        f"⏱ Duration: <b>{days} day(s)</b>\n"
        f"💵 Price/day: <b>${product['price']:.2f}</b>\n"
        f"{disc_line}"
        f"💰 Total: <b>${price:.2f}</b>\n{_LINE}\n"
        f"👛 Your Balance: <b>${balance:.2f}</b>\n\n"
        f"⚠️ This will deduct <b>${price:.2f}</b> from your balance.",
        reply_markup=confirm_final_kb(),
    )


@router_vpn.message(UserFlow.vpn_coupon_input)
async def vpn_coupon_code_entered(message: Message, state: FSMContext):
    if message.text in (CANCEL_BTN, BACK_BTN, HOME_BTN):
        data    = await state.get_data()
        vpn_dm  = data.get("vpn_dm", {})
        vpn_pid = data.get("vpn_pid")
        product = data.get("vpn_product", {})
        days    = data.get("vpn_days", 0)
        price   = data.get("vpn_price", 0)
        user    = await get_user(message.from_user.id) or {}
        disc    = 0
        disc_line = ""
        await state.set_data({"vpn_dm": vpn_dm, "vpn_pid": vpn_pid, "vpn_product": product,
                               "vpn_days": days, "vpn_price": price,
                               "vpn_coupon_id": None, "vpn_discount_pct": 0})
        await state.set_state(UserFlow.vpn_coupon)
        await message.answer(
            f"🛒 <b>Order Summary</b>\n{_SEP}\n"
            f"🌐 {product['name']}\n"
            f"⏱ Duration: <b>{days} day(s)</b>\n"
            f"💵 Price/day: <b>${product['price']:.2f}</b>\n"
            f"{disc_line}"
            f"💰 Total: <b>${price:.2f}</b>\n{_LINE}\n"
            f"👛 Your Balance: <b>${user.get('balance', 0):.2f}</b>\n\n"
            f"Have a coupon code? Enter it below,\nor press ✅ Confirm Purchase to skip.",
            reply_markup=confirm_kb(),
        )
        return
    code   = message.text.strip()
    coupon = await get_coupon(code)
    if not coupon:
        await message.answer(
            f"❌ Invalid coupon code: <code>{code}</code>\n\nTry again or press ❌ Cancel to go back.",
            reply_markup=input_kb(),
        )
        return
    valid, err = validate_coupon(coupon)
    if not valid:
        await message.answer(err + "\n\nTry another code or press ❌ Cancel to go back.", reply_markup=input_kb())
        return
    data    = await state.get_data()
    product = data.get("vpn_product", {})
    days    = data.get("vpn_days", 0)
    disc    = coupon["discount_pct"]
    price   = round(product.get("price", 0) * days * (1 - disc / 100), 4)
    await state.update_data(vpn_coupon_id=coupon["coupon_id"], vpn_discount_pct=disc, vpn_price=price)
    user    = await get_user(message.from_user.id) or {}
    await state.set_state(UserFlow.vpn_coupon)
    await message.answer(
        f"🎟 <b>Coupon Applied!</b>\n{_SEP}\n"
        f"Code: <code>{code.upper()}</code>\n"
        f"Discount: <b>{disc:.0f}%</b>\n"
        f"New Total: <b>${price:.2f}</b>\n\n"
        f"Press ✅ <b>Confirm Purchase</b> to proceed.",
        reply_markup=confirm_kb(),
    )


@router_vpn.message(UserFlow.vpn_confirm)
async def vpn_confirm(message: Message, state: FSMContext):
    if message.text in (BACK_BTN, CANCEL_BTN):
        data    = await state.get_data()
        product = data.get("vpn_product", {})
        days    = data.get("vpn_days", 0)
        price   = data.get("vpn_price", 0)
        disc    = data.get("vpn_discount_pct", 0)
        disc_line = f"🎟 Coupon: <b>-{disc:.0f}%</b>\n" if disc else ""
        user    = await get_user(message.from_user.id) or {}
        await state.set_state(UserFlow.vpn_coupon)
        await message.answer(
            f"🛒 <b>Order Summary</b>\n{_SEP}\n"
            f"🌐 {product['name']}\n"
            f"⏱ Duration: <b>{days} day(s)</b>\n"
            f"💵 Price/day: <b>${product['price']:.2f}</b>\n"
            f"{disc_line}"
            f"💰 Total: <b>${price:.2f}</b>\n{_LINE}\n"
            f"👛 Your Balance: <b>${user.get('balance', 0):.2f}</b>\n\n"
            f"Have a coupon code? Enter it below,\nor press ✅ Confirm Purchase to skip.",
            reply_markup=confirm_kb(),
        )
        return
    if message.text != BTN_CONFIRM:
        await message.answer("❌ Press the Confirm Purchase button or go back.")
        return
    data    = await state.get_data()
    product = data.get("vpn_product")
    days    = data.get("vpn_days")
    price   = data.get("vpn_price")
    vpn_pid = data.get("vpn_pid")
    if not product or days is None or price is None or not vpn_pid:
        await send_main_menu(message, state)
        return
    uid     = message.from_user.id
    user    = await get_user(uid) or {}
    balance = user.get("balance") or 0

    # ── Bonus logic ──
    bonus = await get_bonus(uid)
    bonus_used, remaining_cost = compute_bonus_usage(bonus, vpn_pid, price)

    actual_deduct = remaining_cost

    if balance < actual_deduct:
        await state.clear()
        await message.answer(
            f"❌ <b>Insufficient Balance</b>\n{_SEP}\nRequired: <b>${price:.2f}</b>"
            + (f"  ·  🎁 Bonus: <b>${bonus_used:.2f}</b>" if bonus_used > 0 else "")
            + f"  ·  Yours: <b>${balance:.2f}</b>",
            reply_markup=main_menu_kb(),
        )
        return
    if bonus_used > 0:
        await use_bonus(uid, bonus_used)
    await update_balance(uid, -actual_deduct)
    if data.get("vpn_coupon_id"):
        await use_coupon(data["vpn_coupon_id"])
    oid = await create_vpn_order(uid, user.get("username",""), vpn_pid, product["name"], days, price)
    await state.clear()
    bonus_note = f"\n🎁 Paid from Bonus: ${bonus_used:.2f}" if bonus_used > 0 else ""
    await message.answer(fmt_service_order_placed("🌐", "VPN", oid, product["name"], days, price) + bonus_note, reply_markup=main_menu_kb())
    # ── Referral commission ──
    await pay_referral_commission(message.bot, uid, price, oid)
    for adm in ADMIN_IDS:
        try:
            await message.bot.send_message(
                adm,
                fmt_admin_vpn_order({"order_id": oid, "user_id": uid, "username": user.get("username",""),
                                     "product_name": product["name"], "duration_days": days,
                                     "price": price, "created_at": int(time.time())}),
                reply_markup=order_review_inline(oid, "vpn"),
            )
        except Exception:
            pass


# ══════════════════════════════════════════════════════════════════
# ROUTER — Proxy
# ══════════════════════════════════════════════════════════════════

router_proxy = Router()


@router_proxy.message(F.text == BTN_BUY_PROXY)
async def proxy_start(message: Message, state: FSMContext):
    products = await get_service_products("proxy")
    if not products:
        await message.answer(f"🔐 <b>Proxy Products</b>\n{_SEP}\n❌ No proxy products available.", reply_markup=main_menu_kb())
        return
    dm = {}
    for pid, p in products.items():
        mode = p.get("delivery_mode", "manual")
        if mode == "auto":
            stock = p.get("stock_count", 0)
            label = f"{p.get('emoji','🔐')} {p['name']}  🤖  📦{stock}  ·  ${p['price']:.2f}"
        else:
            label = f"{p.get('emoji','🔐')} {p['name']}  👤  ·  ${p['price']:.2f}/GB"
        dm[label] = pid
    await state.update_data(proxy_dm=dm)
    await state.set_state(UserFlow.proxy_product)
    await message.answer(fmt_product_list_header("🔐 Proxy Products", len(dm)), reply_markup=products_kb(dm))


@router_proxy.message(UserFlow.proxy_product)
async def proxy_product_selected(message: Message, state: FSMContext):
    if message.text == BACK_BTN:
        await send_main_menu(message, state)
        return
    data     = await state.get_data()
    proxy_dm = data.get("proxy_dm", {})
    pid      = proxy_dm.get(message.text)
    if not pid:
        await message.answer("❌ Select a product from the keyboard.", reply_markup=products_kb(proxy_dm))
        return
    product = await get_product(pid)
    await state.update_data(proxy_pid=pid, proxy_product=product)

    if product.get("delivery_mode", "manual") == "auto":
        # ── Auto delivery: ask for quantity like Mail ──────────────
        stock = await get_stock_count(pid)
        if stock == 0:
            await message.answer(
                f"🔐 <b>{product['name']}</b>\n{_SEP}\n❌ Out of stock. Try another product.",
                reply_markup=products_kb(data.get("proxy_dm", {})),
            )
            return
        await state.update_data(proxy_stock=stock)
        await state.set_state(UserFlow.proxy_auto_qty)
        await message.answer(
            f"🔐 <b>{product['name']}</b>\n{_SEP}\n"
            f"💵 Price: <b>${product['price']:.2f}</b> per item\n"
            f"📦 In Stock: <b>{stock}</b>\n\n"
            f"How many do you want? (1–{stock}):",
            reply_markup=input_kb(),
        )
    else:
        # ── Manual delivery: data amount selection ─────────────────
        settings = await get_settings()
        data_opts = settings.get("proxy_data_options", "1 GB,5 GB,10 GB,50 GB")
        await state.update_data(proxy_data_opts=data_opts)
        await state.set_state(UserFlow.proxy_duration)
        await message.answer(
            f"🔐 <b>{product['name']}</b>\n{_SEP}\n"
            f"💵 Price: <b>${product['price']:.2f}</b> per GB\n\nSelect data amount:",
            reply_markup=proxy_data_kb(data_opts),
        )


# ── AUTO PROXY — qty & confirm ──────────────────────────────────────

@router_proxy.message(UserFlow.proxy_auto_qty)
async def proxy_auto_qty_entered(message: Message, state: FSMContext):
    if message.text in (CANCEL_BTN, BACK_BTN, HOME_BTN):
        data     = await state.get_data()
        proxy_dm = data.get("proxy_dm", {})
        proxy_data_opts = data.get("proxy_data_opts", "1 GB,5 GB,10 GB,50 GB")
        await state.set_data({"proxy_dm": proxy_dm, "proxy_data_opts": proxy_data_opts})
        await state.set_state(UserFlow.proxy_product)
        await message.answer(fmt_product_list_header("🔐 Proxy Products", len(proxy_dm)), reply_markup=products_kb(proxy_dm))
        return
    data  = await state.get_data()
    stock = data.get("proxy_stock", 0)
    qty, err = validate_quantity(message.text, stock)
    if err:
        await message.answer(err)
        return
    pid     = data.get("proxy_pid")
    if not pid:
        await send_main_menu(message, state)
        return
    product = await get_product(pid) or data.get("proxy_product", {})
    await state.update_data(proxy_product=product)
    total   = round(qty * product.get("price", 0), 4)
    user    = await get_user(message.from_user.id) or {}
    await state.update_data(proxy_qty=qty, proxy_price=total, proxy_auto_coupon_id=None, proxy_auto_discount_pct=0)
    await state.set_state(UserFlow.proxy_auto_coupon)
    await message.answer(
        f"🛒 <b>Order Summary</b>\n{_SEP}\n"
        f"🔐 {product['name']}\n"
        f"🔢 Quantity: <b>{qty}</b>\n"
        f"💵 Price: <b>${product['price']:.2f}</b> × {qty}\n"
        f"💰 Total: <b>${total:.2f}</b>\n{_LINE}\n"
        f"👛 Your Balance: <b>${user.get('balance', 0):.2f}</b>\n\n"
        f"Have a coupon code? Enter it below,\nor press ✅ Confirm Purchase to skip.",
        reply_markup=confirm_kb(),
    )


@router_proxy.message(UserFlow.proxy_auto_coupon)
async def proxy_auto_coupon_or_confirm(message: Message, state: FSMContext):
    if message.text in (BACK_BTN, CANCEL_BTN, HOME_BTN):
        data     = await state.get_data()
        proxy_dm = data.get("proxy_dm", {})
        proxy_data_opts = data.get("proxy_data_opts", "1 GB,5 GB,10 GB,50 GB")
        proxy_pid = data.get("proxy_pid")
        prod     = data.get("proxy_product", {})
        stock    = data.get("proxy_stock", 0)
        await state.set_data({"proxy_dm": proxy_dm, "proxy_data_opts": proxy_data_opts,
                               "proxy_pid": proxy_pid, "proxy_product": prod, "proxy_stock": stock})
        await state.set_state(UserFlow.proxy_auto_qty)
        await message.answer(
            fmt_product_detail(prod, stock) if prod else fmt_product_list_header("🔐 Proxy Products", len(proxy_dm)),
            reply_markup=input_kb() if prod else products_kb(proxy_dm),
        )
        return
    if message.text == BTN_COUPON:
        await state.set_state(UserFlow.proxy_auto_coupon_input)
        await message.answer(
            f"🎟 <b>Apply Coupon</b>\n{_SEP}\nType your coupon code below:",
            reply_markup=input_kb(),
        )
        return
    if message.text != BTN_CONFIRM:
        await message.answer(
            "❌ Please press <b>✅ Confirm Purchase</b> to proceed, or <b>🎟 Apply Coupon</b> to enter a coupon code.",
            reply_markup=confirm_kb(),
        )
        return
    data         = await state.get_data()
    product      = data.get("proxy_product", {})
    qty          = data.get("proxy_qty", 1)
    total        = data.get("proxy_price", 0)
    discount_pct = data.get("proxy_auto_discount_pct", 0)
    user         = await get_user(message.from_user.id) or {}
    disc_line    = f"🎟 Coupon: <b>-{discount_pct:.0f}%</b>\n" if discount_pct else ""
    await state.set_state(UserFlow.proxy_auto_confirm)
    await message.answer(
        f"🛒 <b>Order Summary</b>\n{_SEP}\n"
        f"🔐 {product['name']}\n"
        f"🔢 Quantity: <b>{qty}</b>\n"
        f"💵 Price: <b>${product['price']:.2f}</b> × {qty}\n"
        f"{disc_line}"
        f"💰 Total: <b>${total:.2f}</b>\n{_LINE}\n"
        f"👛 Your Balance: <b>${user.get('balance', 0):.2f}</b>\n\n"
        f"⚠️ This will deduct <b>${total:.2f}</b> from your balance.",
        reply_markup=confirm_final_kb(),
    )


@router_proxy.message(UserFlow.proxy_auto_coupon_input)
async def proxy_auto_coupon_code_entered(message: Message, state: FSMContext):
    if message.text in (CANCEL_BTN, BACK_BTN, HOME_BTN):
        data    = await state.get_data()
        proxy_dm = data.get("proxy_dm", {})
        proxy_data_opts = data.get("proxy_data_opts", "1 GB,5 GB,10 GB,50 GB")
        proxy_pid = data.get("proxy_pid")
        product = data.get("proxy_product", {})
        stock   = data.get("proxy_stock", 0)
        qty     = data.get("proxy_qty", 1)
        total   = data.get("proxy_price", 0)
        user    = await get_user(message.from_user.id) or {}
        disc    = 0
        disc_line = ""
        await state.set_data({"proxy_dm": proxy_dm, "proxy_data_opts": proxy_data_opts,
                               "proxy_pid": proxy_pid, "proxy_product": product, "proxy_stock": stock,
                               "proxy_qty": qty, "proxy_price": total,
                               "proxy_auto_coupon_id": None, "proxy_auto_discount_pct": 0})
        await state.set_state(UserFlow.proxy_auto_coupon)
        await message.answer(
            f"🛒 <b>Order Summary</b>\n{_SEP}\n"
            f"🔐 {product['name']}\n"
            f"🔢 Quantity: <b>{qty}</b>\n"
            f"💵 Price: <b>${product['price']:.2f}</b> × {qty}\n"
            f"{disc_line}"
            f"💰 Total: <b>${total:.2f}</b>\n{_LINE}\n"
            f"👛 Your Balance: <b>${user.get('balance', 0):.2f}</b>\n\n"
            f"Have a coupon code? Enter it below,\nor press ✅ Confirm Purchase to skip.",
            reply_markup=confirm_kb(),
        )
        return
    code   = message.text.strip()
    coupon = await get_coupon(code)
    if not coupon:
        await message.answer(
            f"❌ Invalid coupon code: <code>{code}</code>\n\nTry again or press ❌ Cancel to go back.",
            reply_markup=input_kb(),
        )
        return
    valid, err = validate_coupon(coupon)
    if not valid:
        await message.answer(err + "\n\nTry another code or press ❌ Cancel to go back.", reply_markup=input_kb())
        return
    data    = await state.get_data()
    product = data.get("proxy_product", {})
    qty     = data.get("proxy_qty", 1)
    disc    = coupon["discount_pct"]
    total   = round(qty * product.get("price", 0) * (1 - disc / 100), 4)
    await state.update_data(proxy_auto_coupon_id=coupon["coupon_id"], proxy_auto_discount_pct=disc, proxy_price=total)
    user    = await get_user(message.from_user.id) or {}
    await state.set_state(UserFlow.proxy_auto_coupon)
    await message.answer(
        f"🎟 <b>Coupon Applied!</b>\n{_SEP}\n"
        f"Code: <code>{code.upper()}</code>\n"
        f"Discount: <b>{disc:.0f}%</b>\n"
        f"New Total: <b>${total:.2f}</b>\n\n"
        f"Press ✅ <b>Confirm Purchase</b> to proceed.",
        reply_markup=confirm_kb(),
    )


@router_proxy.message(UserFlow.proxy_auto_confirm)
async def proxy_auto_confirm(message: Message, state: FSMContext):
    if message.text in (BACK_BTN, CANCEL_BTN, HOME_BTN):
        data     = await state.get_data()
        proxy_dm = data.get("proxy_dm", {})
        proxy_data_opts = data.get("proxy_data_opts", "1 GB,5 GB,10 GB,50 GB")
        proxy_pid = data.get("proxy_pid")
        product  = data.get("proxy_product", {})
        stock    = data.get("proxy_stock", 0)
        qty      = data.get("proxy_qty", 1)
        total    = data.get("proxy_price", 0)
        disc     = data.get("proxy_auto_discount_pct", 0)
        disc_line = f"🎟 Coupon: <b>-{disc:.0f}%</b>\n" if disc else ""
        user     = await get_user(message.from_user.id) or {}
        await state.set_data({"proxy_dm": proxy_dm, "proxy_data_opts": proxy_data_opts,
                               "proxy_pid": proxy_pid, "proxy_product": product, "proxy_stock": stock,
                               "proxy_qty": qty, "proxy_price": total,
                               "proxy_auto_coupon_id": data.get("proxy_auto_coupon_id"),
                               "proxy_auto_discount_pct": disc})
        await state.set_state(UserFlow.proxy_auto_coupon)
        await message.answer(
            f"🛒 <b>Order Summary</b>\n{_SEP}\n"
            f"🔐 {product.get('name', '')}\n"
            f"🔢 Quantity: <b>{qty}</b>\n"
            f"💵 Price: <b>${product.get('price', 0):.2f}</b> × {qty}\n"
            f"{disc_line}"
            f"💰 Total: <b>${total:.2f}</b>\n{_LINE}\n"
            f"👛 Your Balance: <b>${user.get('balance', 0):.2f}</b>\n\n"
            f"Have a coupon code? Enter it below,\nor press ✅ Confirm Purchase to skip.",
            reply_markup=confirm_kb(),
        )
        return
    if message.text != BTN_CONFIRM:
        await message.answer("❌ Press the Confirm Purchase button or go back.", reply_markup=confirm_final_kb())
        return
    data    = await state.get_data()
    pid     = data.get("proxy_pid")
    product = data.get("proxy_product", {})
    qty     = data.get("proxy_qty", 1)
    total   = data.get("proxy_price", 0)
    if not pid or not product:
        await send_main_menu(message, state)
        return
    uid     = message.from_user.id
    user    = await get_user(uid) or {}
    balance = user.get("balance", 0)

    # ── Bonus logic ──
    bonus = await get_bonus(uid)
    bonus_used, remaining_cost = compute_bonus_usage(bonus, pid, total)

    actual_deduct = remaining_cost

    if balance < actual_deduct:
        await state.clear()
        await message.answer(
            f"❌ <b>Insufficient Balance</b>\n{_SEP}\n"
            f"💵 Required: <b>${total:.2f}</b>\n"
            + (f"🎁 Bonus Applied: <b>${bonus_used:.2f}</b>\n" if bonus_used > 0 else "")
            + f"💰 Your Balance: <b>${balance:.2f}</b>\n"
            f"Shortfall: <b>${remaining_cost - balance:.2f}</b>\n\nPlease deposit to continue.",
            reply_markup=main_menu_kb(),
        )
        return
    stock = await get_stock_count(pid)
    if stock < qty:
        await state.clear()
        await message.answer(
            f"❌ <b>Not Enough Stock</b>\n{_SEP}\nOnly {stock} left. Please try a smaller quantity.",
            reply_markup=main_menu_kb(),
        )
        return

    items   = await pop_stock_items(pid, qty)
    if bonus_used > 0:
        await use_bonus(uid, bonus_used)
    new_bal = await update_balance(uid, -actual_deduct)
    if data.get("proxy_auto_coupon_id"):
        await use_coupon(data["proxy_auto_coupon_id"])
    oid     = await create_proxy_order(uid, user.get("username", ""), pid, product["name"], "Auto", total, items=items)
    await state.clear()

    preview = "\n".join(items[:5])
    more    = f"\n<i>... and {len(items) - 5} more</i>" if len(items) > 5 else ""
    bonus_note = f"\n🎁 Paid from Bonus: <b>${bonus_used:.2f}</b>" if bonus_used > 0 else ""
    await message.answer(
        f"✅ <b>Proxy Order Delivered!</b>\n{_SEP}\n"
        f"🆔 Order: <code>{oid}</code>\n"
        f"🔐 {product['name']} × {qty}\n"
        f"💵 Paid: <b>${total:.2f}</b>\n"
        f"👛 Balance left: <b>${new_bal:.2f}</b>\n{bonus_note}{_SEP}\n"
        f"📋 <b>Your Proxies:</b>\n<code>{preview}{more}</code>",
        reply_markup=main_menu_kb(),
    )

    # Send purchased items as xlsx file to the user
    try:
        xlsx_bytes = make_user_order_xlsx(oid, product["name"], qty, total, items)
        date_str = time.strftime("%Y%m%d_%H%M", time.gmtime())
        filename = f"order_{oid}_{date_str}.xlsx"
        await message.answer_document(
            document=BufferedInputFile(xlsx_bytes, filename=filename),
            caption=(
                f"📋 <b>Your Proxies File</b>\n{_SEP}\n"
                f"🔐 {product['name']} \u00d7 {qty}\n"
                f"🆔 Order: <code>{oid}</code>\n\n"
                f"<i>Open this file to see all your proxy details.</i>"
            ),
        )
    except Exception as _xe:
        logger.warning("Failed to send proxy order xlsx to user: %s", _xe)

    # Low stock alert
    await _check_low_stock_alert(message.bot, pid, product['name'])

    # ── Referral commission ──
    await pay_referral_commission(message.bot, uid, total, oid)


# ── MANUAL PROXY — duration & confirm ──────────────────────────────

@router_proxy.message(UserFlow.proxy_duration)
async def proxy_duration_selected(message: Message, state: FSMContext):
    if message.text == BACK_BTN:
        data            = await state.get_data()
        proxy_dm        = data.get("proxy_dm", {})
        proxy_data_opts = data.get("proxy_data_opts", "1 GB,5 GB,10 GB,50 GB")
        await state.set_data({"proxy_dm": proxy_dm, "proxy_data_opts": proxy_data_opts})
        await state.set_state(UserFlow.proxy_product)
        await message.answer(fmt_product_list_header("🔐 Proxy Products", len(proxy_dm)), reply_markup=products_kb(proxy_dm))
        return
    if message.text == BTN_PROXY_DATA_CUSTOM:
        await state.set_state(UserFlow.proxy_custom)
        await message.answer(f"📡 <b>Custom Data Amount</b>\n{_SEP}\nEnter data amount (e.g. 3 GB, 500 MB):", reply_markup=input_kb())
        return
    data = await state.get_data()
    data_opts = data.get("proxy_data_opts", "1 GB,5 GB,10 GB,50 GB")
    all_opts_bold = [_b(o) for o in parse_proxy_data_options(data_opts)]
    if message.text not in all_opts_bold:
        await message.answer("❌ Select a data option from the keyboard.")
        return
    raw_label = parse_proxy_data_options(data_opts)[all_opts_bold.index(message.text)]
    await _proxy_confirm(message, state, raw_label)


@router_proxy.message(UserFlow.proxy_custom)
async def proxy_custom_days(message: Message, state: FSMContext):
    if message.text in (CANCEL_BTN, BACK_BTN):
        data            = await state.get_data()
        proxy_dm        = data.get("proxy_dm", {})
        proxy_data_opts = data.get("proxy_data_opts", "1 GB,5 GB,10 GB,50 GB")
        proxy_pid       = data.get("proxy_pid")
        prod            = data.get("proxy_product", {})
        await state.set_data({"proxy_dm": proxy_dm, "proxy_data_opts": proxy_data_opts,
                               "proxy_pid": proxy_pid, "proxy_product": prod})
        await state.set_state(UserFlow.proxy_duration)
        await message.answer(f"🔐 <b>{prod.get('name','')}</b>\n{_SEP}\nSelect data amount:", reply_markup=proxy_data_kb(proxy_data_opts))
        return
    amount, err = parse_data_amount(message.text)
    if err:
        await message.answer(err)
        return
    raw = message.text.strip()
    unit = "MB" if "MB" in raw.upper() else "GB"
    label = f"{amount:.0f} {unit}" if amount == int(amount) else f"{amount} {unit}"
    await _proxy_confirm(message, state, label)


async def _proxy_confirm(message: Message, state: FSMContext, data_label: str):
    data    = await state.get_data()
    pid     = data.get("proxy_pid")
    if not pid:
        await send_main_menu(message, state)
        return
    product = await get_product(pid) or data.get("proxy_product", {})
    await state.update_data(proxy_product=product)
    amount, _ = parse_data_amount(data_label)
    amount = amount or 1.0
    price   = round(product.get("price", 0) * amount, 4)
    await state.update_data(proxy_days=data_label, proxy_price=price, proxy_coupon_id=None, proxy_discount_pct=0)
    await state.set_state(UserFlow.proxy_coupon)
    user = await get_user(message.from_user.id) or {}
    balance = user.get("balance", 0)
    await message.answer(
        f"🛒 <b>Order Summary</b>\n{_SEP}\n"
        f"🔐 {product['name']}\n"
        f"📡 Data: <b>{data_label}</b>\n"
        f"💵 Price/GB: <b>${product['price']:.2f}</b>\n"
        f"💰 Total: <b>${price:.2f}</b>\n{_LINE}\n"
        f"👛 Your Balance: <b>${balance:.2f}</b>\n\n"
        f"Have a coupon code? Enter it below,\nor press ✅ Confirm Purchase to skip.",
        reply_markup=confirm_kb(),
    )


@router_proxy.message(UserFlow.proxy_coupon)
async def proxy_coupon_or_confirm(message: Message, state: FSMContext):
    if message.text in (BACK_BTN, CANCEL_BTN):
        data            = await state.get_data()
        proxy_dm        = data.get("proxy_dm", {})
        proxy_data_opts = data.get("proxy_data_opts", "1 GB,5 GB,10 GB,50 GB")
        proxy_pid       = data.get("proxy_pid")
        prod            = data.get("proxy_product", {})
        await state.set_data({"proxy_dm": proxy_dm, "proxy_data_opts": proxy_data_opts,
                               "proxy_pid": proxy_pid, "proxy_product": prod})
        await state.set_state(UserFlow.proxy_duration)
        await message.answer(
            f"🔐 <b>{prod.get('name', 'Proxy')}</b>\n{_SEP}\n"
            f"💵 Price/GB: <b>${prod.get('price', 0):.2f}</b>\n\nSelect data amount:",
            reply_markup=proxy_data_kb(proxy_data_opts),
        )
        return
    if message.text == BTN_COUPON:
        await state.set_state(UserFlow.proxy_coupon_input)
        await message.answer(
            f"🎟 <b>Apply Coupon</b>\n{_SEP}\nType your coupon code below:",
            reply_markup=input_kb(),
        )
        return
    if message.text != BTN_CONFIRM:
        await message.answer(
            "❌ Please press <b>✅ Confirm Purchase</b> to proceed, or <b>🎟 Apply Coupon</b> to enter a coupon code.",
            reply_markup=confirm_kb(),
        )
        return
    data         = await state.get_data()
    product      = data.get("proxy_product", {})
    days         = data.get("proxy_days", "")
    price        = data.get("proxy_price", 0)
    if not product or not days:
        await send_main_menu(message, state)
        return
    discount_pct = data.get("proxy_discount_pct", 0)
    user         = await get_user(message.from_user.id) or {}
    balance      = user.get("balance", 0)
    disc_line    = f"🎟 Coupon: <b>-{discount_pct:.0f}%</b>\n" if discount_pct else ""
    await state.set_state(UserFlow.proxy_confirm)
    await message.answer(
        f"🛒 <b>Order Summary</b>\n{_SEP}\n"
        f"🔐 {product['name']}\n"
        f"📡 Data: <b>{days}</b>\n"
        f"💵 Price/GB: <b>${product['price']:.2f}</b>\n"
        f"{disc_line}"
        f"💰 Total: <b>${price:.2f}</b>\n{_LINE}\n"
        f"👛 Your Balance: <b>${balance:.2f}</b>\n\n"
        f"⚠️ This will deduct <b>${price:.2f}</b> from your balance.",
        reply_markup=confirm_final_kb(),
    )


@router_proxy.message(UserFlow.proxy_coupon_input)
async def proxy_coupon_code_entered(message: Message, state: FSMContext):
    if message.text in (CANCEL_BTN, BACK_BTN, HOME_BTN):
        data    = await state.get_data()
        product = data.get("proxy_product", {})
        days    = data.get("proxy_days", "")
        price   = data.get("proxy_price", 0)
        user    = await get_user(message.from_user.id) or {}
        disc    = data.get("proxy_discount_pct", 0)
        disc_line = f"🎟 Coupon: <b>-{disc:.0f}%</b>\n" if disc else ""
        await state.set_state(UserFlow.proxy_coupon)
        await message.answer(
            f"🛒 <b>Order Summary</b>\n{_SEP}\n"
            f"🔐 {product['name']}\n"
            f"📡 Data: <b>{days}</b>\n"
            f"💵 Price/GB: <b>${product['price']:.2f}</b>\n"
            f"{disc_line}"
            f"💰 Total: <b>${price:.2f}</b>\n{_LINE}\n"
            f"👛 Your Balance: <b>${user.get('balance', 0):.2f}</b>\n\n"
            f"Have a coupon code? Enter it below,\nor press ✅ Confirm Purchase to skip.",
            reply_markup=confirm_kb(),
        )
        return
    code   = message.text.strip()
    coupon = await get_coupon(code)
    if not coupon:
        await message.answer(
            f"❌ Invalid coupon code: <code>{code}</code>\n\nTry again or press ❌ Cancel to go back.",
            reply_markup=input_kb(),
        )
        return
    valid, err = validate_coupon(coupon)
    if not valid:
        await message.answer(err + "\n\nTry another code or press ❌ Cancel to go back.", reply_markup=input_kb())
        return
    data    = await state.get_data()
    product = data.get("proxy_product", {})
    days    = data.get("proxy_days", "")
    disc    = coupon["discount_pct"]
    amt, _  = parse_data_amount(days)
    amt     = amt or 1.0
    price   = round(product.get("price", 0) * amt * (1 - disc / 100), 4)
    await state.update_data(proxy_coupon_id=coupon["coupon_id"], proxy_discount_pct=disc, proxy_price=price)
    user    = await get_user(message.from_user.id) or {}
    await state.set_state(UserFlow.proxy_coupon)
    await message.answer(
        f"🎟 <b>Coupon Applied!</b>\n{_SEP}\n"
        f"Code: <code>{code.upper()}</code>\n"
        f"Discount: <b>{disc:.0f}%</b>\n"
        f"New Total: <b>${price:.2f}</b>\n\n"
        f"Press ✅ <b>Confirm Purchase</b> to proceed.",
        reply_markup=confirm_kb(),
    )


@router_proxy.message(UserFlow.proxy_confirm)
async def proxy_confirm(message: Message, state: FSMContext):
    if message.text in (BACK_BTN, CANCEL_BTN):
        data    = await state.get_data()
        product = data.get("proxy_product", {})
        days    = data.get("proxy_days", "")
        price   = data.get("proxy_price", 0)
        disc    = data.get("proxy_discount_pct", 0)
        disc_line = f"🎟 Coupon: <b>-{disc:.0f}%</b>\n" if disc else ""
        user    = await get_user(message.from_user.id) or {}
        await state.set_state(UserFlow.proxy_coupon)
        await message.answer(
            f"🛒 <b>Order Summary</b>\n{_SEP}\n"
            f"🔐 {product['name']}\n"
            f"📡 Data: <b>{days}</b>\n"
            f"💵 Price/GB: <b>${product['price']:.2f}</b>\n"
            f"{disc_line}"
            f"💰 Total: <b>${price:.2f}</b>\n{_LINE}\n"
            f"👛 Your Balance: <b>${user.get('balance', 0):.2f}</b>\n\n"
            f"Have a coupon code? Enter it below,\nor press ✅ Confirm Purchase to skip.",
            reply_markup=confirm_kb(),
        )
        return
    if message.text != BTN_CONFIRM:
        await message.answer("❌ Press the Confirm Purchase button or go back.", reply_markup=confirm_final_kb())
        return
    data    = await state.get_data()
    product = data.get("proxy_product", {})
    days    = data.get("proxy_days", "")
    price   = data.get("proxy_price", 0)
    if not product or not days:
        await send_main_menu(message, state)
        return
    uid     = message.from_user.id
    user    = await get_user(uid) or {}
    balance = user.get("balance") or 0
    proxy_pid = data.get("proxy_pid", "")

    # ── Bonus logic ──
    bonus = await get_bonus(uid)
    bonus_used, remaining_cost = compute_bonus_usage(bonus, proxy_pid, price)

    actual_deduct = remaining_cost

    if balance < actual_deduct:
        await state.clear()
        await message.answer(
            f"❌ <b>Insufficient Balance</b>\n{_SEP}\nRequired: <b>${price:.2f}</b>"
            + (f"  ·  🎁 Bonus: <b>${bonus_used:.2f}</b>" if bonus_used > 0 else "")
            + f"  ·  Yours: <b>${balance:.2f}</b>",
            reply_markup=main_menu_kb(),
        )
        return
    if bonus_used > 0:
        await use_bonus(uid, bonus_used)
    await update_balance(uid, -actual_deduct)
    if data.get("proxy_coupon_id"):
        await use_coupon(data.get("proxy_coupon_id"))
    oid = await create_proxy_order(uid, user.get("username",""), proxy_pid, product.get("name",""), days, price)
    await state.clear()
    bonus_note = f"\n🎁 Paid from Bonus: <b>${bonus_used:.2f}</b>" if bonus_used > 0 else ""
    await message.answer(
        f"🔐 <b>Proxy Order Placed!</b>\n{_SEP}\n"
        f"🆔 Order: <code>{oid}</code>\n"
        f"📦 {product['name']}\n"
        f"📡 Data: <b>{days}</b>\n"
        f"💵 Paid: <b>${price:.2f}</b>\n{bonus_note}{_SEP}\n"
        f"⏳ Status: <b>Pending</b>\n"
        f"Our team will deliver your order shortly.\n"
        f"You'll receive a notification when ready. 🔔",
        reply_markup=main_menu_kb()
    )
    # ── Referral commission ──
    await pay_referral_commission(message.bot, uid, price, oid)
    for adm in ADMIN_IDS:
        try:
            await message.bot.send_message(
                adm,
                f"🔐 <b>New Proxy Order</b>\n{_SEP}\n"
                f"🆔 <code>{oid}</code>\n"
                f"👤 {uid}  @{user.get('username','')}\n"
                f"📦 {product['name']}\n"
                f"📡 Data: <b>{days}</b>\n"
                f"💵 ${price:.2f}\n"
                f"🕐 {_dt(int(time.time()))}",
                reply_markup=order_review_inline(oid, "proxy"),
            )
        except Exception:
            pass


# ══════════════════════════════════════════════════════════════════
# ROUTER — Deposit
# ══════════════════════════════════════════════════════════════════

router_deposit = Router()


@router_deposit.message(F.text == BTN_DEPOSIT)
async def deposit_start(message: Message, state: FSMContext):
    await state.set_state(UserFlow.dep_method)
    user = await get_user(message.from_user.id) or {}
    await message.answer(
        f"💵 <b>Add Funds</b>\n{_SEP}\n"
        f"💰 Current Balance: <b>${user.get('balance',0):.2f}</b>\n\nChoose your payment method:",
        reply_markup=deposit_method_kb(),
    )


@router_deposit.message(UserFlow.dep_method)
async def deposit_method_chosen(message: Message, state: FSMContext):
    if message.text == BACK_BTN:
        await send_main_menu(message, state)
        return
    method_key = DEPOSIT_MAP.get(message.text)
    if not method_key:
        await message.answer("❌ Select a payment method from the keyboard.")
        return
    settings = await get_settings()
    if method_key == "bkash":
        number  = settings.get("bkash_number", "")
        min_amt = settings.get("bkash_min", 100)
        label   = "bKash"
    elif method_key == "nagad":
        number  = settings.get("nagad_number", "")
        min_amt = settings.get("nagad_min", 150)
        label   = "Nagad"
    else:
        number  = settings.get("binance_uid", "")
        min_amt = settings.get("binance_min", 5)
        label   = "Binance"
    rate   = settings.get("usd_rate", 125)
    is_usd = (method_key == "binance")
    await state.update_data(method_key=method_key, method_label=label, min_amt=min_amt, rate=rate, is_usd=is_usd, pay_number=number)
    await state.set_state(UserFlow.dep_amount)
    await message.answer(fmt_deposit_info(label, number, rate, min_amt, is_usd), reply_markup=input_kb())


@router_deposit.message(UserFlow.dep_amount)
async def deposit_amount_entered(message: Message, state: FSMContext):
    if message.text in (CANCEL_BTN, BACK_BTN):
        await state.set_state(UserFlow.dep_method)
        await message.answer("💵 Select payment method:", reply_markup=deposit_method_kb())
        return
    data   = await state.get_data()
    is_usd = data.get("is_usd", False)
    if is_usd:
        amount_usd, err = validate_price(message.text)
        if err:
            await message.answer(err)
            return
        if amount_usd <= 0:
            await message.answer("❌ Amount must be positive.")
            return
        if amount_usd < data.get("min_amt", 0):
            await message.answer(f"❌ Minimum Binance deposit is <b>${data.get('min_amt', 0):.2f} USD</b>")
            return
        rate       = data.get("rate") or 125
        amount_bdt = round(amount_usd * rate, 2)
    else:
        amount_bdt, err = validate_amount_bdt(message.text, data.get("min_amt", 0))
        if err:
            await message.answer(err)
            return
        rate = data.get("rate") or 0
        if rate <= 0:
            await message.answer("❌ Exchange rate is not configured. Contact admin.")
            return
        amount_usd = round(amount_bdt / rate, 4)
    await state.update_data(amount_bdt=amount_bdt, amount_usd=amount_usd)
    await state.set_state(UserFlow.dep_trx)
    await message.answer(fmt_deposit_confirm(data.get("method_label", ""), amount_bdt, amount_usd, is_usd), reply_markup=input_kb())


@router_deposit.message(UserFlow.dep_trx)
async def deposit_trx_entered(message: Message, state: FSMContext):
    if message.text in (CANCEL_BTN, BACK_BTN):
        data = await state.get_data()
        await state.set_state(UserFlow.dep_amount)
        await message.answer(fmt_deposit_info(data.get("method_label", ""), data.get("pay_number", ""), data.get("rate", 0), data.get("min_amt", 0), data.get("is_usd", False)), reply_markup=input_kb())
        return
    trx = message.text.strip()
    if await check_trx_duplicate(trx):
        await message.answer(
            f"❌ <b>Duplicate Transaction</b>\n{_SEP}\n"
            f"This TRX ID has already been submitted.\nContact support if this is an error."
        )
        return
    await state.update_data(trx_id=trx)
    await state.set_state(UserFlow.dep_screenshot)
    await message.answer(
        f"📸 <b>Upload Screenshot</b>\n{_SEP}\n"
        f"Please send a clear screenshot of your payment.\n\n<i>This helps our team verify faster. ⚡</i>",
        reply_markup=input_kb(),
    )


@router_deposit.message(UserFlow.dep_screenshot)
async def deposit_screenshot(message: Message, state: FSMContext):
    if message.text in (CANCEL_BTN, BACK_BTN):
        data = await state.get_data()
        await state.set_state(UserFlow.dep_trx)
        await message.answer(
            f"📝 <b>Transaction ID</b>\n{_SEP}\nPlease enter your TRX ID / Transaction reference:",
            reply_markup=input_kb(),
        )
        return
    if not message.photo:
        await message.answer("❌ Please send a photo/screenshot of your payment.")
        return
    data      = await state.get_data()
    uid       = message.from_user.id
    user      = await get_user(uid) or {}
    photo     = message.photo[-1]
    file_id   = photo.file_id          # always use Telegram file_id — no Firebase needed

    # Try Firebase upload for a permanent URL (optional, non-blocking)
    screenshot_url = ""
    try:
        file       = await message.bot.get_file(photo.file_id)
        file_bytes = (await message.bot.download_file(file.file_path)).read()
        screenshot_url = await upload_screenshot(file_bytes, f"{uid}_{int(time.time())}.jpg")
    except Exception as e:
        logger.warning("Screenshot Firebase upload skipped: %s", e)

    method_label = data.get("method_label", "")
    amount_bdt   = data.get("amount_bdt", 0)
    amount_usd   = data.get("amount_usd", 0)
    trx_id       = data.get("trx_id", "")
    dep_id = await create_deposit(
        uid, user.get("username", ""), method_label,
        amount_bdt, amount_usd, trx_id, screenshot_url,
    )
    await state.clear()
    await message.answer(
        fmt_deposit_submitted(dep_id, method_label, amount_bdt, amount_usd, data.get("is_usd", False)),
        reply_markup=main_menu_kb(),
    )
    for adm in ADMIN_IDS:
        try:
            dep = {
                "deposit_id": dep_id, "user_id": uid, "username": user.get("username", ""),
                "method": method_label, "amount_bdt": amount_bdt,
                "amount_usd": amount_usd, "trx_id": trx_id,
                "screenshot_url": screenshot_url, "created_at": int(time.time()),
            }
            text      = fmt_admin_deposit_review(dep)
            kb_inline = deposit_review_inline(dep_id)
            # Always send the photo using Telegram file_id — reliable even if Firebase failed
            await message.bot.send_photo(adm, photo=file_id, caption=text, reply_markup=kb_inline)
        except Exception as ex:
            logger.warning("Admin deposit notify error (uid=%s): %s", adm, ex)
            try:
                await message.bot.send_message(adm, text, reply_markup=kb_inline)
            except Exception:
                pass


# ══════════════════════════════════════════════════════════════════
# ROUTER — Temp Mail
# ══════════════════════════════════════════════════════════════════

router_tempmail = Router()

_TM_GEN  = _b("📧 Generate Email")
_TM_INB  = _b("📬 Check Inbox")
_TM_REF  = _b("🔄 Refresh")
_TM_MY   = _b("📋 My Email")
_TM_DEL  = _b("🗑 Delete Email")
_TM_DOM  = _b("🌐 Change Domain")
_TM_RND  = _b("🎲 Random Domain")
_TM_BTNS = {_TM_GEN, _TM_INB, _TM_REF, _TM_MY, _TM_DEL, _TM_DOM, _TM_RND}


@router_tempmail.message(F.text == BTN_TEMP_MAIL)
async def tempmail_entry(message: Message, state: FSMContext):
    uid = message.from_user.id
    await state.set_state(UserFlow.temp_mail_menu)
    sess = _tempmail_sessions.get(uid)
    if sess:
        email = sess.get("email", "")
        txt = (
            f"📧 <b>Temp Mail</b>\n{_SEP}\n"
            f"✅ Active email:\n<code>{html_lib.escape(email)}</code>\n\n"
            f"Choose an action:"
        )
    else:
        txt = (
            f"📧 <b>Temp Mail</b>\n{_SEP}\n"
            f"Get a free disposable email address instantly.\n\n"
            f"Press <b>📧 Generate Email</b> to create one."
        )
    await message.answer(txt, reply_markup=_tm_menu_kb())


@router_tempmail.message(UserFlow.temp_mail_menu, F.text.in_({BACK_BTN, HOME_BTN}))
async def tempmail_back(message: Message, state: FSMContext):
    await send_main_menu(message, state)


@router_tempmail.message(UserFlow.temp_mail_menu, F.text == _TM_GEN)
async def tempmail_generate(message: Message, state: FSMContext):
    uid = message.from_user.id
    # use domain saved in session, or default
    chosen_domain = (_tempmail_sessions.get(uid) or {}).get("chosen_domain", _TEMPMAIL_DOMAIN)
    await message.answer(f"⏳ Generating email on <b>@{chosen_domain}</b>…")
    email, token = await _tm_create_email(chosen_domain)
    if not email:
        await message.answer("❌ Failed to generate email. Please try again.", reply_markup=_tm_menu_kb())
        return
    _tempmail_sessions[uid] = {
        "email": email, "token": token,
        "chosen_domain": chosen_domain,
        "seen_ids": set(), "messages": {}, "msg_index": 0,
    }
    ts = datetime.now().strftime("%Y-%m-%d %H:%M UTC")
    await message.answer(
        f"✅ <b>Email Generated!</b>\n{_SEP}\n"
        f"📧 <code>{html_lib.escape(email)}</code>\n\n"
        f"<i>Tap the address above to copy it</i>\n"
        f"🕐 Created: {ts}",
        reply_markup=_tm_menu_kb(),
    )


@router_tempmail.message(UserFlow.temp_mail_menu, F.text == _TM_MY)
async def tempmail_myemail(message: Message, state: FSMContext):
    uid  = message.from_user.id
    sess = _tempmail_sessions.get(uid)
    if not sess:
        await message.answer("❌ No active email. Press <b>📧 Generate Email</b> first.", reply_markup=_tm_menu_kb())
        return
    await message.answer(
        f"📋 <b>Your current email:</b>\n{_SEP}\n"
        f"<code>{html_lib.escape(sess['email'])}</code>\n\n"
        f"<i>Tap to copy</i>",
        reply_markup=_tm_menu_kb(),
    )


@router_tempmail.message(UserFlow.temp_mail_menu, F.text.in_({_TM_INB, _TM_REF}))
async def tempmail_inbox(message: Message, state: FSMContext):
    uid     = message.from_user.id
    refresh = message.text == _TM_REF
    sess    = _tempmail_sessions.get(uid)
    if not sess:
        await message.answer("❌ No email found. Generate one first.", reply_markup=_tm_menu_kb())
        return

    email = sess["email"]
    await message.answer(f"🔍 Checking inbox for <code>{html_lib.escape(email)}</code>…")
    messages = await _tm_get_inbox(email)

    if not messages:
        await message.answer("📭 <b>Inbox is empty.</b>\n\n<i>Emails may take a moment to arrive.</i>", reply_markup=_tm_menu_kb())
        return

    seen     = sess.get("seen_ids", set())
    new_msgs = [m for m in messages if str(m.get("id", "")) not in seen]

    if refresh and not new_msgs:
        await message.answer(
            f"🔄 No new messages. Total: <b>{len(messages)}</b> email(s).",
            reply_markup=_tm_menu_kb(),
        )
        return

    show_msgs = new_msgs if new_msgs else messages

    stored   = sess.get("messages", {})
    mi       = sess.get("msg_index", 0)
    id_map: Dict[str, str] = {}

    for msg in messages:
        full_id  = str(msg.get("id", ""))
        existing = next((k for k, v in stored.items() if str(v.get("id", "")) == full_id), None)
        if existing:
            id_map[full_id] = existing
        else:
            short_id = f"m{mi}"
            mi += 1
            stored[short_id] = msg
            id_map[full_id]  = short_id

    _tempmail_sessions[uid]["messages"]  = stored
    _tempmail_sessions[uid]["msg_index"] = mi

    for i, msg in enumerate(show_msgs[:10]):
        full_id     = str(msg.get("id", ""))
        short_id    = id_map.get(full_id, f"m{i}")
        sender      = html_lib.escape(str(msg.get("from", "Unknown"))[:60])
        subject     = html_lib.escape(str(msg.get("subject", "(No Subject)"))[:80])
        received    = str(msg.get("created_at", ""))[:19]
        raw_preview = (msg.get("body_text") or "")[:100].replace("\n", " ")
        preview     = html_lib.escape(raw_preview)

        card = (
            f"📩 <b>Mail #{i+1}</b>\n"
            f"👤 <b>From:</b> {sender}\n"
            f"📌 <b>Subject:</b> {subject}\n"
        )
        if received:
            card += f"🕐 {received}\n"
        if preview:
            card += f"\n<i>{preview}…</i>"

        inline_kb = InlineKeyboardMarkup(inline_keyboard=[[
            InlineKeyboardButton(text="📖 Read Full Email", callback_data=f"tm_read:{uid}:{short_id}")
        ]])
        await message.answer(card, reply_markup=inline_kb)
        seen.add(full_id)
        await asyncio.sleep(0.3)

    _tempmail_sessions[uid]["seen_ids"] = seen
    label = "new" if new_msgs else "total"
    await message.answer(
        f"✅ Showing <b>{len(show_msgs)}</b> {label} email(s).",
        reply_markup=_tm_menu_kb(),
    )


@router_tempmail.callback_query(F.data.startswith("tm_read:"))
async def tempmail_read_callback(call: CallbackQuery):
    await call.answer()
    parts = call.data.split(":", 2)
    if len(parts) < 3:
        await call.answer("❌ Invalid.", show_alert=True)
        return
    uid_str, short_id = parts[1], parts[2]
    try:
        uid = int(uid_str)
    except ValueError:
        return
    sess = _tempmail_sessions.get(uid)
    if not sess:
        await call.message.answer("❌ Session expired. Please generate a new email.")
        return
    msg = sess.get("messages", {}).get(short_id)
    if not msg:
        await call.message.answer("❌ Email not found. Try checking inbox again.")
        return

    sender   = html_lib.escape(str(msg.get("from", "Unknown")))
    subject  = html_lib.escape(str(msg.get("subject", "(No Subject)")))
    received = str(msg.get("created_at", ""))[:19]

    header = (
        f"📨 <b>Full Email</b>\n{_SEP}\n"
        f"👤 <b>From:</b> {sender}\n"
        f"📌 <b>Subject:</b> {subject}\n"
    )
    if received:
        header += f"🕐 <b>Received:</b> {received}\n"
    header += _SEP

    await call.message.answer(header)

    body_formatted = _tm_format_body(
        msg.get("body_text", ""),
        msg.get("body_html", ""),
    )
    chunk_size = 3800
    chunks = [body_formatted[i:i+chunk_size] for i in range(0, len(body_formatted), chunk_size)]
    for chunk in chunks:
        if chunk.strip():
            await call.message.answer(chunk)
            await asyncio.sleep(0.2)


@router_tempmail.message(UserFlow.temp_mail_menu, F.text == _TM_DEL)
async def tempmail_delete(message: Message, state: FSMContext):
    uid  = message.from_user.id
    sess = _tempmail_sessions.pop(uid, None)
    if sess:
        deleted = html_lib.escape(sess["email"])
        await message.answer(
            f"🗑 <b>Email deleted.</b>\n{_SEP}\n"
            f"<code>{deleted}</code>\n\n"
            f"<i>Generate a new one anytime.</i>",
            reply_markup=_tm_menu_kb(),
        )
    else:
        await message.answer("❌ No active email to delete.", reply_markup=_tm_menu_kb())


@router_tempmail.message(UserFlow.temp_mail_menu, F.text == _TM_DOM)
async def tempmail_change_domain(message: Message, state: FSMContext):
    uid = message.from_user.id
    current = (_tempmail_sessions.get(uid) or {}).get("chosen_domain", _TEMPMAIL_DOMAIN)
    await message.answer("⏳ Fetching available domains…")
    domains = await _tm_get_domains()
    if _TEMPMAIL_DOMAIN not in domains:
        domains.insert(0, _TEMPMAIL_DOMAIN)
    await state.set_state(UserFlow.temp_mail_domain)
    await state.update_data(tm_domains=domains)
    await message.answer(
        f"🌐 <b>Choose Domain</b>\n{_SEP}\n"
        f"Current: <code>@{current}</code>\n\n"
        f"Select a domain for your next email.\n"
        f"<i>({len(domains)} domains available)</i>",
        reply_markup=_tm_domain_kb(domains),
    )


@router_tempmail.message(UserFlow.temp_mail_menu, F.text == _TM_RND)
async def tempmail_random_domain(message: Message, state: FSMContext):
    """Pick a random domain from the API list instantly from the main menu."""
    uid = message.from_user.id
    await message.answer("🎲 Picking a random domain…")
    domains = await _tm_get_domains()
    if not domains:
        domains = [_TEMPMAIL_DOMAIN]
    chosen = random.choice(domains)
    sess = _tempmail_sessions.get(uid) or {}
    sess["chosen_domain"] = chosen
    _tempmail_sessions[uid] = sess
    await message.answer(
        f"🎲 <b>Random domain selected:</b> <code>@{chosen}</code>\n\n"
        f"Press 📧 <b>Generate Email</b> to create a new address with this domain.",
        reply_markup=_tm_menu_kb(),
    )


@router_tempmail.message(UserFlow.temp_mail_domain)
async def tempmail_domain_selected(message: Message, state: FSMContext):
    if message.text in (BACK_BTN, HOME_BTN, CANCEL_BTN):
        await state.set_state(UserFlow.temp_mail_menu)
        uid = message.from_user.id
        current = (_tempmail_sessions.get(uid) or {}).get("chosen_domain", _TEMPMAIL_DOMAIN)
        await message.answer(
            f"📧 <b>Temp Mail</b>\n{_SEP}\nDomain unchanged: <code>@{current}</code>",
            reply_markup=_tm_menu_kb(),
        )
        return
    data    = await state.get_data()
    domains = data.get("tm_domains", [_TEMPMAIL_DOMAIN])
    uid     = message.from_user.id
    # Handle random button inside domain chooser screen too
    if message.text == _TM_RND:
        chosen = random.choice(domains)
        sess = _tempmail_sessions.get(uid) or {}
        sess["chosen_domain"] = chosen
        _tempmail_sessions[uid] = sess
        await state.set_state(UserFlow.temp_mail_menu)
        await message.answer(
            f"🎲 <b>Random domain selected:</b> <code>@{chosen}</code>\n\n"
            f"Press 📧 <b>Generate Email</b> to use it.",
            reply_markup=_tm_menu_kb(),
        )
        return
    chosen = message.text.strip().lstrip("@")
    if chosen not in domains:
        await message.answer("❌ Please select a domain from the list.")
        return
    # save chosen domain; preserve existing session data if present
    sess = _tempmail_sessions.get(uid) or {}
    sess["chosen_domain"] = chosen
    _tempmail_sessions[uid] = sess
    await state.set_state(UserFlow.temp_mail_menu)
    await message.answer(
        f"✅ <b>Domain set to</b> <code>@{chosen}</code>\n\n"
        f"Press 📧 <b>Generate Email</b> to create a new address with this domain.",
        reply_markup=_tm_menu_kb(),
    )


# ══════════════════════════════════════════════════════════════════
# ROUTER — Admin Panel
# ══════════════════════════════════════════════════════════════════

router_admin = Router()


@router_admin.message(Command("admin"))
async def cmd_admin(message: Message, state: FSMContext):
    if not is_admin(message.from_user.id):
        await message.answer("🚫 Access denied.")
        return
    await state.clear()
    await state.set_state(AdminFlow.menu)
    await message.answer(f"🔐 <b>Admin Panel</b>\n{_SEP}\nWelcome back, admin!", reply_markup=admin_main_kb())


# ── Sub-menu navigation ────────────────────────────────────────────

@router_admin.message(AdminFlow.menu, F.text == BTN_ADM_ORDERS_MENU)
async def admin_orders_submenu(message: Message, state: FSMContext):
    await state.set_state(AdminFlow.orders_submenu)
    await message.answer("📋 <b>Orders Management</b>", reply_markup=admin_orders_submenu_kb())

@router_admin.message(AdminFlow.menu, F.text == BTN_ADM_REPORTS_MENU)
async def admin_reports_submenu(message: Message, state: FSMContext):
    await state.set_state(AdminFlow.reports_submenu)
    await message.answer("📊 <b>Reports & Analytics</b>", reply_markup=admin_reports_submenu_kb())

@router_admin.message(AdminFlow.menu, F.text == BTN_ADM_TOOLS_MENU)
async def admin_tools_submenu(message: Message, state: FSMContext):
    await state.set_state(AdminFlow.tools_submenu)
    await message.answer("🔧 <b>Tools & Automation</b>", reply_markup=admin_tools_submenu_kb())

@router_admin.message(AdminFlow.orders_submenu, F.text == BACK_BTN)
async def admin_orders_back(message: Message, state: FSMContext):
    await state.set_state(AdminFlow.menu)
    await message.answer("🔐 <b>Admin Panel</b>", reply_markup=admin_main_kb())

@router_admin.message(AdminFlow.reports_submenu, F.text == BACK_BTN)
async def admin_reports_back(message: Message, state: FSMContext):
    await state.set_state(AdminFlow.menu)
    await message.answer("🔐 <b>Admin Panel</b>", reply_markup=admin_main_kb())

@router_admin.message(AdminFlow.tools_submenu, F.text == BACK_BTN)
async def admin_tools_back(message: Message, state: FSMContext):
    await state.set_state(AdminFlow.menu)
    await message.answer("🔐 <b>Admin Panel</b>", reply_markup=admin_main_kb())


# ── Dashboard ──────────────────────────────────────────────────────

@router_admin.message(AdminFlow.menu, F.text == BTN_ADM_DASHBOARD)
async def admin_dashboard(message: Message):
    stats = await get_dashboard_stats()
    await message.answer(fmt_admin_dashboard(stats), reply_markup=admin_main_kb())
    # Admin quick-action inline buttons
    admin_actions = _quick_actions_inline([
        ("\U0001F4CA Analytics", "adm_quick:analytics"),
        ("\U0001F4E6 Stock", "adm_quick:stock"),
        ("\U0001F4B3 Deposits", "adm_quick:deposits"),
    ])
    await message.answer("\u26A1 <b>Quick Actions</b>", reply_markup=admin_actions)
    # Auto-check low stock alerts
    settings = await get_settings()
    threshold = int(settings.get("low_stock_threshold", 5))
    products = await get_all_products()
    low_items = []
    _pkg_emoji = "\U0001F4E6"
    for pid, p in products.items():
        if p.get("hidden"):
            continue
        count = p.get("stock_count", 0)
        if count <= threshold:
            stock_dot = _status_dot(count > 0)
            emoji = p.get("emoji", _pkg_emoji)
            low_items.append(f"  {stock_dot} {emoji} {p.get('name', pid)}: <b>{count}</b>")
    if low_items:
        alert_text = f"\U0001F6A8 <b>Low Stock Alert</b>\n{_SEP}\nThreshold: {threshold}\n" + "\n".join(low_items[:10])
        await message.answer(alert_text)


# ── Admin Quick Action Callbacks ───────────────────────────────────

@router_admin.callback_query(F.data == "adm_quick:analytics")
async def adm_quick_analytics_cb(callback: CallbackQuery, state: FSMContext):
    if not is_admin(callback.from_user.id):
        await callback.answer("\U0001F6AB Denied", show_alert=True)
        return
    await state.set_state(AdminFlow.menu)
    await callback.message.answer("\U0001F4CA Opening analytics...", reply_markup=admin_main_kb())
    await callback.answer()


@router_admin.callback_query(F.data == "adm_quick:stock")
async def adm_quick_stock_cb(callback: CallbackQuery):
    if not is_admin(callback.from_user.id):
        await callback.answer("\U0001F6AB Denied", show_alert=True)
        return
    products = await get_all_products()
    lines = []
    _pkg_e = "\U0001F4E6"
    for pid, p in products.items():
        if p.get("hidden"):
            continue
        count = p.get("stock_count", 0)
        dot = _status_dot(count > 0)
        bar = _progress_bar(count, max(count, 50), width=6)
        emoji = p.get("emoji", _pkg_e)
        lines.append(f"{dot} {emoji} {p.get('name', pid)}: <b>{count}</b> {bar}")
    body = "\n".join(lines[:15]) if lines else "No products."
    await callback.message.answer(f"\U0001F4E6 <b>Stock Overview</b>\n{_SEP}\n{body}")
    await callback.answer()


@router_admin.callback_query(F.data == "adm_quick:deposits")
async def adm_quick_deposits_cb(callback: CallbackQuery):
    if not is_admin(callback.from_user.id):
        await callback.answer("\U0001F6AB Denied", show_alert=True)
        return
    await callback.message.answer("\U0001F4B3 Check pending deposits in the Admin menu.", reply_markup=admin_main_kb())
    await callback.answer()


# ── Deposit inline callbacks ───────────────────────────────────────

@router_admin.callback_query(F.data.startswith("dep_ok:"))
async def approve_deposit(call: CallbackQuery):
    if not is_admin(call.from_user.id):
        await call.answer("🚫 Denied", show_alert=True)
        return
    dep_id = call.data.split(":", 1)[1]
    dep    = await get_deposit(dep_id)
    if not dep or dep.get("status") != "pending":
        await call.answer("Already processed.", show_alert=True)
        return
    await update_deposit(dep_id, {"status": "approved"})
    new_bal = await update_balance(dep["user_id"], dep["amount_usd"])
    await log_admin_action(call.from_user.id, "approve_deposit", f"ID:{dep_id} User:{dep['user_id']} ${dep['amount_usd']:.2f}")
    try:
        await call.message.edit_caption((call.message.caption or "") + "\n\n✅ APPROVED", reply_markup=None)
    except Exception:
        try:
            await call.message.edit_text((call.message.text or "") + "\n\n✅ APPROVED", reply_markup=None)
        except Exception:
            pass
    await call.answer("✅ Approved")
    try:
        await call.bot.send_message(
            dep["user_id"],
            f"✅ <b>Deposit Approved!</b>\n{_SEP}\n"
            f"💵 Amount: <b>${dep['amount_usd']:.2f}</b>\n"
            f"💰 New Balance: <b>${new_bal:.2f}</b>",
        )
    except Exception:
        pass


@router_admin.callback_query(F.data.startswith("dep_no:"))
async def reject_deposit(call: CallbackQuery, state: FSMContext):
    if not is_admin(call.from_user.id):
        await call.answer("🚫 Denied", show_alert=True)
        return
    dep_id = call.data.split(":", 1)[1]
    dep    = await get_deposit(dep_id)
    if not dep or dep.get("status") != "pending":
        await call.answer("Already processed.", show_alert=True)
        return
    _pending_reject[call.from_user.id] = {"dep_id": dep_id, "dep": dep, "msg": call.message}
    await state.set_state(AdminFlow.dep_reject_reason)
    await call.answer("✏️ Enter reject reason")
    await call.message.answer(
        f"❌ <b>Reject Deposit</b>\n{_SEP}\n"
        f"User: @{dep.get('username','—')}  |  Amount: <b>${dep['amount_usd']:.2f}</b>\n\n"
        f"✏️ <b>Type the rejection reason to send to the user:</b>\n"
        f"<i>(e.g. Wrong TRX ID, screenshot unclear)</i>\n\n"
        f"Press /cancel to abort.",
    )


# ── Deposits list ──────────────────────────────────────────────────

@router_admin.message(AdminFlow.menu, F.text == BTN_ADM_DEPOSITS)
async def admin_deposits(message: Message):
    pending = await get_pending_deposits()
    if not pending:
        await message.answer("✅ No pending deposits.", reply_markup=admin_main_kb())
        return
    await message.answer(f"💳 <b>{len(pending)} Pending Deposit(s)</b>", reply_markup=admin_main_kb())
    for dep in pending[:10]:
        try:
            text      = fmt_admin_deposit_review(dep)
            kb_inline = deposit_review_inline(dep["deposit_id"])
            ss        = dep.get("screenshot_url", "")
            if ss and ss.startswith("http"):
                await message.answer_photo(photo=ss, caption=text, reply_markup=kb_inline)
            else:
                await message.answer(text, reply_markup=kb_inline)
        except Exception as e:
            logger.warning("Deposit display error: %s", e)


# ── VPN order callbacks ────────────────────────────────────────────

@router_admin.message(AdminFlow.menu, F.text == BTN_ADM_VPN_ORDERS)
@router_admin.message(AdminFlow.orders_submenu, F.text == BTN_ADM_VPN_ORDERS)
async def admin_vpn_orders(message: Message):
    pending = await get_pending_vpn_orders()
    if not pending:
        await message.answer("✅ No pending VPN orders.", reply_markup=admin_main_kb())
        return
    await message.answer(f"🌐 <b>{len(pending)} Pending VPN Order(s)</b>", reply_markup=admin_main_kb())
    for o in pending[:10]:
        await message.answer(fmt_admin_vpn_order(o), reply_markup=order_review_inline(o["order_id"], "vpn"))


@router_admin.callback_query(F.data.startswith("vpn_ok:"))
async def deliver_vpn(call: CallbackQuery, state: FSMContext):
    if not is_admin(call.from_user.id):
        await call.answer("🚫 Denied", show_alert=True)
        return
    oid = call.data.split(":", 1)[1]
    o   = await db_get(f"vpn_orders/{oid}")
    if not o or o.get("status") != "pending":
        await call.answer("Already processed.", show_alert=True)
        return
    _pending_fulfill[call.from_user.id] = {"oid": oid, "uid": o["user_id"], "kind": "vpn",
                                            "product_name": o["product_name"], "days": o["duration_days"]}
    await state.set_state(AdminFlow.vpn_fulfill)
    await call.answer("📝 Enter credentials now")
    await call.message.answer(
        f"📝 <b>VPN Credentials</b>\n{_SEP}\n"
        f"Order: <code>{oid}</code>\n"
        f"Product: <b>{o['product_name']}</b>  ·  {o['duration_days']} days\n\n"
        f"✏️ <b>Type the credentials to send to the user:</b>\n"
        f"<i>(e.g. username, password, server info)</i>\n\n"
        f"Press /cancel to abort.",
    )


@router_admin.callback_query(F.data.startswith("vpn_no:"))
async def cancel_vpn(call: CallbackQuery):
    if not is_admin(call.from_user.id):
        await call.answer("🚫 Denied", show_alert=True)
        return
    oid = call.data.split(":", 1)[1]
    o   = await db_get(f"vpn_orders/{oid}")
    if not o or o.get("status") != "pending":
        await call.answer("Already processed.", show_alert=True)
        return
    uid   = o.get("user_id")
    price = o.get("price", 0)
    await update_vpn_order(oid, {"status": "cancelled"})
    if uid:
        await update_balance(uid, price)
    await call.message.edit_reply_markup(reply_markup=None)
    await call.answer("❌ Cancelled & refunded")
    try:
        await call.bot.send_message(
            uid,
            f"❌ <b>VPN Order Cancelled</b>\n{_SEP}\n"
            f"📦 {o.get('product_name','?')} order was cancelled.\n"
            f"💵 Refund: <b>${price:.2f}</b> added to your balance.",
        )
    except Exception:
        pass


# ── Proxy order callbacks ──────────────────────────────────────────

@router_admin.message(AdminFlow.menu, F.text == BTN_ADM_PROXY_ORDERS)
@router_admin.message(AdminFlow.orders_submenu, F.text == BTN_ADM_PROXY_ORDERS)
async def admin_proxy_orders(message: Message):
    pending = await get_pending_proxy_orders()
    if not pending:
        await message.answer("✅ No pending proxy orders.", reply_markup=admin_main_kb())
        return
    await message.answer(f"🔐 <b>{len(pending)} Pending Proxy Order(s)</b>", reply_markup=admin_main_kb())
    for o in pending[:10]:
        await message.answer(fmt_admin_proxy_order(o), reply_markup=order_review_inline(o["order_id"], "proxy"))


@router_admin.callback_query(F.data.startswith("proxy_ok:"))
async def deliver_proxy(call: CallbackQuery, state: FSMContext):
    if not is_admin(call.from_user.id):
        await call.answer("🚫 Denied", show_alert=True)
        return
    oid = call.data.split(":", 1)[1]
    o   = await db_get(f"proxy_orders/{oid}")
    if not o or o.get("status") != "pending":
        await call.answer("Already processed.", show_alert=True)
        return
    _pending_fulfill[call.from_user.id] = {"oid": oid, "uid": o["user_id"], "kind": "proxy",
                                            "product_name": o["product_name"], "days": o["duration_days"]}
    await state.set_state(AdminFlow.proxy_fulfill)
    await call.answer("📝 Enter credentials now")
    await call.message.answer(
        f"📝 <b>Proxy Credentials</b>\n{_SEP}\n"
        f"Order: <code>{oid}</code>\n"
        f"Product: <b>{o['product_name']}</b>  ·  📡 {o.get('duration_days', '—')}\n\n"
        f"✏️ <b>Type the proxy details to send to the user:</b>\n"
        f"<i>(e.g. IP:Port, username, password)</i>\n\n"
        f"Press /cancel to abort.",
    )


@router_admin.callback_query(F.data.startswith("proxy_no:"))
async def cancel_proxy(call: CallbackQuery):
    if not is_admin(call.from_user.id):
        await call.answer("🚫 Denied", show_alert=True)
        return
    oid = call.data.split(":", 1)[1]
    o   = await db_get(f"proxy_orders/{oid}")
    if not o or o.get("status") != "pending":
        await call.answer("Already processed.", show_alert=True)
        return
    uid   = o.get("user_id")
    price = o.get("price", 0)
    await update_proxy_order(oid, {"status": "cancelled"})
    if uid:
        await update_balance(uid, price)
    await call.message.edit_reply_markup(reply_markup=None)
    await call.answer("❌ Cancelled & refunded")
    try:
        await call.bot.send_message(
            uid,
            f"❌ <b>Proxy Order Cancelled</b>\n{_SEP}\n"
            f"📦 {o.get('product_name','?')} order was cancelled.\n"
            f"💵 Refund: <b>${price:.2f}</b> added to your balance.",
        )
    except Exception:
        pass


# ── VPN Credentials Fulfillment ────────────────────────────────────

@router_admin.message(AdminFlow.vpn_fulfill)
async def vpn_credentials_received(message: Message, state: FSMContext):
    if message.text in (BACK_BTN, HOME_BTN) or (message.text and message.text.lower() == "/cancel"):
        _pending_fulfill.pop(message.from_user.id, None)
        await state.set_state(AdminFlow.menu)
        await message.answer("❌ Cancelled.", reply_markup=admin_main_kb())
        return
    if not message.text:
        await message.answer("❌ Please send the credentials as <b>text</b> (no files or photos).")
        return
    ctx = _pending_fulfill.pop(message.from_user.id, None)
    if not ctx:
        await state.set_state(AdminFlow.menu)
        await message.answer("❌ Session expired.", reply_markup=admin_main_kb())
        return
    await update_vpn_order(ctx["oid"], {"status": "delivered"})
    await state.set_state(AdminFlow.menu)
    await message.answer(
        "✅ <b>Delivered!</b> Credentials sent to user.",
        reply_markup=admin_main_kb(),
    )
    try:
        await message.bot.send_message(
            ctx["uid"],
            f"✅ <b>VPN Order Delivered!</b>\n{_SEP}\n"
            f"📦 <b>{ctx['product_name']}</b>  ·  {ctx['days']} day(s)\n\n"
            f"🔑 <b>Your Credentials:</b>\n<pre>{message.text}</pre>",
        )
    except Exception:
        pass


# ── Proxy Credentials Fulfillment ──────────────────────────────────

@router_admin.message(AdminFlow.proxy_fulfill)
async def proxy_credentials_received(message: Message, state: FSMContext):
    if message.text in (BACK_BTN, HOME_BTN) or (message.text and message.text.lower() == "/cancel"):
        _pending_fulfill.pop(message.from_user.id, None)
        await state.set_state(AdminFlow.menu)
        await message.answer("❌ Cancelled.", reply_markup=admin_main_kb())
        return
    if not message.text:
        await message.answer("❌ Please send the proxy details as <b>text</b> (no files or photos).")
        return
    ctx = _pending_fulfill.pop(message.from_user.id, None)
    if not ctx:
        await state.set_state(AdminFlow.menu)
        await message.answer("❌ Session expired.", reply_markup=admin_main_kb())
        return
    await update_proxy_order(ctx["oid"], {"status": "delivered"})
    await state.set_state(AdminFlow.menu)
    await message.answer(
        "✅ <b>Delivered!</b> Proxy details sent to user.",
        reply_markup=admin_main_kb(),
    )
    try:
        await message.bot.send_message(
            ctx["uid"],
            f"✅ <b>Proxy Order Delivered!</b>\n{_SEP}\n"
            f"📦 <b>{ctx['product_name']}</b>  ·  📡 {ctx['days']}\n\n"
            f"🔑 <b>Your Proxy Details:</b>\n<pre>{message.text}</pre>",
        )
    except Exception:
        pass


# ── Deposit Reject Reason ───────────────────────────────────────────

@router_admin.message(AdminFlow.dep_reject_reason)
async def deposit_reject_reason_received(message: Message, state: FSMContext):
    if message.text in (BACK_BTN, HOME_BTN) or (message.text and message.text.lower() == "/cancel"):
        _pending_reject.pop(message.from_user.id, None)
        await state.set_state(AdminFlow.menu)
        await message.answer("❌ Cancelled.", reply_markup=admin_main_kb())
        return
    ctx = _pending_reject.pop(message.from_user.id, None)
    if not ctx:
        await state.set_state(AdminFlow.menu)
        await message.answer("❌ Session expired.", reply_markup=admin_main_kb())
        return
    dep_id = ctx["dep_id"]
    dep    = ctx["dep"]
    reason = message.text or "No reason given."
    await update_deposit(dep_id, {"status": "rejected", "reject_reason": reason})
    await log_admin_action(message.from_user.id, "reject_deposit", f"ID:{dep_id} User:{dep['user_id']} Reason:{reason[:50]}")
    orig_msg = ctx["msg"]
    try:
        await orig_msg.edit_caption(
            (orig_msg.caption or "") + f"\n\n❌ REJECTED\nReason: {reason}", reply_markup=None
        )
    except Exception:
        try:
            await orig_msg.edit_text(
                (orig_msg.text or "") + f"\n\n❌ REJECTED\nReason: {reason}", reply_markup=None
            )
        except Exception:
            pass
    await state.set_state(AdminFlow.menu)
    await message.answer("❌ Deposit rejected.", reply_markup=admin_main_kb())
    try:
        await message.bot.send_message(
            dep["user_id"],
            f"❌ <b>Deposit Rejected</b>\n{_SEP}\n"
            f"Amount: <b>${dep['amount_usd']:.2f}</b> was not credited.\n\n"
            f"📋 <b>Reason:</b> {reason}\n\n"
            f"Contact support if you think this is an error.",
        )
    except Exception:
        pass


# ── Products ───────────────────────────────────────────────────────

@router_admin.message(AdminFlow.menu, F.text == BTN_ADM_PRODUCTS)
async def admin_products(message: Message, state: FSMContext):
    products = await get_all_products()
    dm = build_admin_products_dm(products)
    await state.update_data(products=products, dm=dm)
    await state.set_state(AdminFlow.products_list)
    await message.answer(
        f"📦 <b>Products</b>\n{_SEP}\n{len(products)} product(s) total.\nSelect one to manage:",
        reply_markup=admin_products_kb(products),
    )


@router_admin.message(AdminFlow.products_list)
async def admin_products_action(message: Message, state: FSMContext):
    if message.text in (BACK_BTN, HOME_BTN):
        await state.set_state(AdminFlow.menu)
        await message.answer("🔐 <b>Admin Panel</b>", reply_markup=admin_main_kb())
        return
    if message.text == BTN_ADD_PRODUCT:
        await state.set_state(AdminFlow.product_add_name)
        await message.answer("Enter product name:", reply_markup=input_kb())
        return
    if message.text == BTN_DOWNLOAD_PRODUCTS:
        products = await get_all_products()
        if not products:
            await message.answer("❌ No products found.")
            return
        all_stocks = {}
        for pid in products:
            all_stocks[pid] = await db_get(f"stocks/{pid}") or {}
        xlsx_bytes = make_products_xlsx(products, all_stocks)
        date_str   = time.strftime("%Y%m%d_%H%M", time.gmtime())
        total_items = sum(len(v) for v in all_stocks.values())
        await message.answer_document(
            document=BufferedInputFile(xlsx_bytes, filename=f"products_{date_str}.xlsx"),
            caption=(
                f"📦 <b>All Products</b>\n{_SEP}\n"
                f"Total: <b>{len(products)}</b> product(s)\n"
                f"📋 Stock Items: <b>{total_items}</b>\n\n"
                f"<i>Edit this file and upload it via 📤 Import Products to add new products in bulk.</i>"
            ),
        )
        return
    if message.text == BTN_IMPORT_PRODUCTS:
        await state.set_state(AdminFlow.products_import)
        await message.answer(
            f"📤 <b>Import Products</b>\n{_SEP}\n"
            f"Send an <b>.xlsx</b> file to import products in bulk.\n\n"
            f"<b>Required columns:</b>\n"
            f"• <b>Name</b> — product name\n"
            f"• <b>Emoji</b> — icon (e.g. 📮 🌐 🔐)\n"
            f"• <b>Category</b> — mail / vpn / proxy\n"
            f"• <b>Price (USD)</b> — price per unit\n"
            f"• <b>Delivery Mode</b> — manual / auto (proxy only)\n"
            f"• <b>Hidden</b> — yes / no\n\n"
            f"💡 Tip: Download the current products file first to use it as a template.",
            reply_markup=_kb([CANCEL_BTN, HOME_BTN]),
        )
        return
    data = await state.get_data()
    pid  = data.get("dm", {}).get(message.text)
    if not pid:
        await message.answer("❌ Select a product from the keyboard.")
        return
    product = await get_product(pid)
    await state.update_data(pid=pid, product=product)
    await state.set_state(AdminFlow.product_detail)
    hidden        = product.get("hidden", False)
    category      = product.get("category", "mail")
    delivery_mode = product.get("delivery_mode", "manual")
    desc          = product.get("description", "").strip()
    mode_label = ""
    if category == "proxy":
        mode_label = f"\n🚀 Delivery: <b>{'🤖 Auto (Stock)' if delivery_mode == 'auto' else '👤 Manual (Admin)'}</b>"
    desc_line = f"\n📝 <i>{desc}</i>" if desc else "\n📝 <i>No description</i>"
    await message.answer(
        f"{product.get('emoji','📦')} <b>{product['name']}</b>\n{_SEP}\n"
        f"💵 Price: ${product['price']:.2f}\n"
        f"📦 Stock: {product.get('stock_count',0)}\n"
        f"🏷 Category: {category.upper()}{mode_label}\n"
        f"👁 Visible: {'No 🙈' if hidden else 'Yes 👁'}\n"
        f"🛍 Sold: {product.get('total_sold',0)}"
        f"{desc_line}",
        reply_markup=admin_product_actions_kb(hidden, category, delivery_mode),
    )


@router_admin.message(AdminFlow.product_detail)
async def admin_product_detail_action(message: Message, state: FSMContext):
    data    = await state.get_data()
    pid     = data.get("pid")
    product = data.get("product", {})
    if message.text in (BACK_BTN, HOME_BTN):
        await _back_to_products(message, state)
        return
    if message.text in (BTN_HIDE_PROD, BTN_SHOW_PROD):
        new_h   = not product.get("hidden", False)
        await update_product(pid, {"hidden": new_h})
        updated = await get_product(pid)
        await state.update_data(product=updated)
        await message.answer(
            f"✅ Product is now {'hidden 🙈' if new_h else 'visible 👁'}.",
            reply_markup=admin_product_actions_kb(
                updated.get("hidden", False),
                updated.get("category", "mail"),
                updated.get("delivery_mode", "manual"),
            ),
        )
        return
    if message.text in (BTN_TOGGLE_AUTO, BTN_TOGGLE_MANUAL):
        if product.get("category", "mail") != "proxy":
            await message.answer("❌ Only proxy products have a delivery mode.")
            return
        new_mode = "auto" if message.text == BTN_TOGGLE_AUTO else "manual"
        await update_product(pid, {"delivery_mode": new_mode})
        updated = await get_product(pid)
        await state.update_data(product=updated)
        label = "🤖 Auto (Stock)" if new_mode == "auto" else "👤 Manual (Admin)"
        await message.answer(
            f"✅ Delivery mode → <b>{label}</b>",
            reply_markup=admin_product_actions_kb(updated.get("hidden", False), "proxy", new_mode),
        )
        return
    if message.text == BTN_ADD_STOCK:
        cat  = product.get("category", "mail")
        dmode = product.get("delivery_mode", "manual")
        if not (cat == "mail" or (cat == "proxy" and dmode == "auto")):
            await message.answer("❌ Stock management is only for Mail or Auto-Proxy products.")
            return
        stock = await get_stock_count(pid)
        settings = await get_settings()
        warn = "\n⚠️ <b>Low Stock Alert!</b>" if stock <= settings.get("low_stock_threshold", 5) else ""
        await state.update_data(stock_pid=pid, stock_product=product)
        await state.set_state(AdminFlow.stock_detail)
        await message.answer(
            f"📥 <b>{product['name']}</b>\n{_SEP}\n"
            f"📦 Current Stock: <b>{stock}</b>{warn}\n\nChoose an action:",
            reply_markup=admin_stock_actions_kb(),
        )
        return
    if message.text == BTN_DELETE:
        await delete_product(pid)
        await message.answer("🗑 Deleted.")
        await _back_to_products(message, state)
        return
    field_map = {
        BTN_EDIT_NAME:  "name",
        BTN_EDIT_PRICE: "price",
        BTN_EDIT_EMOJI: "emoji",
        BTN_EDIT_DESC:  "description",
    }
    field = field_map.get(message.text)
    if field:
        await state.update_data(edit_field=field)
        await state.set_state(AdminFlow.product_edit)
        prompts = {
            "name":        "Enter new product name:",
            "price":       "Enter new price (USD):",
            "emoji":       "Enter new emoji:",
            "description": "Enter new description (or send — to remove it):",
        }
        await message.answer(prompts[field], reply_markup=input_kb())
        return
    await message.answer("❌ Select from keyboard.")


@router_admin.message(AdminFlow.product_edit)
async def receive_product_edit(message: Message, state: FSMContext):
    if message.text in (CANCEL_BTN, HOME_BTN):
        await _back_to_products(message, state)
        return
    data  = await state.get_data()
    field = data.get("edit_field", "name")
    pid   = data.get("pid")
    value = message.text.strip()
    if field == "price":
        price, err = validate_price(value)
        if err:
            await message.answer(err)
            return
        await update_product(pid, {"price": price})
    elif field == "description":
        # "—" or "-" means remove description
        new_desc = "" if value in ("—", "-") else value
        await update_product(pid, {"description": new_desc})
    else:
        await update_product(pid, {field: value})
    updated = await get_product(pid)
    await state.update_data(product=updated)
    await state.set_state(AdminFlow.product_detail)
    field_label = "Description" if field == "description" else field.capitalize()
    await message.answer(
        f"✅ <b>{field_label} updated!</b>\n{updated.get('emoji','📦')} {updated['name']} — ${updated['price']:.2f}",
        reply_markup=admin_product_actions_kb(
            updated.get("hidden", False),
            updated.get("category", "mail"),
            updated.get("delivery_mode", "manual"),
        ),
    )


@router_admin.message(AdminFlow.products_import)
async def admin_products_import(message: Message, state: FSMContext):
    if message.text in (CANCEL_BTN, HOME_BTN):
        await _back_to_products(message, state)
        return
    if not message.document:
        await message.answer(
            "❌ Please send an <b>.xlsx</b> file.\n\nPress ❌ Cancel to go back.",
            reply_markup=_kb([CANCEL_BTN, HOME_BTN]),
        )
        return
    fname = message.document.file_name or ""
    if not fname.lower().endswith(".xlsx"):
        await message.answer(
            f"❌ Unsupported file type: <code>{fname}</code>\nOnly <b>.xlsx</b> files are supported.\n\nPress ❌ Cancel to go back.",
            reply_markup=_kb([CANCEL_BTN, HOME_BTN]),
        )
        return
    try:
        file       = await message.bot.get_file(message.document.file_id)
        file_bytes = (await message.bot.download_file(file.file_path)).read()
    except Exception as e:
        await message.answer(f"❌ Failed to download file: {e}", reply_markup=_kb([CANCEL_BTN, HOME_BTN]))
        return
    parsed = parse_products_xlsx(file_bytes)
    if not parsed:
        await message.answer(
            "❌ No valid products found in the file.\n\n"
            "Make sure the file has <b>Name</b> and <b>Price (USD)</b> columns.\n\n"
            "Press ❌ Cancel to go back.",
            reply_markup=_kb([CANCEL_BTN, HOME_BTN]),
        )
        return
    created = 0
    skipped = 0
    stock_added = 0
    for p in parsed:
        try:
            pid = await create_product(p["name"], p["price"], p["emoji"], p["category"],
                                  p.get("delivery_mode", "manual"), p.get("hidden", False))
            created += 1
            # Import stock items if present
            if p.get("stock_items"):
                added = await add_stock_items(pid, p["stock_items"], bot=message.bot)
                stock_added += added
        except Exception:
            skipped += 1
    await _back_to_products(message, state)
    lines = [f"  • {p['emoji']} {p['name']} [{p['category'].upper()}] — ${p['price']:.2f}" for p in parsed[:10]]
    preview = "\n".join(lines) + (f"\n  <i>... and {len(parsed) - 10} more</i>" if len(parsed) > 10 else "")
    stock_line = f"\n📋 Stock imported: <b>{stock_added}</b> item(s)" if stock_added > 0 else ""
    await message.answer(
        f"✅ <b>Import Complete!</b>\n{_SEP}\n"
        f"✅ Created: <b>{created}</b> product(s)\n"
        f"❌ Skipped: <b>{skipped}</b>{stock_line}\n\n"
        f"<b>Imported products:</b>\n{preview}\n\n"
        f"<i>ℹ️ Note: Import always creates new products. It does not update existing ones.</i>",
    )


@router_admin.message(AdminFlow.product_add_name)
async def add_product_name(message: Message, state: FSMContext):
    if message.text in (CANCEL_BTN, HOME_BTN):
        await _back_to_products(message, state)
        return
    await state.update_data(pname=message.text.strip())
    await state.set_state(AdminFlow.product_add_emoji)
    await message.answer("Enter emoji (e.g. 📮 🌐 🔐):", reply_markup=input_kb())


@router_admin.message(AdminFlow.product_add_emoji)
async def add_product_emoji(message: Message, state: FSMContext):
    if message.text in (CANCEL_BTN, HOME_BTN):
        await _back_to_products(message, state)
        return
    await state.update_data(pemoji=message.text.strip())
    await state.set_state(AdminFlow.product_add_price)
    await message.answer("Enter price in USD:", reply_markup=input_kb())


@router_admin.message(AdminFlow.product_add_price)
async def add_product_price(message: Message, state: FSMContext):
    if message.text in (CANCEL_BTN, HOME_BTN):
        await _back_to_products(message, state)
        return
    price, err = validate_price(message.text)
    if err:
        await message.answer(err)
        return
    await state.update_data(pprice=price)
    await state.set_state(AdminFlow.product_add_cat)
    await message.answer("Select category:", reply_markup=admin_category_kb())


@router_admin.message(AdminFlow.product_add_cat)
async def add_product_category(message: Message, state: FSMContext):
    if message.text in (CANCEL_BTN, HOME_BTN):
        await _back_to_products(message, state)
        return
    cat = CATEGORY_MAP.get(message.text)
    if not cat:
        await message.answer("❌ Select from keyboard.", reply_markup=admin_category_kb())
        return
    await state.update_data(pcat=cat)
    await state.set_state(AdminFlow.product_add_desc)
    await message.answer(
        "📝 <b>Product Description</b> (optional)\n\n"
        "Enter a short description shown to users when they view this product.\n"
        "<i>Example: Fast residential proxies, 99% uptime guarantee</i>\n\n"
        "Or press <b>⏭ Skip</b> to leave it blank.",
        reply_markup=_kb(["⏭ Skip"], [CANCEL_BTN]),
    )


@router_admin.message(AdminFlow.product_add_desc)
async def add_product_desc(message: Message, state: FSMContext):
    if message.text in (CANCEL_BTN, HOME_BTN):
        await _back_to_products(message, state)
        return
    data = await state.get_data()
    desc = "" if message.text.strip() in ("⏭ Skip", "skip", "-") else message.text.strip()
    await create_product(
        data.get("pname", ""), data.get("pprice", 0.0),
        data.get("pemoji", "📦"), data.get("pcat", "mail"),
        description=desc,
    )
    desc_preview = f"\n📝 {desc}" if desc else ""
    await message.answer(
        f"✅ <b>Product Created!</b>\n{_SEP}\n"
        f"{data.get('pemoji','📦')} {data.get('pname','')} — ${data.get('pprice',0):.2f} [{data.get('pcat','mail').upper()}]"
        f"{desc_preview}"
    )
    await _back_to_products(message, state)


async def _back_to_products(message: Message, state: FSMContext):
    products = await get_all_products()
    dm = build_admin_products_dm(products)
    await state.update_data(products=products, dm=dm)
    await state.set_state(AdminFlow.products_list)
    await message.answer("📦 <b>Products</b>", reply_markup=admin_products_kb(products))


# ── Stock ──────────────────────────────────────────────────────────

@router_admin.message(AdminFlow.menu, F.text == BTN_ADM_STOCK)
async def admin_stock(message: Message, state: FSMContext):
    products = await get_all_products()
    stockable = {
        k: v for k, v in products.items()
        if v.get("category") == "mail"
        or (v.get("category") == "proxy" and v.get("delivery_mode") == "auto")
    }
    if not stockable:
        await message.answer(
            "📥 No stockable products yet.\nCreate Mail products or Auto-Proxy products first.",
            reply_markup=admin_main_kb(),
        )
        return
    dm = build_admin_stock_dm(stockable)
    await state.update_data(stock_products=stockable, stock_dm=dm)
    await state.set_state(AdminFlow.stock_list)
    await message.answer(
        f"📥 <b>Stock Management</b>\n{_SEP}\n"
        f"📮 Mail & 🤖 Auto-Proxy products:\nSelect a product:",
        reply_markup=admin_stock_products_kb(stockable),
    )


@router_admin.message(AdminFlow.stock_list)
async def admin_stock_list_action(message: Message, state: FSMContext):
    if message.text in (BACK_BTN, HOME_BTN):
        await state.set_state(AdminFlow.menu)
        await message.answer("🔐 <b>Admin Panel</b>", reply_markup=admin_main_kb())
        return
    if message.text == BTN_DOWNLOAD_ALL_STOCK:
        data     = await state.get_data()
        products = data.get("stock_products", {})
        if not products:
            await message.answer("❌ No stockable products found.", reply_markup=admin_main_kb())
            return
        await message.answer("⏳ Exporting all stock, please wait…")
        try:
            all_stocks = {}
            for pid_key in products:
                all_stocks[pid_key] = await db_get(f"stocks/{pid_key}") or {}
            total_items = sum(len(v) for v in all_stocks.values())
            xlsx_bytes  = make_all_stock_export_xlsx(products, all_stocks)
            date_str    = time.strftime("%Y%m%d_%H%M", time.gmtime())
            await message.answer_document(
                document=BufferedInputFile(xlsx_bytes, filename=f"all_stock_{date_str}.xlsx"),
                caption=(
                    f"📦 <b>All Stock Export</b>\n{_SEP}\n"
                    f"📂 Products: <b>{len(products)}</b>\n"
                    f"📋 Total Items: <b>{total_items}</b>\n\n"
                    f"<i>Each product has its own sheet. Column D (Raw) is the original account string for re-importing.</i>"
                ),
            )
        except Exception as e:
            logger.error("All-stock export error: %s", e)
            await message.answer(f"❌ Export failed: {e}")
        return
    if message.text == BTN_IMPORT_ALL_STOCK:
        await state.set_state(AdminFlow.stock_import_all)
        await message.answer(
            f"📤 <b>Import All Stock</b>\n{_SEP}\n"
            f"Send a <b>.xlsx</b> file exported by <b>Download All Stock</b>.\n\n"
            f"<i>Each sheet name will be matched to a product. "
            f"Column D (Raw) is used for import.</i>",
            reply_markup=input_kb(),
        )
        return
    data = await state.get_data()
    pid  = data.get("stock_dm", {}).get(message.text)
    if not pid:
        await message.answer("❌ Select from keyboard.")
        return
    product  = await get_product(pid)
    stock    = await get_stock_count(pid)
    settings = await get_settings()
    warn     = "\n⚠️ <b>Low Stock Alert!</b>" if stock <= settings.get("low_stock_threshold", 5) else ""
    await state.update_data(stock_pid=pid, stock_product=product)
    await state.set_state(AdminFlow.stock_detail)
    await message.answer(
        f"📥 <b>{product['name']}</b>\n{_SEP}\n"
        f"📦 Current Stock: <b>{stock}</b>{warn}\n\nChoose an action:",
        reply_markup=admin_stock_actions_kb(),
    )


@router_admin.message(AdminFlow.stock_detail)
async def admin_stock_detail_action(message: Message, state: FSMContext):
    data    = await state.get_data()
    pid     = data.get("stock_pid")
    product = data.get("stock_product", {})
    if message.text == BACK_BTN:
        mail_p = data.get("stock_products", {})
        await state.set_state(AdminFlow.stock_list)
        await message.answer("📥 <b>Stock Management</b>", reply_markup=admin_stock_products_kb(mail_p))
        return
    if message.text == HOME_BTN:
        await state.set_state(AdminFlow.menu)
        await message.answer("🔐 <b>Admin Panel</b>", reply_markup=admin_main_kb())
        return
    if message.text == BTN_GET_TEMPLATE:
        try:
            tmpl = make_stock_template_xlsx(product["name"])
            safe_name = product["name"].replace(" ", "_")[:20]
            await message.answer_document(
                document=BufferedInputFile(tmpl, filename=f"stock_template_{safe_name}.xlsx"),
                caption=(
                    f"📋 <b>Stock Template — {product['name']}</b>\n{_SEP}\n"
                    f"1️⃣  Open this file in Excel / Google Sheets\n"
                    f"2️⃣  Delete the 3 example rows (green)\n"
                    f"3️⃣  Fill in real accounts: Email in <b>column A</b>, Password in <b>column B</b>\n"
                    f"4️⃣  Save and upload back with <b>📤 Upload File</b>"
                ),
                reply_markup=admin_stock_actions_kb(),
            )
        except Exception as _te:
            logger.warning("Template send failed: %s", _te)
            await message.answer("❌ Failed to generate template.", reply_markup=admin_stock_actions_kb())
        return
    if message.text == BTN_DOWNLOAD_STOCK:
        stock_data = await db_get(f"stocks/{pid}") or {}
        if not stock_data:
            await message.answer(
                f"📭 <b>No stock found</b>\n{_SEP}\n"
                f"Product <b>{product['name']}</b> has no stock items yet.",
                reply_markup=admin_stock_actions_kb(),
            )
            return
        try:
            xlsx_bytes = make_stock_export_xlsx(product["name"], stock_data)
            safe_name  = re.sub(r"[^\w]", "_", product["name"])[:20]
            date_str   = time.strftime("%Y%m%d_%H%M", time.gmtime())
            await message.answer_document(
                document=BufferedInputFile(xlsx_bytes, filename=f"stock_{safe_name}_{date_str}.xlsx"),
                caption=(
                    f"📥 <b>Stock Export — {product['name']}</b>\n{_SEP}\n"
                    f"📦 Total Items: <b>{len(stock_data)}</b>\n\n"
                    f"<i>Column D (Raw) is the original account string — use it when re-importing via 📤 Upload File.</i>"
                ),
                reply_markup=admin_stock_actions_kb(),
            )
        except Exception as e:
            logger.error("Stock export error: %s", e)
            await message.answer(f"❌ Failed to generate stock file: {e}", reply_markup=admin_stock_actions_kb())
        return
    if message.text == BTN_UPLOAD_FILE:
        await state.set_state(AdminFlow.stock_uploading)
        await message.answer(
            f"📤 <b>Upload Stock</b>\n{_SEP}\n"
            f"Product: <b>{product['name']}</b>\n\n"
            f"Send a <b>.txt</b> / <b>.csv</b> / <b>.xlsx</b> file\n"
            f"Format: <code>email:password</code> (one per row/line)\n\n"
            f"<i>For xlsx: put email in column A, password in column B</i>",
            reply_markup=input_kb(),
        )
        return
    if message.text == BTN_MANUAL_ADD:
        await state.set_state(AdminFlow.stock_manual)
        await message.answer(
            f"✏️ <b>Manual Add</b>\n{_SEP}\n"
            f"Product: <b>{product['name']}</b>\n\n"
            f"Type accounts one per line:\n<code>email@gmail.com:password</code>",
            reply_markup=input_kb(),
        )
        return
    if message.text == BTN_CLEAR_STOCK:
        await clear_stock(pid)
        await message.answer(
            f"🗑 Stock cleared for <b>{product['name']}</b>.",
            reply_markup=admin_stock_actions_kb(),
        )
        return
    await message.answer("❌ Select from keyboard.")


@router_admin.message(AdminFlow.stock_uploading, F.document)
async def receive_stock_file(message: Message, state: FSMContext):
    data    = await state.get_data()
    pid     = data.get("stock_pid")
    product = data.get("stock_product", {})
    doc: Document = message.document

    # Check file type before downloading
    fname = (doc.file_name or "file.txt").lower()
    if not (fname.endswith(".txt") or fname.endswith(".csv") or fname.endswith(".xlsx")):
        await message.answer(
            f"❌ Unsupported file type: <b>{doc.file_name}</b>\n"
            f"Please send a <b>.txt</b>, <b>.csv</b>, or <b>.xlsx</b> file.",
            reply_markup=admin_stock_actions_kb(),
        )
        await state.set_state(AdminFlow.stock_detail)
        return

    await message.answer("⏳ Processing file…")
    try:
        file = await message.bot.get_file(doc.file_id)
        raw  = (await message.bot.download_file(file.file_path)).read()
    except Exception as e:
        await message.answer(f"❌ Failed to download file: {e}", reply_markup=admin_stock_actions_kb())
        await state.set_state(AdminFlow.stock_detail)
        return

    try:
        items = parse_stock_file(raw, doc.file_name or "file.txt")
    except Exception as e:
        await message.answer(f"❌ Parse error: {e}", reply_markup=admin_stock_actions_kb())
        await state.set_state(AdminFlow.stock_detail)
        return

    if not items:
        await message.answer(
            "❌ No valid accounts found in the file.\n"
            "Make sure the file has email:password format.",
            reply_markup=admin_stock_actions_kb(),
        )
        await state.set_state(AdminFlow.stock_detail)
        return

    added     = await add_stock_items(pid, items, bot=message.bot)
    new_count = await get_stock_count(pid)
    await log_admin_action(message.from_user.id, "add_stock", f"Product:{product.get('name', pid)} +{added} items")
    await state.set_state(AdminFlow.stock_detail)
    await message.answer(
        f"✅ <b>{added} accounts added</b> ({len(items)} lines read)\n{_SEP}\n"
        f"Product: <b>{product['name']}</b>\n📦 Total Stock: <b>{new_count}</b>",
        reply_markup=admin_stock_actions_kb(),
    )


@router_admin.message(AdminFlow.stock_uploading)
async def stock_not_file(message: Message, state: FSMContext):
    if message.text in (CANCEL_BTN, BACK_BTN, HOME_BTN):
        await state.set_state(AdminFlow.stock_detail)
        await message.answer("Cancelled.", reply_markup=admin_stock_actions_kb())
        return
    await message.answer("❌ Please send a <b>.txt</b>, <b>.csv</b> or <b>.xlsx</b> file — not a text message.")


@router_admin.message(AdminFlow.stock_import_all, F.document)
async def receive_import_all_stock_file(message: Message, state: FSMContext):
    doc = message.document
    fname = (doc.file_name or "").lower()
    if not fname.endswith(".xlsx"):
        await message.answer(
            "❌ Please send a <b>.xlsx</b> file (the format exported by Download All Stock).",
            reply_markup=input_kb(),
        )
        return

    await message.answer("⏳ Importing stock from all sheets, please wait...")
    try:
        file = await message.bot.get_file(doc.file_id)
        raw  = (await message.bot.download_file(file.file_path)).read()
    except Exception as e:
        data = await state.get_data()
        products = data.get("stock_products", {})
        await state.set_state(AdminFlow.stock_list)
        await message.answer(f"❌ Failed to download file: {e}", reply_markup=admin_stock_products_kb(products))
        return

    try:
        wb = openpyxl.load_workbook(io.BytesIO(raw), read_only=True, data_only=True)
    except Exception as e:
        data = await state.get_data()
        products = data.get("stock_products", {})
        await state.set_state(AdminFlow.stock_list)
        await message.answer(f"❌ Cannot read xlsx file: {e}", reply_markup=admin_stock_products_kb(products))
        return

    data = await state.get_data()
    products = data.get("stock_products", {})

    # Build lookup: normalized product name (truncated to 28 chars) -> pid
    name_to_pid = {}
    for pid, p in products.items():
        pname = p.get("name", "")
        safe_title = re.sub(r"[\\/*?:\[\]]", "", pname)[:28].lower().strip()
        name_to_pid[safe_title] = pid

    results = []
    total_imported = 0
    matched_sheets = 0

    for sheet_name in wb.sheetnames:
        if sheet_name.lower() == "summary":
            continue

        # Match sheet name to product (case-insensitive, trimmed to 28 chars)
        lookup_key = sheet_name.lower().strip()
        pid = name_to_pid.get(lookup_key)
        if not pid:
            # Try partial match: sheet name might be truncated
            for norm_name, candidate_pid in name_to_pid.items():
                if norm_name.startswith(lookup_key) or lookup_key.startswith(norm_name):
                    pid = candidate_pid
                    break

        if not pid:
            results.append(f"⚠️ <b>{sheet_name}</b>: no matching product found")
            continue

        ws = wb[sheet_name]
        items = []
        first_row = True
        for row in ws.iter_rows(values_only=True):
            if first_row:
                first_row = False
                continue  # Skip header row
            # Prefer column D (Raw), fallback to columns B:C (Email:Password)
            cols = list(row)
            raw_val = cols[3] if len(cols) > 3 else None
            if raw_val and str(raw_val).strip() and str(raw_val).strip().lower() != "none":
                items.append(str(raw_val).strip())
            elif len(cols) >= 3:
                email = str(cols[1]).strip() if cols[1] else ""
                password = str(cols[2]).strip() if cols[2] else ""
                if email and email.lower() != "none":
                    if password and password.lower() != "none":
                        items.append(f"{email}:{password}")
                    else:
                        items.append(email)

        if items:
            added = await add_stock_items(pid, items, bot=message.bot)
            total_imported += added
            matched_sheets += 1
            product_name = products[pid].get("name", pid)
            results.append(f"✅ <b>{product_name}</b>: {added} items imported")
        else:
            product_name = products[pid].get("name", pid)
            results.append(f"⚠️ <b>{product_name}</b>: no items found in sheet")

    wb.close()

    # Refresh products for updated stock counts
    all_products = await get_all_products()
    stockable = {pid: p for pid, p in all_products.items() if p.get("category") in ("mail", "proxy")}
    for pid in stockable:
        stockable[pid]["stock_count"] = await get_stock_count(pid)
    await state.update_data(stock_products=stockable, stock_dm=build_admin_stock_dm(stockable))

    report = "\n".join(results) if results else "No sheets processed."
    await state.set_state(AdminFlow.stock_list)
    await message.answer(
        f"📤 <b>Import All Stock - Complete</b>\n{_SEP}\n"
        f"📂 Sheets matched: <b>{matched_sheets}</b>\n"
        f"📋 Total items imported: <b>{total_imported}</b>\n\n"
        f"{report}",
        reply_markup=admin_stock_products_kb(stockable),
    )


@router_admin.message(AdminFlow.stock_import_all)
async def stock_import_all_not_file(message: Message, state: FSMContext):
    if message.text in (CANCEL_BTN, BACK_BTN, HOME_BTN):
        data = await state.get_data()
        products = data.get("stock_products", {})
        await state.set_state(AdminFlow.stock_list)
        await message.answer("Cancelled.", reply_markup=admin_stock_products_kb(products))
        return
    await message.answer("❌ Please send a <b>.xlsx</b> file exported by <b>Download All Stock</b>.")


@router_admin.message(AdminFlow.stock_manual)
async def receive_manual_stock(message: Message, state: FSMContext):
    if message.text in (CANCEL_BTN, BACK_BTN, HOME_BTN):
        await state.set_state(AdminFlow.stock_detail)
        await message.answer("Cancelled.", reply_markup=admin_stock_actions_kb())
        return
    data    = await state.get_data()
    pid     = data.get("stock_pid")
    product = data.get("stock_product", {})
    lines   = [ln.strip() for ln in message.text.splitlines() if ln.strip()]
    if not lines:
        await message.answer("❌ No valid lines found.")
        return
    added     = await add_stock_items(pid, lines, bot=message.bot)
    new_count = await get_stock_count(pid)
    await state.set_state(AdminFlow.stock_detail)
    await message.answer(
        f"✅ <b>{added} accounts added</b>\n{_SEP}\nProduct: <b>{product['name']}</b>\n📦 Total Stock: <b>{new_count}</b>",
        reply_markup=admin_stock_actions_kb(),
    )


# ── Users ──────────────────────────────────────────────────────────

@router_admin.message(AdminFlow.menu, F.text == BTN_ADM_USERS)
async def admin_users(message: Message, state: FSMContext):
    await _show_user_menu(message, state)


@router_admin.message(AdminFlow.user_menu)
async def admin_user_menu(message: Message, state: FSMContext):
    if message.text == BACK_BTN or message.text == HOME_BTN:
        await state.set_state(AdminFlow.menu)
        await message.answer("🔐 <b>Admin Panel</b>", reply_markup=admin_main_kb())
        return
    if message.text == BTN_SEARCH_USER:
        await state.set_state(AdminFlow.user_search)
        await message.answer("🔍 Enter a Telegram User ID to look up:", reply_markup=_kb([BACK_BTN, HOME_BTN]))
        return
    if message.text == BTN_BONUS_ALL_USERS:
        await state.set_state(AdminFlow.bonus_all_amount)
        await message.answer("🎁 <b>Bonus All Users</b>\n\nEnter bonus amount (USD):", reply_markup=input_kb())
        return
    if message.text == BTN_BONUS_TOP10:
        await state.set_state(AdminFlow.bonus_top10_amount)
        await message.answer("🏆 <b>Bonus Top 10 Active Users</b>\n\nEnter bonus amount (USD):", reply_markup=input_kb())
        return
    await message.answer("❌ Select from keyboard.")


@router_admin.message(AdminFlow.user_search)
async def admin_search_user(message: Message, state: FSMContext):
    if message.text in (BACK_BTN, HOME_BTN):
        await _show_user_menu(message, state)
        return
    try:
        uid = int(message.text.strip())
    except ValueError:
        await message.answer("❌ Enter a valid numeric user ID.")
        return
    user = await get_user(uid)
    if not user:
        await message.answer("❌ User not found.")
        return
    mail_orders, vpn_orders, proxy_orders = await asyncio.gather(
        get_user_orders(uid),
        get_user_vpn_orders(uid),
        get_user_proxy_orders(uid),
    )
    await state.update_data(target_uid=uid, target_user=user)
    await state.set_state(AdminFlow.user_detail)
    await message.answer(
        fmt_user_info(user, len(mail_orders), len(vpn_orders), len(proxy_orders)),
        reply_markup=admin_user_actions_kb(user.get("is_banned", False)),
    )


@router_admin.message(AdminFlow.user_detail)
async def admin_user_action(message: Message, state: FSMContext):
    data = await state.get_data()
    uid  = data.get("target_uid")
    user = data.get("target_user", {})
    if message.text == BACK_BTN:
        await _show_user_menu(message, state)
        return
    if message.text == HOME_BTN:
        await state.set_state(AdminFlow.menu)
        await message.answer("🔐 <b>Admin Panel</b>", reply_markup=admin_main_kb())
        return
    if message.text in (BTN_BAN_USER, BTN_UNBAN_USER):
        new_b   = not user.get("is_banned", False)
        await ban_user(uid, new_b)
        await log_admin_action(message.from_user.id, "ban_user" if new_b else "unban_user", f"User:{uid}")
        updated = await get_user(uid)
        await state.update_data(target_user=updated)
        await message.answer(
            f"✅ User <code>{uid}</code> {'banned 🚫' if new_b else 'unbanned ✅'}",
            reply_markup=admin_user_actions_kb(updated.get("is_banned", False)),
        )
        try:
            await message.bot.send_message(
                uid,
                "🚫 Your account has been banned." if new_b else "✅ Your account has been unbanned.",
            )
        except Exception:
            pass
        return
    if message.text == BTN_ADD_BAL:
        await state.set_state(AdminFlow.user_add_bal)
        await message.answer(f"💰 Enter USD amount to add to <code>{uid}</code>:", reply_markup=input_kb())
        return
    if message.text == BTN_REMOVE_BAL:
        await state.set_state(AdminFlow.user_remove_bal)
        await message.answer(f"💸 Enter USD amount to remove from <code>{uid}</code>:", reply_markup=input_kb())
        return
    if message.text == BTN_ADD_BONUS:
        await state.set_state(AdminFlow.user_bonus_amount)
        current_bonus = await get_bonus(uid)
        if current_bonus["amount"] > 0:
            await message.answer(
                f"⚠️ User already has a bonus: <b>${current_bonus['amount']:.2f}</b>\n"
                f"Setting a new bonus will replace it."
            )
        await message.answer(f"🎁 Enter bonus amount (USD) for <code>{uid}</code>:", reply_markup=input_kb())
        return
    if message.text == BTN_REMOVE_BONUS:
        current_bonus = await get_bonus(uid)
        if current_bonus["amount"] <= 0:
            await message.answer("❌ User has no bonus to remove.")
            return
        await db_delete(f"users/{uid}/bonus")
        updated = await get_user(uid)
        await state.update_data(target_user=updated)
        await message.answer(
            f"✅ Bonus removed from user <code>{uid}</code>.\n"
            f"Previous bonus was: <b>${current_bonus['amount']:.2f}</b>",
            reply_markup=admin_user_actions_kb(updated.get("is_banned", False)),
        )
        try:
            await message.bot.send_message(
                uid,
                f"🗑 <b>Bonus Removed</b>\n{_SEP}\n"
                f"Your bonus of <b>${current_bonus['amount']:.2f}</b> has been removed by admin.",
            )
        except Exception:
            pass
        return
    await message.answer("❌ Select from keyboard.")


@router_admin.message(AdminFlow.user_add_bal)
async def admin_add_bal(message: Message, state: FSMContext):
    if message.text in (CANCEL_BTN, BACK_BTN, HOME_BTN):
        data = await state.get_data()
        user = data.get("target_user", {})
        await state.set_state(AdminFlow.user_detail)
        await message.answer(fmt_user_info(user), reply_markup=admin_user_actions_kb(user.get("is_banned", False)))
        return
    amount, err = validate_price(message.text)
    if err:
        await message.answer(err)
        return
    data    = await state.get_data()
    uid     = data.get("target_uid")
    new_bal = await update_balance(uid, amount)
    updated = await get_user(uid)
    await state.update_data(target_user=updated)
    await state.set_state(AdminFlow.user_detail)
    await message.answer(
        f"✅ Added <b>${amount:.2f}</b> → New Balance: <b>${new_bal:.2f}</b>",
        reply_markup=admin_user_actions_kb(updated.get("is_banned", False)),
    )
    try:
        await message.bot.send_message(
            uid,
            f"💰 <b>Balance Added!</b>\n{_SEP}\n➕ ${amount:.2f}\n💰 New Balance: ${new_bal:.2f}",
        )
    except Exception:
        pass


@router_admin.message(AdminFlow.user_remove_bal)
async def admin_remove_bal(message: Message, state: FSMContext):
    if message.text in (CANCEL_BTN, BACK_BTN, HOME_BTN):
        data = await state.get_data()
        user = data.get("target_user", {})
        await state.set_state(AdminFlow.user_detail)
        await message.answer(fmt_user_info(user), reply_markup=admin_user_actions_kb(user.get("is_banned", False)))
        return
    amount, err = validate_price(message.text)
    if err:
        await message.answer(err)
        return
    data    = await state.get_data()
    uid     = data.get("target_uid")
    target_user = data.get("target_user", {})
    new_bal = await update_balance(uid, -amount)
    updated = await get_user(uid)
    await state.update_data(target_user=updated)
    await state.set_state(AdminFlow.user_detail)
    warning = ""
    if new_bal < 0:
        warning = "\n\n\u26a0\ufe0f <b>Warning:</b> Balance went negative!"
    await message.answer(
        f"\u2705 Removed <b>${amount:.2f}</b> \u2192 New Balance: <b>${new_bal:.2f}</b>{warning}",
        reply_markup=admin_user_actions_kb(updated.get("is_banned", False)),
    )


# ── Admin Bonus ────────────────────────────────────────────────────

@router_admin.message(AdminFlow.user_bonus_amount)
async def admin_bonus_amount(message: Message, state: FSMContext):
    if message.text in (CANCEL_BTN, BACK_BTN, HOME_BTN):
        data = await state.get_data()
        user = data.get("target_user", {})
        await state.set_state(AdminFlow.user_detail)
        await message.answer(fmt_user_info(user), reply_markup=admin_user_actions_kb(user.get("is_banned", False)))
        return
    amount, err = validate_price(message.text)
    if err:
        await message.answer(err)
        return
    if amount <= 0:
        await message.answer("❌ Bonus amount must be greater than 0.")
        return
    await state.update_data(bonus_amount=amount, bonus_selected_products=[])
    products = await get_all_products()
    if not products:
        await message.answer("❌ No products found. Add products first.")
        data = await state.get_data()
        user = data.get("target_user", {})
        await state.set_state(AdminFlow.user_detail)
        await message.answer(fmt_user_info(user), reply_markup=admin_user_actions_kb(user.get("is_banned", False)))
        return
    buttons = []
    for pid, prod in products.items():
        buttons.append([InlineKeyboardButton(
            text=f"{prod.get('emoji', '📦')} {prod['name']}",
            callback_data=f"bonus_prod:{pid}",
        )])
    buttons.append([InlineKeyboardButton(text="✅ Done", callback_data="bonus_done")])
    kb = InlineKeyboardMarkup(inline_keyboard=buttons)
    await state.set_state(AdminFlow.user_bonus_products)
    await message.answer(
        f"🎁 Bonus: <b>${amount:.2f}</b>\n\n"
        f"Select which products this bonus can be used on:\n"
        f"(Tap to toggle selection, then press Done)",
        reply_markup=kb,
    )


@router_admin.callback_query(F.data.startswith("bonus_prod:"))
async def admin_bonus_toggle_product(call: CallbackQuery, state: FSMContext):
    if not is_admin(call.from_user.id):
        await call.answer("Not authorized", show_alert=True)
        return
    current_state = await state.get_state()
    if current_state != AdminFlow.user_bonus_products.state:
        await call.answer()
        return
    pid = call.data.split(":", 1)[1]
    data = await state.get_data()
    selected = data.get("bonus_selected_products", [])
    if pid in selected:
        selected.remove(pid)
    else:
        selected.append(pid)
    await state.update_data(bonus_selected_products=selected)
    products = await get_all_products()
    buttons = []
    for p_id, prod in products.items():
        check = "✅ " if p_id in selected else ""
        buttons.append([InlineKeyboardButton(
            text=f"{check}{prod.get('emoji', '📦')} {prod['name']}",
            callback_data=f"bonus_prod:{p_id}",
        )])
    buttons.append([InlineKeyboardButton(text="✅ Done", callback_data="bonus_done")])
    kb = InlineKeyboardMarkup(inline_keyboard=buttons)
    await call.message.edit_reply_markup(reply_markup=kb)
    await call.answer()


@router_admin.callback_query(F.data == "bonus_done")
async def admin_bonus_done(call: CallbackQuery, state: FSMContext):
    if not is_admin(call.from_user.id):
        await call.answer("Not authorized", show_alert=True)
        return
    current_state = await state.get_state()
    if current_state != AdminFlow.user_bonus_products.state:
        await call.answer()
        return
    data = await state.get_data()
    uid = data.get("target_uid")
    amount = data.get("bonus_amount", 0)
    selected = data.get("bonus_selected_products", [])
    if not selected:
        await call.answer("Please select at least one product.", show_alert=True)
        return
    # Validate that selected product IDs still exist
    products = await get_all_products()
    valid_selected = [p for p in selected if p in products]
    if not valid_selected:
        await call.answer("Selected products no longer exist. Please re-select.", show_alert=True)
        return
    await set_bonus(uid, amount, valid_selected)
    updated = await get_user(uid)
    await state.update_data(target_user=updated)
    await state.set_state(AdminFlow.user_detail)
    prod_names = [products[p]["name"] for p in valid_selected if p in products]
    await call.message.edit_text(
        f"✅ Bonus set!\n"
        f"🎁 Amount: <b>${amount:.2f}</b>\n"
        f"📦 Products: {', '.join(prod_names)}"
    )
    await call.message.answer(
        fmt_user_info(updated),
        reply_markup=admin_user_actions_kb(updated.get("is_banned", False)),
    )
    try:
        await call.bot.send_message(
            uid,
            f"🎁 <b>Bonus Added!</b>\n{_SEP}\n"
            f"💵 Amount: <b>${amount:.2f}</b>\n"
            f"📦 Eligible Products: {', '.join(prod_names)}\n\n"
            f"Use this bonus to purchase the above products!",
        )
    except Exception:
        pass
    await call.answer()


# ── Bonus All Users ────────────────────────────────────────────────

@router_admin.message(AdminFlow.bonus_all_amount)
async def admin_bonus_all_amount(message: Message, state: FSMContext):
    if message.text in (CANCEL_BTN, BACK_BTN, HOME_BTN):
        await _show_user_menu(message, state)
        return
    amount, err = validate_price(message.text)
    if err:
        await message.answer(err)
        return
    if amount <= 0:
        await message.answer("❌ Bonus amount must be greater than 0.")
        return
    await state.update_data(bonus_all_amount=amount, bonusall_selected_products=[])
    products = await get_all_products()
    if not products:
        await message.answer("❌ No products found. Add products first.")
        await _show_user_menu(message, state)
        return
    buttons = []
    for pid, prod in products.items():
        buttons.append([InlineKeyboardButton(
            text=f"{prod.get('emoji', '📦')} {prod['name']}",
            callback_data=f"bonusall_prod:{pid}",
        )])
    buttons.append([InlineKeyboardButton(text="✅ Done", callback_data="bonusall_done")])
    kb = InlineKeyboardMarkup(inline_keyboard=buttons)
    await state.set_state(AdminFlow.bonus_all_products)
    await message.answer(
        f"🎁 Bonus All Users: <b>${amount:.2f}</b>\n\n"
        f"Select which products this bonus can be used on:\n"
        f"(Tap to toggle selection, then press Done)",
        reply_markup=kb,
    )


@router_admin.callback_query(F.data.startswith("bonusall_prod:"))
async def admin_bonusall_toggle_product(call: CallbackQuery, state: FSMContext):
    if not is_admin(call.from_user.id):
        await call.answer("Not authorized", show_alert=True)
        return
    current_state = await state.get_state()
    if current_state != AdminFlow.bonus_all_products.state:
        await call.answer()
        return
    pid = call.data.split(":", 1)[1]
    data = await state.get_data()
    selected = data.get("bonusall_selected_products", [])
    if pid in selected:
        selected.remove(pid)
    else:
        selected.append(pid)
    await state.update_data(bonusall_selected_products=selected)
    products = await get_all_products()
    buttons = []
    for p_id, prod in products.items():
        check = "✅ " if p_id in selected else ""
        buttons.append([InlineKeyboardButton(
            text=f"{check}{prod.get('emoji', '📦')} {prod['name']}",
            callback_data=f"bonusall_prod:{p_id}",
        )])
    buttons.append([InlineKeyboardButton(text="✅ Done", callback_data="bonusall_done")])
    kb = InlineKeyboardMarkup(inline_keyboard=buttons)
    await call.message.edit_reply_markup(reply_markup=kb)
    await call.answer()


@router_admin.callback_query(F.data == "bonusall_done")
async def admin_bonusall_done(call: CallbackQuery, state: FSMContext):
    if not is_admin(call.from_user.id):
        await call.answer("Not authorized", show_alert=True)
        return
    current_state = await state.get_state()
    if current_state != AdminFlow.bonus_all_products.state:
        await call.answer()
        return
    data = await state.get_data()
    amount = data.get("bonus_all_amount", 0)
    selected = data.get("bonusall_selected_products", [])
    if not selected:
        await call.answer("Please select at least one product.", show_alert=True)
        return
    products = await get_all_products()
    valid_selected = [p for p in selected if p in products]
    if not valid_selected:
        await call.answer("Selected products no longer exist. Please re-select.", show_alert=True)
        return
    await call.answer()
    all_users = await get_all_users()
    count = 0
    for uid_str in all_users:
        try:
            uid_int = int(uid_str)
            await set_bonus(uid_int, amount, valid_selected)
            count += 1
            try:
                prod_names = [products[p]["name"] for p in valid_selected if p in products]
                await call.bot.send_message(
                    uid_int,
                    f"🎁 <b>Bonus Added!</b>\n{_SEP}\n"
                    f"💵 Amount: <b>${amount:.2f}</b>\n"
                    f"📦 Eligible Products: {', '.join(prod_names)}\n\n"
                    f"Use this bonus to purchase the above products!",
                )
            except Exception:
                pass
            await asyncio.sleep(0.05)
        except Exception:
            pass
    prod_names = [products[p]["name"] for p in valid_selected if p in products]
    await call.message.edit_text(
        f"✅ Bonus set for <b>{count}</b> user(s)!\n"
        f"🎁 Amount: <b>${amount:.2f}</b>\n"
        f"📦 Products: {', '.join(prod_names)}"
    )
    await _show_user_menu(call.message, state)


# ── Bonus Top 10 ──────────────────────────────────────────────────

@router_admin.message(AdminFlow.bonus_top10_amount)
async def admin_bonus_top10_amount(message: Message, state: FSMContext):
    if message.text in (CANCEL_BTN, BACK_BTN, HOME_BTN):
        await _show_user_menu(message, state)
        return
    amount, err = validate_price(message.text)
    if err:
        await message.answer(err)
        return
    if amount <= 0:
        await message.answer("❌ Bonus amount must be greater than 0.")
        return
    await state.update_data(bonus_top10_amount=amount, bonustop_selected_products=[])
    products = await get_all_products()
    if not products:
        await message.answer("❌ No products found. Add products first.")
        await _show_user_menu(message, state)
        return
    buttons = []
    for pid, prod in products.items():
        buttons.append([InlineKeyboardButton(
            text=f"{prod.get('emoji', '📦')} {prod['name']}",
            callback_data=f"bonustop_prod:{pid}",
        )])
    buttons.append([InlineKeyboardButton(text="✅ Done", callback_data="bonustop_done")])
    kb = InlineKeyboardMarkup(inline_keyboard=buttons)
    await state.set_state(AdminFlow.bonus_top10_products)
    await message.answer(
        f"🏆 Bonus Top 10: <b>${amount:.2f}</b>\n\n"
        f"Select which products this bonus can be used on:\n"
        f"(Tap to toggle selection, then press Done)",
        reply_markup=kb,
    )


@router_admin.callback_query(F.data.startswith("bonustop_prod:"))
async def admin_bonustop_toggle_product(call: CallbackQuery, state: FSMContext):
    if not is_admin(call.from_user.id):
        await call.answer("Not authorized", show_alert=True)
        return
    current_state = await state.get_state()
    if current_state != AdminFlow.bonus_top10_products.state:
        await call.answer()
        return
    pid = call.data.split(":", 1)[1]
    data = await state.get_data()
    selected = data.get("bonustop_selected_products", [])
    if pid in selected:
        selected.remove(pid)
    else:
        selected.append(pid)
    await state.update_data(bonustop_selected_products=selected)
    products = await get_all_products()
    buttons = []
    for p_id, prod in products.items():
        check = "✅ " if p_id in selected else ""
        buttons.append([InlineKeyboardButton(
            text=f"{check}{prod.get('emoji', '📦')} {prod['name']}",
            callback_data=f"bonustop_prod:{p_id}",
        )])
    buttons.append([InlineKeyboardButton(text="✅ Done", callback_data="bonustop_done")])
    kb = InlineKeyboardMarkup(inline_keyboard=buttons)
    await call.message.edit_reply_markup(reply_markup=kb)
    await call.answer()


@router_admin.callback_query(F.data == "bonustop_done")
async def admin_bonustop_done(call: CallbackQuery, state: FSMContext):
    if not is_admin(call.from_user.id):
        await call.answer("Not authorized", show_alert=True)
        return
    current_state = await state.get_state()
    if current_state != AdminFlow.bonus_top10_products.state:
        await call.answer()
        return
    data = await state.get_data()
    amount = data.get("bonus_top10_amount", 0)
    selected = data.get("bonustop_selected_products", [])
    if not selected:
        await call.answer("Please select at least one product.", show_alert=True)
        return
    products = await get_all_products()
    valid_selected = [p for p in selected if p in products]
    if not valid_selected:
        await call.answer("Selected products no longer exist. Please re-select.", show_alert=True)
        return
    await call.answer()
    all_users = await get_all_users()
    # Sort by order_count descending, take top 10
    sorted_users = sorted(
        all_users.items(),
        key=lambda x: x[1].get("order_count", 0) if isinstance(x[1], dict) else 0,
        reverse=True,
    )[:10]
    count = 0
    for uid_str, u_data in sorted_users:
        try:
            uid_int = int(uid_str)
            await set_bonus(uid_int, amount, valid_selected)
            count += 1
            try:
                prod_names = [products[p]["name"] for p in valid_selected if p in products]
                await call.bot.send_message(
                    uid_int,
                    f"🏆 <b>Bonus Added!</b>\n{_SEP}\n"
                    f"You are one of the top 10 active users!\n"
                    f"💵 Amount: <b>${amount:.2f}</b>\n"
                    f"📦 Eligible Products: {', '.join(prod_names)}\n\n"
                    f"Use this bonus to purchase the above products!",
                )
            except Exception:
                pass
            await asyncio.sleep(0.05)
        except Exception:
            pass
    prod_names = [products[p]["name"] for p in valid_selected if p in products]
    await call.message.edit_text(
        f"✅ Bonus set for top <b>{count}</b> active user(s)!\n"
        f"🎁 Amount: <b>${amount:.2f}</b>\n"
        f"📦 Products: {', '.join(prod_names)}"
    )
    await _show_user_menu(call.message, state)


# ── Coupons ────────────────────────────────────────────────────────

@router_admin.message(AdminFlow.menu, F.text == BTN_ADM_COUPONS)
@router_admin.message(AdminFlow.tools_submenu, F.text == BTN_ADM_COUPONS)
async def admin_coupons(message: Message, state: FSMContext):
    coupons = await get_all_coupons()
    dm = build_coupons_dm(coupons)
    await state.update_data(coupons=coupons, coupons_dm=dm)
    await state.set_state(AdminFlow.coupons_list)
    await message.answer(f"🎟 <b>Coupons</b>\n{_SEP}\n{len(coupons)} coupon(s) total.", reply_markup=admin_coupons_kb(coupons))


@router_admin.message(AdminFlow.coupons_list)
async def admin_coupons_action(message: Message, state: FSMContext):
    if message.text in (BACK_BTN, HOME_BTN):
        await state.set_state(AdminFlow.menu)
        await message.answer("🔐 <b>Admin Panel</b>", reply_markup=admin_main_kb())
        return
    if message.text == BTN_CREATE_COUPON:
        await state.set_state(AdminFlow.coupon_code)
        await message.answer("Enter coupon code (alphanumeric, e.g. SAVE20):", reply_markup=input_kb())
        return
    data = await state.get_data()
    cid  = data.get("coupons_dm", {}).get(message.text)
    if not cid:
        await message.answer("❌ Select from keyboard.")
        return
    coupons = data.get("coupons", [])
    coupon  = next((c for c in coupons if c["coupon_id"] == cid), None)
    if not coupon:
        await message.answer("❌ Not found.")
        return
    expires = time.strftime("%d %b %Y", time.gmtime(coupon.get("expires_at", 0)))
    await state.update_data(selected_cid=cid)
    await state.set_state(AdminFlow.coupon_detail)
    await message.answer(
        f"🎟 <b>{coupon['code']}</b>\n{_SEP}\n"
        f"🏷 Discount: <b>{coupon['discount_pct']:.0f}%</b>\n"
        f"🔢 Used: <b>{coupon['used_count']}/{coupon['max_uses']}</b>\n"
        f"📅 Expires: <b>{expires}</b>\n"
        f"Status: {'✅ Active' if coupon.get('active') else '❌ Inactive'}",
        reply_markup=admin_coupon_actions_kb(),
    )


@router_admin.message(AdminFlow.coupon_detail)
async def admin_coupon_action(message: Message, state: FSMContext):
    data = await state.get_data()
    cid  = data.get("selected_cid")
    if message.text in (BACK_BTN, HOME_BTN):
        await _back_to_coupons(message, state)
        return
    if message.text == BTN_DELETE_COUPON:
        await delete_coupon(cid)
        await _back_to_coupons(message, state)
        return
    await message.answer("❌ Select from keyboard.")


@router_admin.message(AdminFlow.coupon_code)
async def coupon_create_code(message: Message, state: FSMContext):
    if message.text in (CANCEL_BTN, HOME_BTN):
        await _back_to_coupons(message, state)
        return
    code = message.text.strip().upper()
    if not code.isalnum():
        await message.answer("❌ Alphanumeric only (letters and numbers).")
        return
    await state.update_data(c_code=code)
    await state.set_state(AdminFlow.coupon_discount)
    await message.answer("Enter discount % (1–100):", reply_markup=input_kb())


@router_admin.message(AdminFlow.coupon_discount)
async def coupon_create_discount(message: Message, state: FSMContext):
    if message.text in (CANCEL_BTN, HOME_BTN):
        await _back_to_coupons(message, state)
        return
    pct, err = validate_discount(message.text)
    if err:
        await message.answer(err)
        return
    await state.update_data(c_pct=pct)
    await state.set_state(AdminFlow.coupon_max_uses)
    await message.answer("Enter max number of uses:", reply_markup=input_kb())


@router_admin.message(AdminFlow.coupon_max_uses)
async def coupon_create_max_uses(message: Message, state: FSMContext):
    if message.text in (CANCEL_BTN, HOME_BTN):
        await _back_to_coupons(message, state)
        return
    try:
        max_uses = int(message.text.strip())
        assert max_uses > 0
    except (ValueError, AssertionError):
        await message.answer("❌ Enter a positive integer.")
        return
    await state.update_data(c_max=max_uses)
    await state.set_state(AdminFlow.coupon_expiry)
    await message.answer("Enter validity in days (e.g. 30):", reply_markup=input_kb())


@router_admin.message(AdminFlow.coupon_expiry)
async def coupon_create_expiry(message: Message, state: FSMContext):
    if message.text in (CANCEL_BTN, HOME_BTN):
        await _back_to_coupons(message, state)
        return
    try:
        days = int(message.text.strip())
        assert days > 0
    except (ValueError, AssertionError):
        await message.answer("❌ Enter a positive number of days.")
        return
    data       = await state.get_data()
    expires_at = int(time.time()) + days * 86400
    c_code = data.get("c_code", "")
    c_pct  = data.get("c_pct", 0)
    c_max  = data.get("c_max", 1)
    await create_coupon(c_code, c_pct, c_max, expires_at)
    await message.answer(
        f"✅ <b>Coupon Created!</b>\n{_SEP}\n"
        f"Code: <code>{c_code}</code>\n"
        f"Discount: <b>{c_pct:.0f}%</b>\n"
        f"Max Uses: <b>{c_max}</b>\n"
        f"Valid for: <b>{days} days</b>",
    )
    await _back_to_coupons(message, state)


async def _back_to_coupons(message: Message, state: FSMContext):
    coupons = await get_all_coupons()
    dm = build_coupons_dm(coupons)
    await state.update_data(coupons=coupons, coupons_dm=dm)
    await state.set_state(AdminFlow.coupons_list)
    await message.answer("🎟 <b>Coupons</b>", reply_markup=admin_coupons_kb(coupons))


# ── Analytics ──────────────────────────────────────────────────────

@router_admin.message(AdminFlow.menu, F.text == BTN_ADM_ANALYTICS)
@router_admin.message(AdminFlow.reports_submenu, F.text == BTN_ADM_ANALYTICS)
async def admin_analytics(message: Message):
    data = await get_analytics_data()
    detail_kb = InlineKeyboardMarkup(inline_keyboard=[
        [InlineKeyboardButton(text="📊 Detailed Breakdown", callback_data="adm_analytics_detail")]
    ])
    await message.answer(fmt_analytics(data), reply_markup=detail_kb)


# ── Order Lookup ───────────────────────────────────────────────────

@router_admin.message(AdminFlow.menu, F.text == BTN_ADM_ORDER_LOOKUP)
@router_admin.message(AdminFlow.orders_submenu, F.text == BTN_ADM_ORDER_LOOKUP)
async def admin_order_lookup_start(message: Message, state: FSMContext):
    await state.set_state(AdminFlow.order_lookup)
    await message.answer(
        f"🔍 <b>Order Lookup</b>\n{_SEP}\n"
        f"Enter an Order ID to search:\n"
        f"<i>(e.g. ORD-0001, VPN-0001, PXY-0001)</i>",
        reply_markup=_kb([BACK_BTN, HOME_BTN]),
    )


@router_admin.message(AdminFlow.order_lookup)
async def admin_order_lookup_search(message: Message, state: FSMContext):
    if message.text in (BACK_BTN, HOME_BTN):
        await state.set_state(AdminFlow.menu)
        await message.answer("🔐 <b>Admin Panel</b>", reply_markup=admin_main_kb())
        return
    oid = (message.text or "").strip()
    if not oid:
        await message.answer("❌ Please enter a valid Order ID.")
        return
    # Validate format to prevent path injection
    if not re.match(r'^[A-Z]{2,4}-\d{3,4}$', oid):
        await message.answer(
            "❌ Invalid Order ID format.\n"
            "Use format like: ORD-0001, VPN-0001, PXY-0001"
        )
        return
    # Prefix-aware routing to reduce redundant Firebase lookups
    if oid.startswith("ORD-"):
        order = await db_get(f"orders/{oid}")
        if order:
            user = await get_user(order.get("user_id", 0)) or {}
            items_preview = ""
            items_list = order.get("items", [])
            if items_list:
                preview = "\n".join(str(i) for i in items_list[:5])
                more = f"\n<i>... +{len(items_list) - 5} more</i>" if len(items_list) > 5 else ""
                items_preview = f"\n📋 <b>Items:</b>\n<code>{preview}{more}</code>"
            await message.answer(
                f"📮 <b>Mail Order Details</b>\n{_SEP}\n"
                f"🆔 Order: <code>{oid}</code>\n"
                f"👤 User: @{user.get('username', '—')} (<code>{order.get('user_id')}</code>)\n"
                f"📦 Product: <b>{order.get('product_name', '—')}</b>\n"
                f"🔢 Qty: <b>{order.get('qty', 0)}</b>\n"
                f"💵 Total: <b>${order.get('total_price', 0):.2f}</b>\n"
                f"📊 Status: <b>{order.get('status', '—')}</b>\n"
                f"🕒 Date: {_dt(order.get('created_at', 0))}"
                f"{items_preview}",
                reply_markup=_kb([BACK_BTN, HOME_BTN]),
            )
            return
    elif oid.startswith("VPN-"):
        vpn_order = await db_get(f"vpn_orders/{oid}")
        if vpn_order:
            user = await get_user(vpn_order.get("user_id", 0)) or {}
            await message.answer(
                f"🌐 <b>VPN Order Details</b>\n{_SEP}\n"
                f"🆔 Order: <code>{oid}</code>\n"
                f"👤 User: @{user.get('username', '—')} (<code>{vpn_order.get('user_id')}</code>)\n"
                f"📦 Product: <b>{vpn_order.get('product_name', '—')}</b>\n"
                f"⏱ Duration: <b>{vpn_order.get('duration_days', 0)} day(s)</b>\n"
                f"💵 Price: <b>${vpn_order.get('price', 0):.2f}</b>\n"
                f"📊 Status: <b>{vpn_order.get('status', '—')}</b>\n"
                f"🕒 Date: {_dt(vpn_order.get('created_at', 0))}",
                reply_markup=_kb([BACK_BTN, HOME_BTN]),
            )
            return
    elif oid.startswith("PXY-"):
        proxy_order = await db_get(f"proxy_orders/{oid}")
        if proxy_order:
            user = await get_user(proxy_order.get("user_id", 0)) or {}
            items_preview = ""
            items_list = proxy_order.get("items", [])
            if items_list:
                preview = "\n".join(str(i) for i in items_list[:5])
                more = f"\n<i>... +{len(items_list) - 5} more</i>" if len(items_list) > 5 else ""
                items_preview = f"\n📋 <b>Items:</b>\n<code>{preview}{more}</code>"
            await message.answer(
                f"🔐 <b>Proxy Order Details</b>\n{_SEP}\n"
                f"🆔 Order: <code>{oid}</code>\n"
                f"👤 User: @{user.get('username', '—')} (<code>{proxy_order.get('user_id')}</code>)\n"
                f"📦 Product: <b>{proxy_order.get('product_name', '—')}</b>\n"
                f"📡 Duration: <b>{proxy_order.get('duration_days', '—')}</b>\n"
                f"💵 Price: <b>${proxy_order.get('price', 0):.2f}</b>\n"
                f"📊 Status: <b>{proxy_order.get('status', '—')}</b>\n"
                f"🕒 Date: {_dt(proxy_order.get('created_at', 0))}"
                f"{items_preview}",
                reply_markup=_kb([BACK_BTN, HOME_BTN]),
            )
            return
    else:
        # Fallback for legacy or unknown prefix: search all collections
        order = await db_get(f"orders/{oid}")
        if order:
            user = await get_user(order.get("user_id", 0)) or {}
            items_preview = ""
            items_list = order.get("items", [])
            if items_list:
                preview = "\n".join(str(i) for i in items_list[:5])
                more = f"\n<i>... +{len(items_list) - 5} more</i>" if len(items_list) > 5 else ""
                items_preview = f"\n📋 <b>Items:</b>\n<code>{preview}{more}</code>"
            await message.answer(
                f"📮 <b>Mail Order Details</b>\n{_SEP}\n"
                f"🆔 Order: <code>{oid}</code>\n"
                f"👤 User: @{user.get('username', '—')} (<code>{order.get('user_id')}</code>)\n"
                f"📦 Product: <b>{order.get('product_name', '—')}</b>\n"
                f"🔢 Qty: <b>{order.get('qty', 0)}</b>\n"
                f"💵 Total: <b>${order.get('total_price', 0):.2f}</b>\n"
                f"📊 Status: <b>{order.get('status', '—')}</b>\n"
                f"🕒 Date: {_dt(order.get('created_at', 0))}"
                f"{items_preview}",
                reply_markup=_kb([BACK_BTN, HOME_BTN]),
            )
            return
        vpn_order = await db_get(f"vpn_orders/{oid}")
        if vpn_order:
            user = await get_user(vpn_order.get("user_id", 0)) or {}
            await message.answer(
                f"🌐 <b>VPN Order Details</b>\n{_SEP}\n"
                f"🆔 Order: <code>{oid}</code>\n"
                f"👤 User: @{user.get('username', '—')} (<code>{vpn_order.get('user_id')}</code>)\n"
                f"📦 Product: <b>{vpn_order.get('product_name', '—')}</b>\n"
                f"⏱ Duration: <b>{vpn_order.get('duration_days', 0)} day(s)</b>\n"
                f"💵 Price: <b>${vpn_order.get('price', 0):.2f}</b>\n"
                f"📊 Status: <b>{vpn_order.get('status', '—')}</b>\n"
                f"🕒 Date: {_dt(vpn_order.get('created_at', 0))}",
                reply_markup=_kb([BACK_BTN, HOME_BTN]),
            )
            return
        proxy_order = await db_get(f"proxy_orders/{oid}")
        if proxy_order:
            user = await get_user(proxy_order.get("user_id", 0)) or {}
            items_preview = ""
            items_list = proxy_order.get("items", [])
            if items_list:
                preview = "\n".join(str(i) for i in items_list[:5])
                more = f"\n<i>... +{len(items_list) - 5} more</i>" if len(items_list) > 5 else ""
                items_preview = f"\n📋 <b>Items:</b>\n<code>{preview}{more}</code>"
            await message.answer(
                f"🔐 <b>Proxy Order Details</b>\n{_SEP}\n"
                f"🆔 Order: <code>{oid}</code>\n"
                f"👤 User: @{user.get('username', '—')} (<code>{proxy_order.get('user_id')}</code>)\n"
                f"📦 Product: <b>{proxy_order.get('product_name', '—')}</b>\n"
                f"📡 Duration: <b>{proxy_order.get('duration_days', '—')}</b>\n"
                f"💵 Price: <b>${proxy_order.get('price', 0):.2f}</b>\n"
                f"📊 Status: <b>{proxy_order.get('status', '—')}</b>\n"
                f"🕒 Date: {_dt(proxy_order.get('created_at', 0))}"
                f"{items_preview}",
                reply_markup=_kb([BACK_BTN, HOME_BTN]),
            )
            return
    await message.answer(
        f"❌ <b>Order Not Found</b>\n{_SEP}\n"
        f"No order found with ID: <code>{oid}</code>\n"
        f"Please check the ID and try again.",
        reply_markup=_kb([BACK_BTN, HOME_BTN]),
    )


# ── Export Mail Orders ─────────────────────────────────────────────

@router_admin.message(AdminFlow.menu, F.text == BTN_ADM_EXPORT)
@router_admin.message(AdminFlow.orders_submenu, F.text == BTN_ADM_EXPORT)
async def admin_export_orders(message: Message):
    _init_mail_shop_file()

    if not MAIL_SHOP_FILE.exists():
        await message.answer("📋 No mail orders file found yet.", reply_markup=admin_main_kb())
        return

    wb = openpyxl.load_workbook(str(MAIL_SHOP_FILE))
    ws = wb.active
    total_rows = ws.max_row - 1  # subtract header

    if total_rows <= 0:
        await message.answer("📋 No mail orders recorded yet.", reply_markup=admin_main_kb())
        return

    buf = io.BytesIO()
    wb.save(buf)
    buf.seek(0)

    await message.answer_document(
        document=BufferedInputFile(buf.read(), filename="My_Mail_Shop_Orders.xlsx"),
        caption=(
            f"📋 <b>My Mail Shop Orders</b>\n{_SEP}\n"
            f"📦 Total Orders: <b>{total_rows}</b>\n"
            f"🗓 Updated: {time.strftime('%d %b %Y %H:%M')} UTC"
        ),
        reply_markup=admin_main_kb(),
    )


# ── Broadcast ──────────────────────────────────────────────────────

@router_admin.message(AdminFlow.menu, F.text == BTN_ADM_BROADCAST)
@router_admin.message(AdminFlow.tools_submenu, F.text == BTN_ADM_BROADCAST)
async def admin_broadcast_start(message: Message, state: FSMContext):
    await state.set_state(AdminFlow.broadcast)
    await message.answer(
        f"📢 <b>Broadcast</b>\n{_SEP}\n"
        f"Send a text message or photo+caption.\nIt will be sent to ALL users.\n\n"
        f"Press {BACK_BTN} to cancel.",
        reply_markup=_kb([BACK_BTN, HOME_BTN]),
    )


@router_admin.message(AdminFlow.broadcast)
async def receive_broadcast(message: Message, state: FSMContext):
    if message.text in (BACK_BTN, HOME_BTN):
        await state.set_state(AdminFlow.menu)
        await message.answer("❌ Cancelled.", reply_markup=admin_main_kb())
        return
    if not message.text and not message.photo:
        await message.answer("❌ Send text or a photo with caption.")
        return
    await state.set_state(AdminFlow.menu)
    users  = await get_all_users()
    uids   = list(users.keys())
    status = await message.answer(f"📢 Broadcasting to <b>{len(uids)}</b> users…")
    sent = failed = 0
    for uid_str in uids:
        try:
            uid = int(uid_str)
            if message.photo:
                await message.bot.send_photo(uid, photo=message.photo[-1].file_id, caption=message.caption or "")
            else:
                await message.bot.send_message(uid, message.text)
            sent += 1
        except Exception:
            failed += 1
        await asyncio.sleep(0.05)
    try:
        await status.edit_text(
            f"📢 <b>Broadcast Complete!</b>\n{_SEP}\n✅ Sent: <b>{sent}</b>\n❌ Failed: <b>{failed}</b>"
        )
    except Exception:
        await message.answer(
            f"📢 <b>Broadcast Complete!</b>\n{_SEP}\n✅ Sent: <b>{sent}</b>\n❌ Failed: <b>{failed}</b>"
        )
    await message.answer("🔐 <b>Admin Panel</b>", reply_markup=admin_main_kb())


# ── Settings ───────────────────────────────────────────────────────

@router_admin.message(AdminFlow.menu, F.text == BTN_ADM_SETTINGS)
async def admin_settings(message: Message, state: FSMContext):
    settings = await get_settings()
    await state.set_state(AdminFlow.settings_menu)
    await message.answer(fmt_settings(settings), reply_markup=admin_settings_kb())


@router_admin.message(AdminFlow.settings_menu)
async def admin_settings_action(message: Message, state: FSMContext):
    if message.text in (BACK_BTN, HOME_BTN):
        data = await state.get_data()
        if data.get("settings_category"):
            # Back from sub-category -> show categories
            await state.update_data(settings_category=None)
            settings = await get_settings()
            await message.answer(fmt_settings(settings), reply_markup=admin_settings_kb())
            return
        # Back from categories -> admin main
        await state.set_state(AdminFlow.menu)
        await message.answer("🔐 <b>Admin Panel</b>", reply_markup=admin_main_kb())
        return
    # Check if it's a category button
    if message.text in SETTINGS_CATEGORIES:
        keys = SETTINGS_CATEGORIES[message.text]
        await state.update_data(settings_category=message.text)
        await message.answer(f"⚙️ <b>{message.text}</b>", reply_markup=settings_category_kb(keys))
        return
    # Check if it's a setting item
    info = SETTINGS_MAP.get(message.text)
    if not info:
        await message.answer("❌ Select from keyboard.")
        return
    key, type_, prompt = info
    await state.update_data(s_key=key, s_type=type_.__name__)
    await state.set_state(AdminFlow.settings_edit)
    await message.answer(
        f"⚙️ <b>{message.text}</b>\n{_SEP}\n{prompt}\n\n<i>Send blank to clear this value.</i>",
        reply_markup=input_kb(),
    )


@router_admin.message(AdminFlow.settings_edit)
async def receive_setting(message: Message, state: FSMContext):
    if message.text in (CANCEL_BTN, BACK_BTN, HOME_BTN):
        data = await state.get_data()
        cat = data.get("settings_category")
        await state.set_state(AdminFlow.settings_menu)
        if cat and cat in SETTINGS_CATEGORIES:
            await message.answer(f"⚙️ <b>{cat}</b>", reply_markup=settings_category_kb(SETTINGS_CATEGORIES[cat]))
        else:
            settings = await get_settings()
            await message.answer(fmt_settings(settings), reply_markup=admin_settings_kb())
        return
    data      = await state.get_data()
    key       = data.get("s_key", "")
    type_name = data.get("s_type", "str")
    raw       = message.text.strip()
    if type_name == "float" and raw:
        try:
            value: Any = float(raw)
        except ValueError:
            await message.answer("❌ Enter a valid number.")
            return
        if key == "usd_rate" and value <= 0:
            await message.answer("❌ USD rate must be greater than 0.")
            return
        if key in ("bkash_min", "nagad_min", "binance_min") and value < 0:
            await message.answer("❌ Minimum deposit cannot be negative.")
            return
        if key == "referral_bonus_pct" and not (0 <= value <= 100):
            await message.answer("❌ Bonus % must be between 0 and 100.")
            return
        if key == "backup_interval_hours" and value < 0:
            await message.answer("❌ Backup interval cannot be negative. Use 0 to disable.")
            return
    else:
        value = raw
        if key == "maintenance_mode":
            normalized = raw.strip().upper()
            if normalized in ("ON", "1", "TRUE", "YES"):
                value = "ON"
            elif normalized in ("OFF", "0", "FALSE", "NO"):
                value = "OFF"
            else:
                await message.answer("❌ Invalid value. Enter ON or OFF (also accepts yes/no, true/false, 1/0).")
                return
    await update_settings({key: value})
    settings = await get_settings()
    await state.set_state(AdminFlow.settings_menu)
    data = await state.get_data()
    cat = data.get("settings_category")
    if cat and cat in SETTINGS_CATEGORIES:
        await message.answer(
            f"✅ <b>Updated!</b>  {key} → <code>{value or '(cleared)'}</code>",
            reply_markup=settings_category_kb(SETTINGS_CATEGORIES[cat]),
        )
    else:
        await message.answer(
            f"✅ <b>Updated!</b>  {key} → <code>{value or '(cleared)'}</code>\n\n" + fmt_settings(settings),
            reply_markup=admin_settings_kb(),
        )


# ══════════════════════════════════════════════════════════════════
# ADMIN — Proxy Data Packages Manager
# ══════════════════════════════════════════════════════════════════

def _pkg_options(settings: dict) -> list:
    raw = settings.get("proxy_data_options", "1 GB,5 GB,10 GB,50 GB")
    return [o.strip() for o in raw.split(",") if o.strip()]

def _pkg_summary(options: list) -> str:
    if not options:
        return "<i>No packages set. Add one below.</i>"
    lines = "\n".join(f"  • <b>{o}</b>" for o in options)
    return lines

@router_admin.message(AdminFlow.menu, F.text == BTN_ADM_PROXY_PKGS)
@router_admin.message(AdminFlow.tools_submenu, F.text == BTN_ADM_PROXY_PKGS)
async def admin_proxy_pkgs(message: Message, state: FSMContext):
    settings = await get_settings()
    opts = _pkg_options(settings)
    await state.set_state(AdminFlow.proxy_pkg_list)
    await message.answer(
        f"📡 <b>Proxy Data Packages</b>\n{_SEP}\n"
        f"Current packages:\n{_pkg_summary(opts)}\n\n"
        f"Tap 🗑 <b>a package</b> to remove it, or ➕ <b>Add Package</b> to add new.",
        reply_markup=proxy_pkg_manage_kb(opts),
    )


@router_admin.message(AdminFlow.proxy_pkg_list)
async def proxy_pkg_list_action(message: Message, state: FSMContext):
    if message.text in (BACK_BTN, HOME_BTN):
        await state.set_state(AdminFlow.menu)
        await message.answer("🔐 <b>Admin Panel</b>", reply_markup=admin_main_kb())
        return

    if message.text == BTN_PKG_ADD:
        await state.set_state(AdminFlow.proxy_pkg_add)
        await message.answer(
            f"📡 <b>Add Data Package</b>\n{_SEP}\n"
            f"Enter the data size to add:\n"
            f"<i>Examples: 1 GB · 5 GB · 500 MB · 2.5 GB</i>",
            reply_markup=input_kb(),
        )
        return

    if message.text.startswith("🗑 "):
        label = message.text[2:].strip()
        settings = await get_settings()
        opts = _pkg_options(settings)
        if label in opts:
            opts.remove(label)
            await update_settings({"proxy_data_options": ",".join(opts)})
        settings = await get_settings()
        opts = _pkg_options(settings)
        await message.answer(
            f"🗑 <b>Removed:</b> {label}\n\n"
            f"📡 <b>Current packages:</b>\n{_pkg_summary(opts)}",
            reply_markup=proxy_pkg_manage_kb(opts),
        )
        return

    await message.answer("❌ Tap a package to remove, or ➕ Add Package.")


@router_admin.message(AdminFlow.proxy_pkg_add)
async def proxy_pkg_add_input(message: Message, state: FSMContext):
    if message.text in (CANCEL_BTN, BACK_BTN, HOME_BTN):
        settings = await get_settings()
        opts = _pkg_options(settings)
        await state.set_state(AdminFlow.proxy_pkg_list)
        await message.answer(
            f"📡 <b>Proxy Data Packages</b>\n{_SEP}\n"
            f"Current packages:\n{_pkg_summary(opts)}",
            reply_markup=proxy_pkg_manage_kb(opts),
        )
        return

    raw = message.text.strip()
    amount, err = parse_data_amount(raw)
    if err or not amount:
        await message.answer(
            f"❌ Invalid format.\nEnter like: <code>5 GB</code> or <code>500 MB</code>",
            reply_markup=input_kb(),
        )
        return

    unit = "MB" if "MB" in raw.upper() else "GB"
    label = f"{int(amount)} {unit}" if amount == int(amount) else f"{amount} {unit}"

    settings = await get_settings()
    opts = _pkg_options(settings)
    if label in opts:
        await message.answer(f"⚠️ <b>{label}</b> is already in the list.", reply_markup=input_kb())
        return

    opts.append(label)
    await update_settings({"proxy_data_options": ",".join(opts)})
    settings = await get_settings()
    opts = _pkg_options(settings)
    await state.set_state(AdminFlow.proxy_pkg_list)
    await message.answer(
        f"✅ <b>Added:</b> {label}\n\n"
        f"📡 <b>Current packages:</b>\n{_pkg_summary(opts)}",
        reply_markup=proxy_pkg_manage_kb(opts),
    )




# ══════════════════════════════════════════════════════════════════
# ADMIN — FEAT-002: Enhanced Analytics (hourly, daily chart, conversion)
# ══════════════════════════════════════════════════════════════════

@router_admin.callback_query(F.data == "adm_analytics_detail")
async def admin_analytics_detail_cb(call: CallbackQuery):
    if not is_admin(call.from_user.id):
        await call.answer("🚫 Denied", show_alert=True)
        return
    orders, vpn, proxy = await asyncio.gather(
        db_get("orders"), db_get("vpn_orders"), db_get("proxy_orders"),
    )
    orders = orders or {}
    vpn = vpn or {}
    proxy = proxy or {}
    now = int(time.time())
    today_start = now - (now % 86400)
    week_start = now - 7 * 86400

    # Hourly breakdown for today
    hourly = [0.0] * 24
    today_count = 0
    total_value = 0.0
    for v in orders.values():
        ts = v.get("created_at", 0)
        price = v.get("total_price", 0)
        if ts >= today_start:
            hour = ((ts - today_start) // 3600) % 24
            hourly[hour] += price
            today_count += 1
            total_value += price
    for v in vpn.values():
        ts = v.get("created_at", 0)
        price = v.get("price", 0)
        if ts >= today_start and v.get("status") == "delivered":
            hour = ((ts - today_start) // 3600) % 24
            hourly[hour] += price
            today_count += 1
            total_value += price
    for v in proxy.values():
        ts = v.get("created_at", 0)
        price = v.get("price", 0)
        if ts >= today_start and v.get("status") == "delivered":
            hour = ((ts - today_start) // 3600) % 24
            hourly[hour] += price
            today_count += 1
            total_value += price

    # Daily chart for past 7 days
    daily = [0.0] * 7
    daily_counts = [0] * 7
    all_sources = list(orders.values()) + \
        [v for v in vpn.values() if v.get("status") == "delivered"] + \
        [v for v in proxy.values() if v.get("status") == "delivered"]
    for v in all_sources:
        ts = v.get("created_at", 0)
        price = v.get("total_price") or v.get("price", 0)
        if ts >= week_start:
            day_idx = min(6, (now - ts) // 86400)
            daily[6 - day_idx] += price
            daily_counts[6 - day_idx] += 1

    # Conversion rate: orders / unique users who visited
    users = await db_get("users") or {}
    total_users = len(users)
    total_orders = len(orders) + len([v for v in vpn.values() if v.get("status") == "delivered"]) + \
        len([v for v in proxy.values() if v.get("status") == "delivered"])
    conversion = (total_orders / total_users * 100) if total_users > 0 else 0
    avg_value = (total_value / today_count) if today_count > 0 else 0

    # Format hourly
    max_h = max(hourly) if any(hourly) else 1
    hourly_lines = []
    for h in range(0, 24, 3):
        val = sum(hourly[h:h+3])
        bar_len = int((val / max_h) * 8) if max_h > 0 else 0
        bar = "█" * bar_len
        hourly_lines.append(f"  {h:02d}-{h+3:02d}h: {bar} ${val:.0f}")

    # Format daily chart
    max_d = max(daily) if any(daily) else 1
    daily_lines = []
    for i in range(7):
        day_ts = now - (6 - i) * 86400
        day_lbl = datetime.fromtimestamp(day_ts, tz=pytz.UTC).strftime("%a")
        bar_len = int((daily[i] / max_d) * 8) if max_d > 0 else 0
        bar = "█" * bar_len
        daily_lines.append(f"  {day_lbl}: {bar} ${daily[i]:.0f} ({daily_counts[i]})")

    text = (
        f"📈 <b>Detailed Analytics</b>\n{_SEP}\n\n"
        f"<b>Hourly Breakdown (Today)</b>\n"
        + "\n".join(hourly_lines) + "\n\n"
        f"<b>Daily Revenue (7 days)</b>\n"
        + "\n".join(daily_lines) + "\n\n"
        f"{_SEP}\n"
        f"📊 <b>Conversion Rate:</b> {conversion:.1f}%\n"
        f"💰 <b>Avg Order Value (today):</b> ${avg_value:.2f}\n"
        f"🛍 <b>Total Orders (all time):</b> {total_orders}\n"
        f"👥 <b>Total Users:</b> {total_users}"
    )
    await call.message.answer(text)
    await call.answer()


# ══════════════════════════════════════════════════════════════════
# ADMIN — FEAT-002: Low Stock Alerts
# ══════════════════════════════════════════════════════════════════

@router_admin.message(AdminFlow.menu, F.text == BTN_ADM_LOW_STOCK)
@router_admin.message(AdminFlow.tools_submenu, F.text == BTN_ADM_LOW_STOCK)
async def admin_low_stock_alerts(message: Message):
    if not is_admin(message.from_user.id):
        return
    settings = await get_settings()
    threshold = int(settings.get("low_stock_threshold", 5))
    products = await get_all_products()
    low_stock = []
    for pid, p in products.items():
        if p.get("hidden"):
            continue
        count = p.get("stock_count", 0)
        if count <= threshold:
            low_stock.append((pid, p.get("name", pid), p.get("emoji", "📦"), count))
    if not low_stock:
        await message.answer(
            f"🚨 <b>Low Stock Alerts</b>\n{_SEP}\n\n"
            f"✅ All products have stock above threshold ({threshold}).",
            reply_markup=admin_main_kb(),
        )
        return
    low_stock.sort(key=lambda x: x[3])
    lines = [f"🚨 <b>Low Stock Alerts</b>\n{_SEP}\n"
             f"⚠️ Threshold: <b>{threshold}</b> items\n"]
    buttons = []
    for pid, name, emoji, count in low_stock[:15]:
        status = "🔴 OUT" if count == 0 else f"🟡 {count}"
        lines.append(f"  {emoji} {name}: <b>{status}</b>")
        buttons.append([InlineKeyboardButton(
            text=f"📥 Restock: {name}", callback_data=f"adm_restock:{pid}"
        )])
    text = "\n".join(lines)
    kb = InlineKeyboardMarkup(inline_keyboard=buttons) if buttons else None
    await message.answer(text, reply_markup=kb)


@router_admin.callback_query(F.data.startswith("adm_restock:"))
async def admin_restock_cb(call: CallbackQuery, state: FSMContext):
    if not is_admin(call.from_user.id):
        await call.answer("🚫 Denied", show_alert=True)
        return
    pid = call.data.split(":", 1)[1]
    product = await get_product(pid)
    if not product:
        await call.answer("Product not found", show_alert=True)
        return
    await state.update_data(stock_pid=pid)
    await state.set_state(AdminFlow.stock_detail)
    await call.message.answer(
        f"📦 <b>Restock: {product.get('name', pid)}</b>\n{_SEP}\n"
        f"Current stock: <b>{product.get('stock_count', 0)}</b>\n\n"
        f"Send a .txt or .xlsx file with stock items, or type items line by line.",
        reply_markup=_kb([BACK_BTN, HOME_BTN]),
    )
    await call.answer()


# ══════════════════════════════════════════════════════════════════
# ADMIN — FEAT-002: Auto Price Sync (category markup rules)
# ══════════════════════════════════════════════════════════════════

@router_admin.message(AdminFlow.menu, F.text == BTN_ADM_PRICE_SYNC)
@router_admin.message(AdminFlow.tools_submenu, F.text == BTN_ADM_PRICE_SYNC)
async def admin_price_sync(message: Message, state: FSMContext):
    if not is_admin(message.from_user.id):
        return
    role = await get_admin_role(message.from_user.id)
    if not check_role_permission(role, "manage_settings"):
        await message.answer("🚫 Moderators cannot manage price sync.", reply_markup=admin_main_kb())
        return
    rules = await db_get("settings/price_rules") or {}
    lines = [f"💲 <b>Auto Price Sync</b>\n{_SEP}\n"
             f"Set markup % per category. New products inherit the category markup.\n"]
    if rules:
        for cat, pct in rules.items():
            lines.append(f"  📂 <b>{cat}</b>: +{pct}%")
    else:
        lines.append("  <i>No rules set yet.</i>")
    lines.append(f"\n{_SEP}\nSend: <code>category:percentage</code>\nExample: <code>mail:15</code>")
    lines.append("Send <code>remove:category</code> to delete a rule.")
    await state.set_state(AdminFlow.price_sync)
    await message.answer("\n".join(lines), reply_markup=_kb([BACK_BTN, HOME_BTN]))


@router_admin.message(AdminFlow.price_sync)
async def admin_price_sync_input(message: Message, state: FSMContext):
    if message.text in (BACK_BTN, HOME_BTN):
        await state.set_state(AdminFlow.menu)
        await message.answer("🔐 <b>Admin Panel</b>", reply_markup=admin_main_kb())
        return
    text = (message.text or "").strip()
    if text.lower().startswith("remove:"):
        cat = text.split(":", 1)[1].strip().lower()
        rules = await db_get("settings/price_rules") or {}
        if cat in rules:
            del rules[cat]
            await db_set("settings/price_rules", rules)
            await message.answer(f"✅ Removed rule for <b>{cat}</b>.")
        else:
            await message.answer(f"❌ No rule found for <b>{cat}</b>.")
        return
    if ":" not in text:
        await message.answer("❌ Format: <code>category:percentage</code>\nExample: <code>mail:15</code>")
        return
    parts = text.split(":", 1)
    cat = parts[0].strip().lower()
    try:
        pct = float(parts[1].strip())
    except ValueError:
        await message.answer("❌ Percentage must be a number.")
        return
    if pct < 0 or pct > 500:
        await message.answer("❌ Percentage must be between 0 and 500.")
        return
    rules = await db_get("settings/price_rules") or {}
    rules[cat] = pct
    await db_set("settings/price_rules", rules)
    await log_admin_action(message.from_user.id, "price_sync", f"Set {cat} markup to {pct}%")
    await message.answer(f"✅ Set <b>{cat}</b> markup to <b>+{pct}%</b>")


# ══════════════════════════════════════════════════════════════════
# ADMIN — FEAT-002: Auto Product Import (batch JSON)
# ══════════════════════════════════════════════════════════════════

@router_admin.message(AdminFlow.menu, F.text == BTN_ADM_AUTO_IMPORT)
@router_admin.message(AdminFlow.tools_submenu, F.text == BTN_ADM_AUTO_IMPORT)
async def admin_auto_import(message: Message, state: FSMContext):
    if not is_admin(message.from_user.id):
        return
    role = await get_admin_role(message.from_user.id)
    if not check_role_permission(role, "manage_products"):
        await message.answer("🚫 Moderators cannot import products.", reply_markup=admin_main_kb())
        return
    await state.set_state(AdminFlow.auto_import)
    await message.answer(
        f"📥 <b>Auto Product Import</b>\n{_SEP}\n\n"
        f"Paste a JSON array of products to import.\n\n"
        f"<b>Format:</b>\n<code>"
        '[\n  {"name": "Gmail", "price": 0.5, "category": "mail", "emoji": "📧"},\n'
        '  {"name": "Yahoo", "price": 0.3, "category": "mail"}\n]</code>\n\n'
        f"Required fields: <b>name, price</b>\n"
        f"Optional: category, emoji, description, hidden",
        reply_markup=_kb([BACK_BTN, HOME_BTN]),
    )


@router_admin.message(AdminFlow.auto_import)
async def admin_auto_import_input(message: Message, state: FSMContext):
    if message.text in (BACK_BTN, HOME_BTN):
        await state.set_state(AdminFlow.menu)
        await message.answer("🔐 <b>Admin Panel</b>", reply_markup=admin_main_kb())
        return
    text = (message.text or "").strip()
    if not text:
        await message.answer("❌ Please paste a JSON array of products.")
        return
    try:
        products_data = json.loads(text)
    except json.JSONDecodeError as e:
        await message.answer(f"❌ Invalid JSON: {e}")
        return
    if not isinstance(products_data, list):
        await message.answer("❌ Expected a JSON array (list of products).")
        return
    if len(products_data) > 50:
        await message.answer("❌ Maximum 50 products per import.")
        return
    # Get price rules for auto markup
    price_rules = await db_get("settings/price_rules") or {}
    imported = 0
    errors = []
    for i, p in enumerate(products_data):
        if not isinstance(p, dict):
            errors.append(f"Item {i+1}: not a dict")
            continue
        name = p.get("name", "").strip()
        price = p.get("price")
        if not name:
            errors.append(f"Item {i+1}: missing name")
            continue
        if price is None:
            errors.append(f"Item {i+1}: missing price")
            continue
        try:
            price = float(price)
        except (ValueError, TypeError):
            errors.append(f"Item {i+1}: invalid price")
            continue
        if price < 0:
            errors.append(f"Item {i+1}: negative price")
            continue
        category = str(p.get("category", "mail")).strip().lower()
        # Apply price rule if exists
        if category in price_rules:
            price = round(price * (1 + price_rules[category] / 100), 2)
        emoji = p.get("emoji", "📮")
        desc = p.get("description", "")
        hidden = bool(p.get("hidden", False))
        await create_product(name, price, emoji, category, "manual", hidden, desc)
        imported += 1
    err_text = ""
    if errors:
        err_text = "\n\n<b>Errors:</b>\n" + "\n".join(f"  - {e}" for e in errors[:10])
    await log_admin_action(message.from_user.id, "auto_import", f"Imported {imported} products")
    await state.set_state(AdminFlow.menu)
    await message.answer(
        f"✅ <b>Import Complete</b>\n{_SEP}\n"
        f"📦 Imported: <b>{imported}</b> products\n"
        f"❌ Errors: <b>{len(errors)}</b>{err_text}",
        reply_markup=admin_main_kb(),
    )


# ══════════════════════════════════════════════════════════════════
# ADMIN — FEAT-002: Multiple Admin Roles
# ══════════════════════════════════════════════════════════════════

@router_admin.message(AdminFlow.menu, F.text == BTN_ADM_ROLES)
@router_admin.message(AdminFlow.tools_submenu, F.text == BTN_ADM_ROLES)
async def admin_roles_menu(message: Message, state: FSMContext):
    if not is_admin(message.from_user.id):
        return
    role = await get_admin_role(message.from_user.id)
    if role != "owner":
        await message.answer("🚫 Only owners can manage roles.", reply_markup=admin_main_kb())
        return
    roles_data = await get_all_admin_roles()
    lines = [f"👑 <b>Admin Roles</b>\n{_SEP}\n"
             f"<b>Role Hierarchy:</b> Owner > Admin > Moderator\n"]
    # Show existing ADMIN_IDS as owners
    for aid in ADMIN_IDS:
        lines.append(f"  👑 <code>{aid}</code> - <b>owner</b> (config)")
    # Show additional roles from Firebase
    for uid_str, rdata in roles_data.items():
        uid_int = int(uid_str) if uid_str.isdigit() else 0
        if uid_int in ADMIN_IDS:
            continue
        r = rdata.get("role", "none") if isinstance(rdata, dict) else str(rdata)
        if r != "none":
            emoji_r = "🛡" if r == "admin" else "👁"
            lines.append(f"  {emoji_r} <code>{uid_str}</code> - <b>{r}</b>")
    lines.append(f"\n{_SEP}\nTo assign a role, send:\n<code>user_id:role</code>")
    lines.append("Roles: <b>admin</b>, <b>moderator</b>, <b>none</b> (remove)")
    await state.set_state(AdminFlow.roles_menu)
    await message.answer("\n".join(lines), reply_markup=_kb([BACK_BTN, HOME_BTN]))


@router_admin.message(AdminFlow.roles_menu)
async def admin_roles_input(message: Message, state: FSMContext):
    if message.text in (BACK_BTN, HOME_BTN):
        await state.set_state(AdminFlow.menu)
        await message.answer("🔐 <b>Admin Panel</b>", reply_markup=admin_main_kb())
        return
    text = (message.text or "").strip()
    if ":" not in text:
        await message.answer("❌ Format: <code>user_id:role</code>\nExample: <code>123456789:admin</code>")
        return
    parts = text.split(":", 1)
    try:
        target_uid = int(parts[0].strip())
    except ValueError:
        await message.answer("❌ Invalid user ID.")
        return
    role_val = parts[1].strip().lower()
    if role_val not in ("admin", "moderator", "none"):
        await message.answer("❌ Valid roles: <b>admin</b>, <b>moderator</b>, <b>none</b>")
        return
    if target_uid in ADMIN_IDS and role_val != "owner":
        await message.answer("⚠️ Cannot change role of config-defined admins (they are always owners).")
        return
    await set_admin_role(target_uid, role_val)
    await log_admin_action(message.from_user.id, "set_role", f"Set {target_uid} to {role_val}")
    emoji_r = "👑" if role_val == "owner" else ("🛡" if role_val == "admin" else ("👁" if role_val == "moderator" else "❌"))
    await message.answer(
        f"✅ {emoji_r} User <code>{target_uid}</code> role set to <b>{role_val}</b>",
        reply_markup=_kb([BACK_BTN, HOME_BTN]),
    )


# ══════════════════════════════════════════════════════════════════
# ADMIN — FEAT-002: Advanced User Search
# ══════════════════════════════════════════════════════════════════

@router_admin.callback_query(F.data == "adm_adv_search")
async def admin_advanced_search_cb(call: CallbackQuery, state: FSMContext):
    if not is_admin(call.from_user.id):
        await call.answer("🚫 Denied", show_alert=True)
        return
    await state.set_state(AdminFlow.advanced_search)
    await call.message.answer(
        f"👥 <b>Advanced User Search</b>\n{_SEP}\n\n"
        f"<b>Search options:</b>\n"
        f"  - Username: <code>@username</code>\n"
        f"  - User ID: <code>12345678</code>\n"
        f"  - Balance range: <code>bal:10-50</code>\n"
        f"  - Order count: <code>orders:5+</code>\n"
        f"  - All users: <code>all</code>\n\n"
        f"<i>Results shown as paginated inline list.</i>",
        reply_markup=_kb([BACK_BTN, HOME_BTN]),
    )
    await call.answer()


@router_admin.message(AdminFlow.advanced_search)
async def admin_advanced_search_input(message: Message, state: FSMContext):
    if message.text in (BACK_BTN, HOME_BTN):
        await state.set_state(AdminFlow.menu)
        await message.answer("🔐 <b>Admin Panel</b>", reply_markup=admin_main_kb())
        return
    query = (message.text or "").strip()
    if not query:
        await message.answer("❌ Enter a search query.")
        return
    users = await get_all_users()
    results = []
    if query.startswith("@"):
        uname = query[1:].lower()
        for uid, u in users.items():
            if uname in (u.get("username") or "").lower():
                results.append((uid, u))
    elif query.isdigit():
        uid_search = query
        if uid_search in users:
            results.append((uid_search, users[uid_search]))
    elif query.lower().startswith("bal:"):
        range_str = query[4:].strip()
        try:
            if "-" in range_str:
                low, high = range_str.split("-", 1)
                low, high = float(low), float(high)
            else:
                low, high = float(range_str), 999999
        except ValueError:
            await message.answer("❌ Invalid balance range. Use: <code>bal:10-50</code>")
            return
        for uid, u in users.items():
            bal = u.get("balance", 0)
            if low <= bal <= high:
                results.append((uid, u))
    elif query.lower().startswith("orders:"):
        count_str = query[7:].strip().rstrip("+")
        try:
            min_orders = int(count_str)
        except ValueError:
            await message.answer("❌ Invalid order count. Use: <code>orders:5+</code>")
            return
        for uid, u in users.items():
            if (u.get("order_count") or 0) >= min_orders:
                results.append((uid, u))
    elif query.lower() == "all":
        results = list(users.items())[:50]
    else:
        # Fuzzy search by username or first name
        for uid, u in users.items():
            uname = (u.get("username") or "").lower()
            fname = (u.get("first_name") or "").lower()
            if query.lower() in uname or query.lower() in fname:
                results.append((uid, u))
    if not results:
        await message.answer("❌ No users found matching your search.")
        return
    # Paginate as inline buttons (max 10 per page)
    results = results[:30]
    buttons = []
    for uid, u in results[:10]:
        uname = u.get("username") or "No username"
        bal = u.get("balance", 0)
        buttons.append([InlineKeyboardButton(
            text=f"@{uname} | ${bal:.2f} | ID:{uid}",
            callback_data=f"adm_uinfo:{uid}",
        )])
    if len(results) > 10:
        buttons.append([InlineKeyboardButton(
            text=f"... +{len(results) - 10} more results",
            callback_data="adm_search_more",
        )])
    kb = InlineKeyboardMarkup(inline_keyboard=buttons)
    await message.answer(
        f"👥 <b>Search Results</b> ({len(results)} found)",
        reply_markup=kb,
    )


@router_admin.callback_query(F.data == "adm_search_more")
async def adm_search_more_cb(call: CallbackQuery):
    if not is_admin(call.from_user.id):
        await call.answer("\U0001F6AB Denied", show_alert=True)
        return
    await call.answer(
        "Use more specific search terms to narrow results.",
        show_alert=True,
    )


@router_admin.callback_query(F.data.startswith("adm_uinfo:"))
async def admin_user_info_cb(call: CallbackQuery):
    if not is_admin(call.from_user.id):
        await call.answer("🚫 Denied", show_alert=True)
        return
    uid_str = call.data.split(":", 1)[1]
    user = await db_get(f"users/{uid_str}")
    if not user:
        await call.answer("User not found", show_alert=True)
        return
    text = (
        f"👤 <b>User Details</b>\n{_SEP}\n"
        f"🆔 ID: <code>{uid_str}</code>\n"
        f"👤 Username: @{user.get('username', 'N/A')}\n"
        f"📛 Name: {user.get('first_name', 'N/A')}\n"
        f"💰 Balance: <b>${user.get('balance', 0):.2f}</b>\n"
        f"🛍 Orders: <b>{user.get('order_count', 0)}</b>\n"
        f"💵 Total Spent: <b>${user.get('total_spent', 0):.2f}</b>\n"
        f"🚫 Banned: <b>{'Yes' if user.get('banned') else 'No'}</b>\n"
        f"🕒 Joined: {_dt(user.get('created_at', 0))}"
    )
    await call.message.answer(text)
    await call.answer()


# ══════════════════════════════════════════════════════════════════
# ADMIN — FEAT-002: Transaction Export (CSV/Excel)
# ══════════════════════════════════════════════════════════════════

@router_admin.message(AdminFlow.menu, F.text == BTN_ADM_EXPORT_TX)
@router_admin.message(AdminFlow.reports_submenu, F.text == BTN_ADM_EXPORT_TX)
async def admin_export_transactions(message: Message, state: FSMContext):
    if not is_admin(message.from_user.id):
        return
    await state.set_state(AdminFlow.export_transactions)
    await message.answer(
        f"📊 <b>Transaction Export</b>\n{_SEP}\n\n"
        f"Export all transactions as Excel (.xlsx) with separate sheets.\n\n"
        f"<b>Options:</b>\n"
        f"  - Send <code>all</code> for all-time data\n"
        f"  - Send <code>7d</code> for last 7 days\n"
        f"  - Send <code>30d</code> for last 30 days\n"
        f"  - Send dates: <code>2024-01-01:2024-12-31</code>",
        reply_markup=_kb([BACK_BTN, HOME_BTN]),
    )


@router_admin.message(AdminFlow.export_transactions)
async def admin_export_tx_input(message: Message, state: FSMContext):
    if message.text in (BACK_BTN, HOME_BTN):
        await state.set_state(AdminFlow.menu)
        await message.answer("🔐 <b>Admin Panel</b>", reply_markup=admin_main_kb())
        return
    text = (message.text or "").strip().lower()
    now = int(time.time())
    start_ts = 0
    end_ts = now
    if text == "all":
        start_ts = 0
    elif text == "7d":
        start_ts = now - 7 * 86400
    elif text == "30d":
        start_ts = now - 30 * 86400
    elif ":" in text:
        try:
            parts = text.split(":", 1)
            start_dt = datetime.strptime(parts[0].strip(), "%Y-%m-%d")
            end_dt = datetime.strptime(parts[1].strip(), "%Y-%m-%d")
            start_ts = int(start_dt.timestamp())
            end_ts = int(end_dt.timestamp()) + 86400
        except (ValueError, IndexError):
            await message.answer("❌ Invalid date format. Use: <code>2024-01-01:2024-12-31</code>")
            return
    else:
        await message.answer("❌ Send <code>all</code>, <code>7d</code>, <code>30d</code>, or <code>date:date</code>")
        return

    orders, vpn, proxy, deposits = await asyncio.gather(
        db_get("orders"), db_get("vpn_orders"),
        db_get("proxy_orders"), db_get("deposits"),
    )
    orders = orders or {}
    vpn = vpn or {}
    proxy = proxy or {}
    deposits = deposits or {}

    wb = openpyxl.Workbook()
    header_font = Font(bold=True, color="FFFFFF")
    header_fill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")

    # Sheet 1: Mail Orders
    ws = wb.active
    ws.title = "Mail Orders"
    headers = ["Order ID", "User ID", "Product", "Qty", "Total ($)", "Status", "Date"]
    for col, h in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col, value=h)
        cell.font = header_font
        cell.fill = header_fill
    row = 2
    for oid, o in orders.items():
        ts = o.get("created_at", 0)
        if start_ts <= ts <= end_ts:
            ws.cell(row=row, column=1, value=oid)
            ws.cell(row=row, column=2, value=str(o.get("user_id", "")))
            ws.cell(row=row, column=3, value=o.get("product_name", ""))
            ws.cell(row=row, column=4, value=o.get("qty", 0))
            ws.cell(row=row, column=5, value=round(o.get("total_price", 0), 2))
            ws.cell(row=row, column=6, value=o.get("status", ""))
            ws.cell(row=row, column=7, value=_dt(ts))
            row += 1

    # Sheet 2: VPN Orders
    ws2 = wb.create_sheet("VPN Orders")
    headers2 = ["Order ID", "User ID", "Product", "Duration", "Price ($)", "Status", "Date"]
    for col, h in enumerate(headers2, 1):
        cell = ws2.cell(row=1, column=col, value=h)
        cell.font = header_font
        cell.fill = header_fill
    row = 2
    for oid, o in vpn.items():
        ts = o.get("created_at", 0)
        if start_ts <= ts <= end_ts:
            ws2.cell(row=row, column=1, value=oid)
            ws2.cell(row=row, column=2, value=str(o.get("user_id", "")))
            ws2.cell(row=row, column=3, value=o.get("product_name", ""))
            ws2.cell(row=row, column=4, value=f"{o.get('duration_days', 0)} days")
            ws2.cell(row=row, column=5, value=round(o.get("price", 0), 2))
            ws2.cell(row=row, column=6, value=o.get("status", ""))
            ws2.cell(row=row, column=7, value=_dt(ts))
            row += 1

    # Sheet 3: Proxy Orders
    ws3 = wb.create_sheet("Proxy Orders")
    headers3 = ["Order ID", "User ID", "Product", "Data", "Price ($)", "Status", "Date"]
    for col, h in enumerate(headers3, 1):
        cell = ws3.cell(row=1, column=col, value=h)
        cell.font = header_font
        cell.fill = header_fill
    row = 2
    for oid, o in proxy.items():
        ts = o.get("created_at", 0)
        if start_ts <= ts <= end_ts:
            ws3.cell(row=row, column=1, value=oid)
            ws3.cell(row=row, column=2, value=str(o.get("user_id", "")))
            ws3.cell(row=row, column=3, value=o.get("product_name", ""))
            ws3.cell(row=row, column=4, value=str(o.get("duration_days", "")))
            ws3.cell(row=row, column=5, value=round(o.get("price", 0), 2))
            ws3.cell(row=row, column=6, value=o.get("status", ""))
            ws3.cell(row=row, column=7, value=_dt(ts))
            row += 1

    # Sheet 4: Deposits
    ws4 = wb.create_sheet("Deposits")
    headers4 = ["Deposit ID", "User ID", "Username", "Method", "BDT", "USD", "Status", "Date"]
    for col, h in enumerate(headers4, 1):
        cell = ws4.cell(row=1, column=col, value=h)
        cell.font = header_font
        cell.fill = header_fill
    row = 2
    for did, d in deposits.items():
        ts = d.get("created_at", 0)
        if start_ts <= ts <= end_ts:
            ws4.cell(row=row, column=1, value=did)
            ws4.cell(row=row, column=2, value=str(d.get("user_id", "")))
            ws4.cell(row=row, column=3, value=d.get("username", ""))
            ws4.cell(row=row, column=4, value=d.get("method", ""))
            ws4.cell(row=row, column=5, value=round(d.get("amount_bdt", 0), 2))
            ws4.cell(row=row, column=6, value=round(d.get("amount_usd", 0), 2))
            ws4.cell(row=row, column=7, value=d.get("status", ""))
            ws4.cell(row=row, column=8, value=_dt(ts))
            row += 1

    buf = io.BytesIO()
    wb.save(buf)
    buf.seek(0)
    date_str = time.strftime("%Y%m%d_%H%M", time.gmtime())
    await log_admin_action(message.from_user.id, "export_transactions", f"Filter: {text}")
    await state.set_state(AdminFlow.menu)
    await message.answer_document(
        document=BufferedInputFile(buf.read(), filename=f"transactions_{date_str}.xlsx"),
        caption=f"📊 <b>Transaction Export</b>\n{_SEP}\nFilter: <b>{text}</b>",
        reply_markup=admin_main_kb(),
    )


# ══════════════════════════════════════════════════════════════════
# ADMIN — FEAT-002: Service Status Manager
# ══════════════════════════════════════════════════════════════════

@router_admin.message(AdminFlow.menu, F.text == BTN_ADM_STATUS_MGR)
@router_admin.message(AdminFlow.tools_submenu, F.text == BTN_ADM_STATUS_MGR)
async def admin_service_status(message: Message, state: FSMContext):
    if not is_admin(message.from_user.id):
        return
    products = await get_all_products()
    if not products:
        await message.answer("📦 No products found.", reply_markup=admin_main_kb())
        return
    lines = [f"🔘 <b>Service Status Manager</b>\n{_SEP}\n"
             f"Toggle products online/offline with one tap.\n"]
    buttons = []
    for pid, p in sorted(products.items(), key=lambda x: x[1].get("name", "")):
        status = p.get("status", "online")
        icon = "🟢" if status == "online" else "🔴"
        name = p.get("name", pid)
        buttons.append([InlineKeyboardButton(
            text=f"{icon} {name} [{status}]",
            callback_data=f"adm_toggle:{pid}",
        )])
    buttons.append([InlineKeyboardButton(text="🔄 Bulk: All Online", callback_data="adm_bulk_on")])
    buttons.append([InlineKeyboardButton(text="⏸ Bulk: All Offline", callback_data="adm_bulk_off")])
    kb = InlineKeyboardMarkup(inline_keyboard=buttons)
    await message.answer("\n".join(lines), reply_markup=kb)


@router_admin.callback_query(F.data.startswith("adm_toggle:"))
async def admin_toggle_status_cb(call: CallbackQuery):
    if not is_admin(call.from_user.id):
        await call.answer("🚫 Denied", show_alert=True)
        return
    pid = call.data.split(":", 1)[1]
    product = await get_product(pid)
    if not product:
        await call.answer("Product not found", show_alert=True)
        return
    current = product.get("status", "online")
    new_status = "offline" if current == "online" else "online"
    await update_product(pid, {"status": new_status})
    icon = "🟢" if new_status == "online" else "🔴"
    await log_admin_action(call.from_user.id, "toggle_status", f"{product.get('name', pid)}: {new_status}")
    await call.answer(f"{icon} {product.get('name', pid)}: {new_status}")
    # Refresh the list
    products = await get_all_products()
    buttons = []
    for p_id, p in sorted(products.items(), key=lambda x: x[1].get("name", "")):
        status = p.get("status", "online")
        s_icon = "🟢" if status == "online" else "🔴"
        name = p.get("name", p_id)
        buttons.append([InlineKeyboardButton(
            text=f"{s_icon} {name} [{status}]",
            callback_data=f"adm_toggle:{p_id}",
        )])
    buttons.append([InlineKeyboardButton(text="🔄 Bulk: All Online", callback_data="adm_bulk_on")])
    buttons.append([InlineKeyboardButton(text="⏸ Bulk: All Offline", callback_data="adm_bulk_off")])
    kb = InlineKeyboardMarkup(inline_keyboard=buttons)
    try:
        await call.message.edit_reply_markup(reply_markup=kb)
    except Exception:
        pass


@router_admin.callback_query(F.data == "adm_bulk_on")
async def admin_bulk_online_cb(call: CallbackQuery):
    if not is_admin(call.from_user.id):
        await call.answer("🚫 Denied", show_alert=True)
        return
    products = await get_all_products()
    for pid in products:
        await update_product(pid, {"status": "online"})
    await log_admin_action(call.from_user.id, "bulk_status", "All products set to online")
    await call.answer("✅ All products set to ONLINE")
    # Refresh
    products = await get_all_products()
    buttons = []
    for p_id, p in sorted(products.items(), key=lambda x: x[1].get("name", "")):
        buttons.append([InlineKeyboardButton(
            text=f"🟢 {p.get('name', p_id)} [online]",
            callback_data=f"adm_toggle:{p_id}",
        )])
    buttons.append([InlineKeyboardButton(text="🔄 Bulk: All Online", callback_data="adm_bulk_on")])
    buttons.append([InlineKeyboardButton(text="⏸ Bulk: All Offline", callback_data="adm_bulk_off")])
    try:
        await call.message.edit_reply_markup(reply_markup=InlineKeyboardMarkup(inline_keyboard=buttons))
    except Exception:
        pass


@router_admin.callback_query(F.data == "adm_bulk_off")
async def admin_bulk_offline_cb(call: CallbackQuery):
    if not is_admin(call.from_user.id):
        await call.answer("🚫 Denied", show_alert=True)
        return
    products = await get_all_products()
    for pid in products:
        await update_product(pid, {"status": "offline"})
    await log_admin_action(call.from_user.id, "bulk_status", "All products set to offline")
    await call.answer("✅ All products set to OFFLINE")
    products = await get_all_products()
    buttons = []
    for p_id, p in sorted(products.items(), key=lambda x: x[1].get("name", "")):
        buttons.append([InlineKeyboardButton(
            text=f"🔴 {p.get('name', p_id)} [offline]",
            callback_data=f"adm_toggle:{p_id}",
        )])
    buttons.append([InlineKeyboardButton(text="🔄 Bulk: All Online", callback_data="adm_bulk_on")])
    buttons.append([InlineKeyboardButton(text="⏸ Bulk: All Offline", callback_data="adm_bulk_off")])
    try:
        await call.message.edit_reply_markup(reply_markup=InlineKeyboardMarkup(inline_keyboard=buttons))
    except Exception:
        pass


# ══════════════════════════════════════════════════════════════════
# ADMIN — FEAT-002: Profit Calculator
# ══════════════════════════════════════════════════════════════════

@router_admin.message(AdminFlow.menu, F.text == BTN_ADM_PROFIT)
@router_admin.message(AdminFlow.reports_submenu, F.text == BTN_ADM_PROFIT)
async def admin_profit_calc(message: Message, state: FSMContext):
    if not is_admin(message.from_user.id):
        return
    products = await get_all_products()
    orders, vpn, proxy = await asyncio.gather(
        db_get("orders"), db_get("vpn_orders"), db_get("proxy_orders"),
    )
    orders = orders or {}
    vpn = vpn or {}
    proxy = proxy or {}

    # Calculate per-product revenue
    product_revenue = defaultdict(float)
    product_sold = defaultdict(int)
    for o in orders.values():
        pid = o.get("product_id", "")
        product_revenue[pid] += o.get("total_price", 0)
        product_sold[pid] += o.get("qty", 0)
    for o in vpn.values():
        if o.get("status") == "delivered":
            pid = o.get("product_id", "")
            product_revenue[pid] += o.get("price", 0)
            product_sold[pid] += 1
    for o in proxy.values():
        if o.get("status") == "delivered":
            pid = o.get("product_id", "")
            product_revenue[pid] += o.get("price", 0)
            product_sold[pid] += 1

    total_revenue = sum(product_revenue.values())
    total_cost = 0.0
    lines = [f"\U0001F4B9 <b>Profit Calculator</b>\n{_SEP}\n"]
    lines.append("<b>Per-Product Breakdown:</b>\n")
    for pid, p in products.items():
        rev = product_revenue.get(pid, 0)
        cost_per = p.get("cost_price", 0)
        sold = product_sold.get(pid, 0)
        cost = cost_per * sold
        total_cost += cost
        profit = rev - cost
        margin = (profit / rev * 100) if rev > 0 else 0
        if rev > 0 or sold > 0:
            _pkg_emoji = p.get('emoji', '\U0001F4E6')
            lines.append(
                f"  {_pkg_emoji} <b>{p.get('name', pid)}</b>\n"
                f"    Revenue: ${rev:.2f} | Cost: ${cost:.2f} | Profit: ${profit:.2f} ({margin:.0f}%)"
            )
    total_profit = total_revenue - total_cost
    overall_margin = (total_profit / total_revenue * 100) if total_revenue > 0 else 0
    lines.append(f"\n{_SEP}")
    lines.append(f"\U0001F4B0 <b>Total Revenue:</b> ${total_revenue:.2f}")
    lines.append(f"\U0001F4B8 <b>Total Cost:</b> ${total_cost:.2f}")
    lines.append(f"\U0001F4B9 <b>Net Profit:</b> ${total_profit:.2f}")
    lines.append(f"\U0001F4CA <b>Overall Margin:</b> {overall_margin:.1f}%")
    lines.append(f"\n<i>Set cost per item via:</i> <code>set_cost:PRODUCT_ID:amount</code>")
    await state.set_state(AdminFlow.profit_calc)
    await message.answer("\n".join(lines), reply_markup=_kb([BACK_BTN, HOME_BTN]))


@router_admin.message(AdminFlow.profit_calc)
async def admin_profit_set_cost(message: Message, state: FSMContext):
    if message.text in (BACK_BTN, HOME_BTN):
        await state.set_state(AdminFlow.menu)
        await message.answer("🔐 <b>Admin Panel</b>", reply_markup=admin_main_kb())
        return
    text = (message.text or "").strip()
    if not text.lower().startswith("set_cost:"):
        await message.answer("❌ Use format: <code>set_cost:PRODUCT_ID:amount</code>")
        return
    parts = text.split(":", 2)
    if len(parts) < 3:
        await message.answer("❌ Format: <code>set_cost:PRD-0001:0.10</code>")
        return
    pid = parts[1].strip()
    try:
        cost = float(parts[2].strip())
    except ValueError:
        await message.answer("❌ Invalid cost amount.")
        return
    product = await get_product(pid)
    if not product:
        await message.answer(f"❌ Product <code>{pid}</code> not found.")
        return
    await update_product(pid, {"cost_price": cost})
    await message.answer(
        f"✅ Set cost for <b>{product.get('name', pid)}</b> to <b>${cost:.2f}</b> per item."
    )


# ══════════════════════════════════════════════════════════════════
# ADMIN — FEAT-002: Activity Logs
# ══════════════════════════════════════════════════════════════════

@router_admin.message(AdminFlow.menu, F.text == BTN_ADM_LOGS)
@router_admin.message(AdminFlow.reports_submenu, F.text == BTN_ADM_LOGS)
async def admin_activity_logs(message: Message, state: FSMContext):
    if not is_admin(message.from_user.id):
        return
    logs = await get_activity_logs(limit=15, offset=0)
    if not logs:
        await message.answer(
            f"📜 <b>Activity Logs</b>\n{_SEP}\n\n<i>No activity logs yet.</i>",
            reply_markup=admin_main_kb(),
        )
        return
    lines = [f"📜 <b>Activity Logs</b>\n{_SEP}\n"]
    for log in logs:
        ts = _dt(log.get("timestamp", 0))
        admin_id = log.get("admin_id", "?")
        action = log.get("action", "?")
        details = log.get("details", "")
        detail_str = f" - {details}" if details else ""
        lines.append(f"  <code>{ts}</code>\n  Admin: <code>{admin_id}</code> | <b>{action}</b>{detail_str}\n")
    buttons = []
    if len(logs) >= 15:
        buttons.append([InlineKeyboardButton(text="➡️ Next Page", callback_data="adm_logs:15")])
    kb = InlineKeyboardMarkup(inline_keyboard=buttons) if buttons else None
    await message.answer("\n".join(lines), reply_markup=kb or admin_main_kb())


@router_admin.callback_query(F.data.startswith("adm_logs:"))
async def admin_logs_page_cb(call: CallbackQuery):
    if not is_admin(call.from_user.id):
        await call.answer("🚫 Denied", show_alert=True)
        return
    offset = int(call.data.split(":", 1)[1])
    logs = await get_activity_logs(limit=15, offset=offset)
    if not logs:
        await call.answer("No more logs", show_alert=True)
        return
    lines = [f"📜 <b>Activity Logs</b> (page {offset // 15 + 2})\n{_SEP}\n"]
    for log in logs:
        ts = _dt(log.get("timestamp", 0))
        admin_id = log.get("admin_id", "?")
        action = log.get("action", "?")
        details = log.get("details", "")
        detail_str = f" - {details}" if details else ""
        lines.append(f"  <code>{ts}</code>\n  Admin: <code>{admin_id}</code> | <b>{action}</b>{detail_str}\n")
    buttons = []
    if len(logs) >= 15:
        buttons.append([InlineKeyboardButton(text="➡️ Next Page", callback_data=f"adm_logs:{offset + 15}")])
    if offset > 0:
        buttons.append([InlineKeyboardButton(text="⬅️ Previous", callback_data=f"adm_logs:{max(0, offset - 15)}")])
    kb = InlineKeyboardMarkup(inline_keyboard=buttons) if buttons else None
    await call.message.answer("\n".join(lines), reply_markup=kb)
    await call.answer()


# ══════════════════════════════════════════════════════════════════
# DISPATCHER
# ══════════════════════════════════════════════════════════════════

def build_dp() -> Dispatcher:
    dp = Dispatcher(storage=MemoryStorage())
    dp.message.middleware(AntiSpamMiddleware())
    dp.message.middleware(AuthMiddleware())
    dp.callback_query.middleware(AuthMiddleware())   # FIX: apply to callbacks too

    dp.include_router(router_global)   # HOME first — highest priority
    dp.include_router(router_admin)    # Admin before user so AdminFlow states take priority
    dp.include_router(router_user_panel)  # User panel enhancements (before router_start for callbacks)
    dp.include_router(router_start)
    dp.include_router(router_mail)
    dp.include_router(router_vpn)
    dp.include_router(router_proxy)
    dp.include_router(router_deposit)
    dp.include_router(router_tempmail)
    return dp


# ══════════════════════════════════════════════════════════════════
# BACKGROUND TASK — Periodic Stock Export
# ══════════════════════════════════════════════════════════════════

async def stock_export_loop(bot: Bot) -> None:
    """Periodically export all stock as xlsx and send to the designated group."""
    while True:
        try:
            settings = await get_settings()
            interval = float(settings.get("stock_export_interval", 30))
            if interval <= 0:
                # Export disabled; sleep and re-check later
                await asyncio.sleep(300)
                continue
            if interval < 1:
                interval = 1

            products = await get_all_products()
            if not products:
                await asyncio.sleep(interval * 60)
                continue

            all_stocks = {}
            for pid in products:
                all_stocks[pid] = await db_get(f"stocks/{pid}") or {}

            total_items = sum(len(v) for v in all_stocks.values())
            xlsx_bytes = make_products_xlsx(products, all_stocks)
            date_str = time.strftime("%Y%m%d_%H%M", time.gmtime())

            await bot.send_document(
                chat_id=STOCK_EXPORT_GROUP_ID,
                document=BufferedInputFile(xlsx_bytes, filename=f"all_stock_{date_str}.xlsx"),
                caption=(
                    f"\U0001f4e6 <b>Auto Stock Export</b>\n"
                    f"\U0001f4c2 Products: <b>{len(products)}</b>\n"
                    f"\U0001f4cb Total Items: <b>{total_items}</b>\n"
                    f"\U0001f552 Interval: <b>{interval:.0f} min</b>\n\n"
                    f"<i>Exported at {time.strftime('%Y-%m-%d %H:%M UTC', time.gmtime())}</i>\n"
                    f"<i>This file is importable via Import Products.</i>"
                ),
            )

            await asyncio.sleep(interval * 60)
        except Exception as e:
            logger.error("stock_export_loop error: %s", e)
            await asyncio.sleep(60)


# ══════════════════════════════════════════════════════════════════
# BACKGROUND TASK — Periodic Database Backup
# ══════════════════════════════════════════════════════════════════

async def backup_loop(bot: Bot) -> None:
    """Periodically export the entire Firebase database as JSON and send to all admins."""
    while True:
        try:
            settings = await get_settings()
            interval = float(settings.get("backup_interval_hours", 24))
            if interval <= 0:
                await asyncio.sleep(300)
                continue

            # Fetch all data from Firebase
            users = await db_get("users") or {}
            orders = await db_get("orders") or {}
            vpn_orders = await db_get("vpn_orders") or {}
            proxy_orders = await db_get("proxy_orders") or {}
            deposits = await db_get("deposits") or {}
            products = await db_get("products") or {}
            stocks = await db_get("stocks") or {}
            coupons = await db_get("coupons") or {}
            settings_data = await db_get("settings") or {}

            backup_data = {
                "users": users,
                "orders": orders,
                "vpn_orders": vpn_orders,
                "proxy_orders": proxy_orders,
                "deposits": deposits,
                "products": products,
                "stocks": stocks,
                "coupons": coupons,
                "settings": settings_data,
            }

            json_bytes = json.dumps(backup_data, indent=2, ensure_ascii=False).encode("utf-8")

            # Summary
            user_count = len(users)
            order_count = len(orders)
            total_revenue = sum(
                float(o.get("total_price", 0)) for o in orders.values() if isinstance(o, dict)
            )

            date_str = time.strftime("%Y%m%d_%H%M", time.gmtime())
            caption = (
                f"\U0001f4be <b>Auto Database Backup</b>\n"
                f"\U0001f465 Users: <b>{user_count}</b>\n"
                f"\U0001f4e6 Orders: <b>{order_count}</b>\n"
                f"\U0001f4b0 Total Revenue: <b>{total_revenue:.2f} BDT</b>\n"
                f"\U0001f552 Interval: <b>{interval:.0f}h</b>\n\n"
                f"<i>Backup at {time.strftime('%Y-%m-%d %H:%M UTC', time.gmtime())}</i>"
            )

            for adm in ADMIN_IDS:
                try:
                    await bot.send_document(
                        chat_id=adm,
                        document=BufferedInputFile(json_bytes, filename=f"backup_{date_str}.json"),
                        caption=caption,
                    )
                except Exception:
                    logger.error("backup_loop: failed to send backup to admin %s", adm)

            await asyncio.sleep(interval * 3600)
        except Exception as e:
            logger.error("backup_loop error: %s", e)
            await asyncio.sleep(60)


# ══════════════════════════════════════════════════════════════════
# STARTUP / SHUTDOWN
# ══════════════════════════════════════════════════════════════════

async def on_startup(bot: Bot) -> None:
    global _stock_export_task, _backup_task
    if WEBHOOK_URL:
        full = f"{WEBHOOK_URL.rstrip('/')}{WEBHOOK_PATH}"
        await bot.set_webhook(full)
        logger.info("Webhook → %s", full)
    else:
        await bot.delete_webhook()
        logger.info("Polling mode active")
    me = await bot.get_me()
    logger.info("@%s (id=%s) is online ✅", me.username, me.id)
    _stock_export_task = asyncio.create_task(stock_export_loop(bot))
    _backup_task = asyncio.create_task(backup_loop(bot))


async def on_shutdown(bot: Bot) -> None:
    global _stock_export_task, _backup_task
    if _stock_export_task is not None:
        _stock_export_task.cancel()
        _stock_export_task = None
    if _backup_task is not None:
        _backup_task.cancel()
        _backup_task = None
    if WEBHOOK_URL:
        await bot.delete_webhook()
    logger.info("Bot shut down.")


# ══════════════════════════════════════════════════════════════════
# ENTRY POINT
# ══════════════════════════════════════════════════════════════════

def main() -> None:
    bot = Bot(token=BOT_TOKEN, default=DefaultBotProperties(parse_mode=ParseMode.HTML))
    dp  = build_dp()

    if WEBHOOK_URL:
        dp.startup.register(on_startup)
        dp.shutdown.register(on_shutdown)
        app = web.Application()
        SimpleRequestHandler(dispatcher=dp, bot=bot).register(app, path=WEBHOOK_PATH)
        setup_application(app, dp, bot=bot)
        logger.info("Webhook server %s:%s%s", WEBAPP_HOST, WEBAPP_PORT, WEBHOOK_PATH)
        web.run_app(app, host=WEBAPP_HOST, port=WEBAPP_PORT)
    else:
        asyncio.run(_poll(dp, bot))


async def _poll(dp: Dispatcher, bot: Bot) -> None:
    await on_startup(bot)
    try:
        await dp.start_polling(bot, allowed_updates=dp.resolve_used_update_types())
    finally:
        await on_shutdown(bot)


if __name__ == "__main__":
    main()
